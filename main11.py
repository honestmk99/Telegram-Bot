import telegram.ext
import telegram.ext 
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, ConversationHandler,CallbackQueryHandler
from telegram.ext import (
    Updater,
    CommandHandler,
    MessageHandler,
    Filters,
    PreCheckoutQueryHandler,
    ShippingQueryHandler,
)
from telegram.error import TelegramError
import datetime
import sqlite3
import random
import requests
from PIL import Image  
import PIL  
import qrcode
from datetime import date
from datetime import datetime, timedelta
import datetime
import sqlite3 as sql
import os
import json
import string
from sqlite3 import Error
from telegram import KeyboardButton 
from telegram import (ReplyKeyboardMarkup, ReplyKeyboardRemove)
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
import re
from datetime import datetime
from datetime import date
import xlwt
from openpyxl import load_workbook
wc=["WELCOME TO SCAMILY VALUES‼️🥷🏽\nHere at SCAMILY VALUES you will find EVERYTHING you need to succeed‼️\n\n🤝Here you can automatically buy what you need🤫\n🤝If you have any problems - our support is online 24/7!\n🤝Good Luck‼️"]
BUTTON,AB,PHOTO,DES,PR,PIDA,PID,TYPP,MADD,DELETE,CH,BEL,ASUB,ASUBB,BELA,PDB,PDVM,SHAKA,ADDB,PROFT,WALL,TRADE,ADDRESS,DDFF,HAM,PENORD,FIKA,FIKAA,FIKAS,FIKAAS,FIKAF,FIKAAF,HALP,DEPO,DEPOT,DFAQ,DDEPO,PDVMF,SHAKAF,PDBF,ADDRESSF,TRADEF,PROFTF,JE,SH,MADDA,UOL,BYE,YUO,CUI,DBS,WWRR,CUIT,SH, DBSF,CUIF,CUITF,DDFFF,YUOF,PDVMS,SHAKAS,PDBS,ADDRESSS,YUOS,TRADES,DDFFS,BYES,PROFTS,CUIS,CUITS,CUITQS,DBSS,CUITQ,CUITQF,PSTOCK,PSTOCKA,JOP,JOPS,JOPF=range(79)
def start(update, context):
  userg = update.message.from_user
  try:
    usaf=userg.username
  except:
    usaf=userg.first_name
  user=update.effective_user.id
  user = update.message.from_user
  usa=str(update.effective_user.id)
  user=int(update.effective_user.id)
  if usa == "1926801217" or usa == "1394902938" or usa =="5791144147" :#5515451493
    connection = sqlite3.connect("help.db")  
    cursor = connection.cursor() 
    cursor.execute("SELECT id FROM COMPANY ") 
    jobs = cursor.fetchall()
    if len(jobs) ==0:
        cursor.execute("INSERT INTO COMPANY (contact,id) \
          VALUES ('{}','{}')".format("0",'123'))
        connection.commit()
        connection.close()  
        conn = sqlite3.connect('wel.db')
        conn.execute("INSERT INTO COMPANY (msg) \
                VALUES ('{}')".format('Welcome to admin Panel')) 
        conn.commit() 
    keyboard =[[InlineKeyboardButton("Products", callback_data="pak") ,InlineKeyboardButton("Wallet", callback_data="wallet")],  
                [InlineKeyboardButton("Communications", callback_data="comm")]]

    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to the Admin Dashboard" ,reply_markup=reply_markup)
    return BUTTON
  else: 
      connection = sqlite3.connect("users.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT id FROM COMPANY where id= {}".format(int(user))) 
      jobs = cursor.fetchall()
      if len(jobs) ==0:
          cursor.execute("INSERT INTO COMPANY (ID) \
            VALUES ({})".format(int(user)))
          connection.commit()
          connection.close()

          connection = sqlite3.connect("wallet.db")  
          cursor = connection.cursor() 
          cursor.execute("INSERT INTO COMPANY (ID,balance,link,code,amount,name,withdrawl,wemail,ct,sct,email,Trx,exp,naam,wname,opt,ttr,twd,curr) \
            VALUES ({}, '{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')".format(int(user),"0","0","0","0",usaf,"0","0","0","0","0","0","0",usaf,usaf,"0","0","0","0"))
          connection.commit()
          connection.close()
          conn = sqlite3.connect('oo.db')
          conn.execute("INSERT INTO COMPANY (ID,pid,pname,type,amount) \
                  VALUES ('{}', '{}','{}', '{}','{}')".format(str(update.effective_user.id),"0","0","0","0")) 
          conn.commit()
      keyboard =[[InlineKeyboardButton("🏴󠁧󠁢󠁥󠁮󠁧󠁿 English", callback_data="english")],[InlineKeyboardButton("🇫🇷 French", callback_data="french")],[InlineKeyboardButton("🇪🇸 Spanish", callback_data="spanish")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text='Select Your Language',reply_markup=reply_markup)
      return BUTTON
def button(update, context):
    h=update.effective_user.id
    query = update.callback_query
    a=query.data
    if a== '1':
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Send Name of Product',reply_markup=reply_markup)
        return AB
    elif a=='pak':
      keyboard =[[InlineKeyboardButton("Add Category", callback_data="90"),InlineKeyboardButton("Delete Category", callback_data="del")],
                  [InlineKeyboardButton("+Add Sub-Category", callback_data="90a"),InlineKeyboardButton("X Sub-Category", callback_data="dela")],
                    [InlineKeyboardButton("+Add Product", callback_data="1"),InlineKeyboardButton("XDelete Product", callback_data="2")],
                    [InlineKeyboardButton("Change Product's stock", callback_data="pst")]]

      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Products Section" ,reply_markup=reply_markup)
      return BUTTON
    elif a=='wallet':
      keyboard =[[InlineKeyboardButton("Add New Wallet Address", callback_data="189"),InlineKeyboardButton("Delete Wallet Address", callback_data="d189")],
                  [InlineKeyboardButton("Withdrawals", callback_data="owth"),InlineKeyboardButton("User Wallets", callback_data="uwal")],
                  [InlineKeyboardButton("Withdrawl History", callback_data="wh"),InlineKeyboardButton("View Admin Wallet", callback_data="admw")],
                  [InlineKeyboardButton("🛒 Orders", callback_data="99"),InlineKeyboardButton("Check Deposits", callback_data="cdp")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to admin Dashboard" ,reply_markup=reply_markup)
      return BUTTON
    elif a=='comm':
      keyboard =[[InlineKeyboardButton("Welcome Message", callback_data="912"),InlineKeyboardButton("❓ Set Help / Support", callback_data="913")],
                  [InlineKeyboardButton("Add FAQ", callback_data="faq"),InlineKeyboardButton("Delete FAQ", callback_data="dfaq")],
                  [InlineKeyboardButton("🔊 Announcement", callback_data="916")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to admin Dashboard" ,reply_markup=reply_markup)
      return BUTTON
    elif a=='pst':
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Send Product ID to change stock',reply_markup=reply_markup)
        return PSTOCK
    elif a=='2190':
        connection = sqlite3.connect("orders.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT oid FROM COMPANY where ID= '{}' ".format(str(update.effective_user.id)))
        jobs = cursor.fetchall()
        if len(jobs) !=0:
                    conn = sqlite3.connect('orders.db')
                    cursor = conn.execute("SELECT oid,status from COMPANY where ID= '{}' ".format(str(update.effective_user.id)))
                    conn.commit()
                    keyf=[]
                    for row in cursor:
                        if row[1]=="pending":
                            k="⏳"
                        elif row[1]=="Accept":
                            k="✅"
                        elif row[1]=="Delivered":
                            k="🚚"
                        elif row[1]=="Rejected":
                            k="❌"
                        c=[InlineKeyboardButton(k+row[0], callback_data=row[0])]
                        keyf.append(c)
                    reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
                    context.bot.send_message(chat_id=update.effective_user.id,text="Select Order to check detail",reply_markup=reply_markup) 
                    return DDFF
        else:
          context.bot.send_message(chat_id=update.effective_user.id,text="You Have Not Made Any Orders")
          return BUTTON
    if a== 'english':
        keyf=[]
        connection = sqlite3.connect("cata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT cat FROM COMPANY")
        keyf=[]
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[0])]
            keyf.append(c)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Select Main Category",reply_markup=reply_markup) 
        keyf=[["📝️ Products","🛒 Cart"],["💳 Wallet","🛍 My orders"],["🙋 FAQ","❓ Help/Support"]]
        reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text='Main Menu',reply_markup=reply_markup)
        return PDB
    elif a== 'spanish':
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Productos","🛒 Carro"],["💳 Cartera","🛍 Mis ordenes"],[" 🙋 FAQ ","❓ Servicio de asistencia"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    if a== 'french':
      keyf=[["📝️ Des produits","🛒 Chariot"],["💳 Porte monnaie","🛍 Mes commandes"],["🙋 FAQs","❓ Support d'aide"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Bienvenue sur le tableau de bord utilisateur",reply_markup=reply_markup)

    elif a== '189':
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Send Name of Wallet',reply_markup=reply_markup)
        return DEPO
    elif a== 'd189':
        connection = sqlite3.connect("deposit.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT address,name FROM COMPANY")
        keyf=[]
        for name in cursor:
            c=[InlineKeyboardButton(name[1], callback_data=name[0])]
            keyf.append(c)
        b=[InlineKeyboardButton("🌎Main Menu", callback_data='🌎Main Menu')]
        keyf.append(b)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Select Address to Delete",reply_markup=reply_markup) 
        return DDEPO
    elif a== 'faq':
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Send FAQs Question',reply_markup=reply_markup)
        return FIKA
    elif a== 'dfaq':
        connection = sqlite3.connect("faqs.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT ques FROM COMPANY")
        keyf=[]
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[0])]
            keyf.append(c)
        b=[InlineKeyboardButton("🌎Main Menu", callback_data='🌎Main Menu')]
        keyf.append(b)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Select a FAQ to Delete",reply_markup=reply_markup) 
        return DFAQ
    elif a== 'cdp':
      connection = sqlite3.connect("dep.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT ID,amount,status FROM COMPANY where status='pending'")
      jobs = cursor.fetchall()
      if len(jobs) !=0:
        connection = sqlite3.connect("dep.db")  
        cursor = connection.cursor()
        cursor.execute("SELECT ID,amount,status FROM COMPANY where status='pending'")
        keyf=[]
        for row in cursor:
          m='ID: {}\nStatus: {}'.format(row[0],row[2])
          keyboard =[[InlineKeyboardButton("Accept", callback_data="1abb"),InlineKeyboardButton("Reject", callback_data="1acc")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.send_photo(chat_id=update.effective_user.id,photo=row[1],caption=m,reply_markup=reply_markup)
          return ADDB

      else:
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='No Pending Request',reply_markup=reply_markup)
        return BUTTON      
    elif a== '913':
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Send Help/Support Info',reply_markup=reply_markup)
        return HALP
    elif a=='admw':
        style = xlwt.easyxf('font: bold 1, color black;')
        conn = sqlite3.connect('deposit.db') 
        cursor = conn.execute("SELECT name from COMPANY ")
        conn.commit() 
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('deposit.db') 
            cursor = conn.execute("SELECT name,address from COMPANY ")
            conn.commit()
            xa="Address\n"
            i=0
            j=0
            workbook = xlwt.Workbook()
            sheet = workbook.add_sheet("Address History")
            sheet.write(i, 1, "name",  style)
            sheet.write(i, 2, "address",  style)
            for row in cursor:
                    i=i+1
                    j=j+1
                    inv=row[0]
                    vgy=row[1]
                    xdrt=row[2]
                    sheet.write(i, 1, row[0], style)
                    sheet.write(i, 2, row[1], style)
            workbook.save("Address.xls")
            keyboard =[[InlineKeyboardButton("Main Menu", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
            context.bot.send_document(chat_id=update.effective_user.id,document=open('Address.xls', 'rb'),reply_markup=reply_markup)
        else:
                keyboard =[[InlineKeyboardButton("❌", callback_data="100")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
                context.bot.send_message(chat_id=update.effective_user.id,text="You have no Address history",reply_markup=reply_markup)
                return BUTTON
    elif a=='wh':
        style = xlwt.easyxf('font: bold 1, color black;')
        conn = sqlite3.connect('withdraw.db') 
        cursor = conn.execute("SELECT ID from COMPANY ")
        conn.commit() 
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('withdraw.db') 
            cursor = conn.execute("SELECT ID,balance,email,wid,rid,date,time,status from COMPANY ")
            conn.commit()
            xa="withdrawl\n"
            i=0
            j=0
            workbook = xlwt.Workbook()
            sheet = workbook.add_sheet("Order History")
            sheet.write(i, 1, "Wallet_ID",  style)
            sheet.write(i, 2, "Balance",  style)
            sheet.write(i, 3, "Withdrawl_amount", style)
            sheet.write(i, 4, "Email", style)
            sheet.write(i, 5, "Time", style)
            sheet.write(i, 6, "Date", style)
            sheet.write(i, 7, "Status", style)
            for row in cursor:
                    i=i+1
                    j=j+1
                    inv=row[0]
                    vgy=row[1]
                    xdrt=row[2]
                    sheet.write(i, 1, row[0], style)
                    sheet.write(i, 2, row[1], style)
                    sheet.write(i, 3, row[3],  style)
                    sheet.write(i, 4, row[2],  style)
                    sheet.write(i, 5, row[6],  style)
                    sheet.write(i, 6, row[5],  style)
                    sheet.write(i, 7, row[7],  style)
            workbook.save("withdrawl.xls")
            keyboard =[[InlineKeyboardButton("Main Menu", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
            context.bot.send_document(chat_id=update.effective_user.id,document=open('withdrawl.xls', 'rb'),reply_markup=reply_markup)
        else:
                keyboard =[[InlineKeyboardButton("❌", callback_data="100")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
                context.bot.send_message(chat_id=update.effective_user.id,text="You have no withdrawl history",reply_markup=reply_markup)
                return BUTTON
    elif a=='uwal':
        style = xlwt.easyxf('font: bold 1, color black;')
        conn = sqlite3.connect('wallet.db') 
        cursor = conn.execute("SELECT ID from COMPANY ")
        conn.commit() 
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            f = open("wallet.txt", "w",encoding="utf-8")
            conn = sqlite3.connect('wallet.db') 
            cursor = conn.execute("SELECT ID,name,wname,balance,ttr,twd from COMPANY ")
            conn.commit()
            xa="wallet\n"
            i=0
            j=0
            workbook = xlwt.Workbook()
            sheet = workbook.add_sheet("Order History")
            sheet.write(i, 1, "Wallet_ID",  style)
            sheet.write(i, 2, "User_Name", style)
            sheet.write(i, 3, "Wallet_Name",  style)
            sheet.write(i, 4, "Balance", style)
            sheet.write(i, 5, "Total_added_balance", style)
            sheet.write(i, 6, "Total_withdrawl", style)
            for row in cursor:
                    i=i+1
                    j=j+1
                    inv=row[0]
                    vgy=row[1]
                    xdrt=row[2]
                    sheet.write(i, 1, row[0], style)
                    sheet.write(i, 2, row[1], style)
                    sheet.write(i, 3, row[2],  style)
                    sheet.write(i, 4, row[3],  style)
                    sheet.write(i, 5, row[4],  style)
                    sheet.write(i, 6, row[5],  style)

            workbook.save("wallet.xls")
            keyboard =[[InlineKeyboardButton("Main Menu", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
            context.bot.send_document(chat_id=update.effective_user.id,document=open('wallet.xls', 'rb'),reply_markup=reply_markup)
        else:
                keyboard =[[InlineKeyboardButton("❌", callback_data="100")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
                context.bot.send_message(chat_id=update.effective_user.id,text="You have no orders",reply_markup=reply_markup)
                return BUTTON
    elif a=='owth':
      conn = sqlite3.connect('withdraw.db') 
      cursor = conn.execute("SELECT ID from COMPANY where status='Pending'")
      conn.commit() 
      jobs = cursor.fetchall()
      if len(jobs) !=0:
        conn = sqlite3.connect('withdraw.db')
        cursor = conn.execute("SELECT ID,balance,status,wid,email,rid from COMPANY where status='Pending'")
        conn.commit()
        keyf=[]
        k=''
        for row in cursor:
            if row[2]=="Pending":
                k="⏳,"

            c=[InlineKeyboardButton(k+str(row[0]), callback_data=row[5])]
            keyf.append(c)
        b=[InlineKeyboardButton("🌎Main Menu", callback_data='🌎Main Menu')]
        keyf.append(b)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Select Withdrawal Request",reply_markup=reply_markup) 
        return WWRR
      else:
        keyboard = [[InlineKeyboardButton("🔙 Main Menu", callback_data='MAINN')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.bot.send_message(chat_id=update.effective_user.id,text="You don't have any request",reply_markup=reply_markup)                       
        return WWRR
    elif a== '90p':
        context.bot.send_message(chat_id=update.effective_user.id,text='Send a New Name for your Wallet')
        return UOL
    elif a== '1with':
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT balance from COMPANY where ID= {}   ".format(int(update.effective_user.id))) 
        conn.commit()
        for row in cursor:  
            qoop=float(row[0]) 
        if qoop==0:
          context.bot.send_message(chat_id=update.effective_user.id,text="Sorry You don't have enough balance to withdraw")
        else:
          context.bot.send_message(chat_id=update.effective_user.id,text="Each Credit in your Wallet is Worth $1 USD. Please Send us the Amount you Wish to Withdraw")   
          return DBS 
    elif a== '1withs':
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT balance from COMPANY where ID= {}   ".format(int(update.effective_user.id))) 
        conn.commit()
        for row in cursor:  
            qoop=float(row[0]) 
        if qoop==0:
          context.bot.send_message(chat_id=update.effective_user.id,text="Lo sentimos, no tienes suficiente saldo para retirar")
        else:
          context.bot.send_message(chat_id=update.effective_user.id,text="Cada Crédito en su Monedero vale $1 USD. Envíenos la cantidad que desea retirar")   
          return DBSS
    elif a== '1withf':
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT balance from COMPANY where ID= {}   ".format(int(update.effective_user.id))) 
        conn.commit()
        for row in cursor:  
            qoop=float(row[0]) 
        if qoop==0:

          keyboard =[[InlineKeyboardButton("🔙 Menu principal", callback_data="french")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.send_message(chat_id=update.effective_user.id,text="Désolé, vous n'avez pas assez de solde pour effectuer un retrait",reply_markup=reply_markup)
        else:
          keyboard =[[InlineKeyboardButton("🔙 Menu principal", callback_data="french")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.send_message(chat_id=update.effective_user.id,text="Smontant final que vous souhaitez retirer",reply_markup=reply_markup)   
          return DBSF       
    elif a=='2':
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text='Send Product ID to Delete',reply_markup=reply_markup)
      return DELETE
    elif a=="90":      
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Send Name of Category to Add',reply_markup=reply_markup)       
        return CH
    elif a=="del":      
        keyf=[]
        connection = sqlite3.connect("cata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT cat FROM COMPANY")
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[0])]
            keyf.append(c)
        b=[InlineKeyboardButton("🌎Main Menu", callback_data='🌎Main Menu')]
        keyf.append(b)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Select a Category to Delete",reply_markup=reply_markup) 
        return BEL
    elif a=="dela":      
        keyf=[]
        connection = sqlite3.connect("subcata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT subcat FROM COMPANY")
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[0])]
            keyf.append(c)
        b=[InlineKeyboardButton("🌎Main Menu", callback_data='🌎Main Menu')]
        keyf.append(b)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Select Sub-Category to Delete",reply_markup=reply_markup) 
        return BELA 
    elif a=="90a":      
        connection = sqlite3.connect("cata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT cat FROM COMPANY")
        keyf=[]
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[0])]
            keyf.append(c)
        b=[InlineKeyboardButton("🌎Main Menu", callback_data='🌎Main Menu')]
        keyf.append(b)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Select Main Category",reply_markup=reply_markup)   
        return ASUB
    elif a== '160': #quantity++
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        dv=update.callback_query.message.photo[-1].file_id
        cf=update.callback_query.message.message_id
        d=cc
        df=cc.split("𝐐𝐮𝐚𝐧𝐭𝐢𝐭𝐲:  ")
        df=df[1]
        df=df.strip()
        xdd=df
        dff=float(df)
        dffr=float(df)
        df=dff
        dff=dff+1
        mf=cc.split("𝐏𝐫𝐢𝐜𝐞:   ")
        mf=mf[1]
        mf=mf.split("𝐃𝐞𝐬𝐜𝐫𝐢𝐩𝐭𝐢𝐨𝐧:  ")
        mf=mf[0]
        mf=mf.strip()

        mqq=cc.split("𝐈𝐃:  ")
        mqq=mqq[1]
        mqq=mqq.split("𝐏𝐫𝐢𝐜𝐞:   ")
        mqq=mqq[0]
        mqq=mqq.strip()
        mqqt=cc.split("𝐒𝐭𝐨𝐜𝐤 𝐚𝐯𝐚𝐢𝐥𝐢𝐛𝐢𝐥𝐭𝐲:  ")
        mqqt=mqqt[1]
        mqqt=mqqt.split("𝐐𝐮𝐚𝐧𝐭𝐢𝐭𝐲:")
        mqqt=mqqt[0]
        mqqt=mqqt.strip()
        connection = sqlite3.connect("shop.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT max FROM COMPANY where productID='{}'".format(str(mqq)))
        keyf=[]
        for name in cursor:
          tqp=float(name[0])
        if dffr==tqp:
          keyboard =[[InlineKeyboardButton("➕ Quantity", callback_data="160"),InlineKeyboardButton("➖ Quantity", callback_data="161")],[InlineKeyboardButton("➕ Add to cart", callback_data="16"),InlineKeyboardButton("Back to Products", callback_data="200")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)
          context.bot.send_message(chat_id=update.effective_user.id,text='You Have Reached the Maximum Amount that you can Purchase')  
        elif dffr==float(mqqt):
         
          keyboard =[[InlineKeyboardButton("➕ Quantity", callback_data="160"),InlineKeyboardButton("➖ Quantity", callback_data="161")],[InlineKeyboardButton("➕ Add to cart", callback_data="16"),InlineKeyboardButton("Back to Products", callback_data="200")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)
          context.bot.send_message(chat_id=update.effective_user.id,text=' You Have Reached the Maximum Amount that you can Purchase')
        else:

          fgh=mf
          mf=float(mf)
          vc=float(df)
          op=mf/vc
          np=op*dff
          np=str(np)
          dy=str(dff)
          hjam="𝐐𝐮𝐚𝐧𝐭𝐢𝐭𝐲:  {}".format(xdd)
          hjamn="𝐐𝐮𝐚𝐧𝐭𝐢𝐭𝐲:  {}".format(dy)
          cc=cc.replace(hjam,hjamn)
          cc=cc.replace(fgh,np+"$")
          keyboard =[[InlineKeyboardButton("➕ Quantity", callback_data="160"),InlineKeyboardButton("➖ Quantity", callback_data="161")],[InlineKeyboardButton("➕ Add to cart", callback_data="16"),InlineKeyboardButton("Back to Products", callback_data="200")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)
    elif a== '160s': #quantity++
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        dv=update.callback_query.message.photo[-1].file_id
        cf=update.callback_query.message.message_id
        d=cc
        df=cc.split("𝐂𝐚𝐧𝐭𝐢𝐝𝐚𝐝:  ")
        df=df[1]
        df=df.strip()
        xdd=df
        dff=float(df)
        dffr=float(df)
        df=dff
        dff=dff+1
        mf=cc.split("𝐏𝐫𝐞𝐜𝐢𝐨:   ")
        mf=mf[1]
        mf=mf.split("𝐃𝐞𝐬𝐜𝐫𝐢𝐩𝐜𝐢ó𝐧:  ")
        mf=mf[0]
        mf=mf.strip()

        mqq=cc.split("𝐈𝐃:  ")
        mqq=mqq[1]
        mqq=mqq.split("𝐏𝐫𝐞𝐜𝐢𝐨:   ")
        mqq=mqq[0]
        mqq=mqq.strip()
        mqqt=cc.split("𝐃𝐢𝐬𝐩𝐨𝐧𝐢𝐛𝐢𝐥𝐢𝐝𝐚𝐝 𝐝𝐞 𝐞𝐱𝐢𝐬𝐭𝐞𝐧𝐜𝐢𝐚𝐬: ")
        mqqt=mqqt[1]
        mqqt=mqqt.split("𝐂𝐚𝐧𝐭𝐢𝐝𝐚𝐝:  ")
        mqqt=mqqt[0]
        mqqt=mqqt.strip()
        connection = sqlite3.connect("shop.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT max FROM COMPANY where productID='{}'".format(str(mqq)))
        keyf=[]
        for name in cursor:
          tqp=float(name[0])
        if dffr==tqp:
          keyboard =[[InlineKeyboardButton("➕ Añadir a la cesta", callback_data="16s"),InlineKeyboardButton("🔙 Cancelar", callback_data="spanish")],[InlineKeyboardButton("➕ Cantidad", callback_data="161s")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)
          context.bot.send_message(chat_id=update.effective_user.id,text=' Has alcanzado la cantidad máxima permitida')  
        elif dffr==float(mqqt):
         
          keyboard =[[InlineKeyboardButton("➕ Añadir a la cesta", callback_data="16s"),InlineKeyboardButton("🔙 Cancelar", callback_data="spanish")],[InlineKeyboardButton("➖ Cantidad", callback_data="161s")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)
          context.bot.send_message(chat_id=update.effective_user.id,text=' Has alcanzado la cantidad máxima permitida')
        else:

          fgh=mf
          mf=float(mf)
          vc=float(df)
          op=mf/vc
          np=op*dff
          np=str(np)
          dy=str(dff)
          hjam="𝐂𝐚𝐧𝐭𝐢𝐝𝐚𝐝: {}".format(xdd)
          hjamn="𝐂𝐚𝐧𝐭𝐢𝐝𝐚𝐝: {}".format(dy)
          cc=cc.replace(hjam,hjamn)
          cc=cc.replace(fgh,np+"$")
          keyboard =[[InlineKeyboardButton("➕ Añadir a la cesta", callback_data="16s"),InlineKeyboardButton("🔙 Cancelar", callback_data="spanish")],[InlineKeyboardButton("➕ Cantidad", callback_data="160s"),InlineKeyboardButton("➖ Cantidad", callback_data="161s")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)   
    elif a== '160f': #quantity++
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        dv=update.callback_query.message.photo[-1].file_id
        cf=update.callback_query.message.message_id
        d=cc
        df=cc.split("𝐐𝐮𝐚𝐧𝐭𝐢𝐭é:  ")
        df=df[1]
        df=df.strip()
        xdd=df
        dff=float(df)
        dffr=float(df)
        df=dff
        dff=dff+1
        mf=cc.split("𝐏𝐫𝐢𝐱:   ")
        mf=mf[1]
        mf=mf.split("𝐋𝐚 𝐝𝐞𝐬𝐜𝐫𝐢𝐩𝐭𝐢𝐨𝐧:  ")
        mf=mf[0]
        mf=mf.strip()
        mqqt=cc.split("𝐃𝐢𝐬𝐩𝐨𝐧𝐢𝐛𝐢𝐥𝐢𝐭é 𝐝𝐞𝐬 𝐬𝐭𝐨𝐜𝐤𝐬: ")
        mqqt=mqqt[1]
        mqqt=mqqt.split("𝐐𝐮𝐚𝐧𝐭𝐢𝐭é:  ")
        mqqt=mqqt[0]
        mqqt=mqqt.strip()

        mqq=cc.split("𝐈𝐃:  ")
        mqq=mqq[1]
        mqq=mqq.split("𝐏𝐫𝐢𝐱:   ")
        mqq=mqq[0]
        mqq=mqq.strip()
        connection = sqlite3.connect("shop.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT max FROM COMPANY where productID='{}'".format(str(mqq)))
        keyf=[]
        for name in cursor:
          tqp=float(name[0])
        connection = sqlite3.connect("shop.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT max FROM COMPANY where productID='{}'".format(str(mqq)))
        keyf=[]
        for name in cursor:
          tqp=float(name[0])
        if dffr==tqp:
          keyboard =[[InlineKeyboardButton("➕ Ajouter au panier", callback_data="16f"),InlineKeyboardButton("🔙Menu principal", callback_data="french")],[InlineKeyboardButton("➖ Quantité", callback_data="161f")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)
          context.bot.send_message(chat_id=update.effective_user.id,text=' Vous avez atteint la quantité maximale autorisée')  
        elif dffr==float(mqqt):
         
          keyboard =[[InlineKeyboardButton("➕ Ajouter au panier", callback_data="16f"),InlineKeyboardButton("🔙Menu principal", callback_data="french")],[InlineKeyboardButton("➖ Quantité", callback_data="161f")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)
          context.bot.send_message(chat_id=update.effective_user.id,text=' Vous avez atteint la quantité maximale autorisée')
        else:

          fgh=mf
          mf=float(mf)
          vc=float(df)
          op=mf/vc
          np=op*dff
          np=str(np)
          dy=str(dff)
          hjam="𝐐𝐮𝐚𝐧𝐭𝐢𝐭é:  {}".format(xdd)
          hjamn="𝐐𝐮𝐚𝐧𝐭𝐢𝐭é:  {}".format(dy)
          cc=cc.replace(hjam,hjamn)
          cc=cc.replace(fgh,np+"$")
          keyboard =[[InlineKeyboardButton("➕ Ajouter au panier", callback_data="16f"),InlineKeyboardButton("🔙 Menu principal", callback_data="french")],[InlineKeyboardButton("➕ Quantité", callback_data="160f"),InlineKeyboardButton("➖ Quantité", callback_data="161f")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)
    elif a== '161f': #quantity--
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        dv=update.callback_query.message.photo[-1].file_id
        cf=update.callback_query.message.message_id
        d=cc
        df=cc.split("𝐐𝐮𝐚𝐧𝐭𝐢𝐭é:  ")
        df=df[1]
        df=df.strip()
        xd=df
        df=float(df)
        if df>1:
          d=cc
          df=cc.split("𝐐𝐮𝐚𝐧𝐭𝐢𝐭é:  ")
          df=df[1]
          df=df.strip()
          xdd=df
          dff=float(df)
          df=dff
          dff=dff-1
          mf=cc.split("𝐏𝐫𝐢𝐱:   ")
          mf=mf[1]
          mf=mf.split("𝐋𝐚 𝐝𝐞𝐬𝐜𝐫𝐢𝐩𝐭𝐢𝐨𝐧:  ")
          mf=mf[0]
          mf=mf.strip()
          fgh=mf
          mf=float(mf)
          vc=float(df)
          op=mf/vc
          np=op*dff
          np=str(np)
          dy=str(dff)
          hjam="𝐐𝐮𝐚𝐧𝐭𝐢𝐭é:  {}".format(xdd)
          hjamn="𝐐𝐮𝐚𝐧𝐭𝐢𝐭é:  {}".format(dy)
          cc=cc.replace(hjam,hjamn)
          cc=cc.replace(fgh,np+"$")
          keyboard =[[InlineKeyboardButton("➕ Ajouter au panier", callback_data="16f"),InlineKeyboardButton("🔙Menu principal", callback_data="french")],[InlineKeyboardButton("➕ Quantité", callback_data="160f"),InlineKeyboardButton("➖ Quantité", callback_data="161f")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)     
    elif a== '161': #quantity--
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        dv=update.callback_query.message.photo[-1].file_id
        cf=update.callback_query.message.message_id
        d=cc
        df=cc.split("𝐐𝐮𝐚𝐧𝐭𝐢𝐭𝐲:  ")
        df=df[1]
        df=df.strip()
        xd=df
        df=float(df)
        if df>1:
          d=cc
          df=cc.split("𝐐𝐮𝐚𝐧𝐭𝐢𝐭𝐲:  ")
          df=df[1]
          df=df.strip()
          xdd=df
          dff=float(df)
          df=dff
          dff=dff-1
          mf=cc.split("𝐏𝐫𝐢𝐜𝐞:   ")
          mf=mf[1]
          mf=mf.split("𝐃𝐞𝐬𝐜𝐫𝐢𝐩𝐭𝐢𝐨𝐧:  ")
          mf=mf[0]
          mf=mf.strip()
          fgh=mf
          mf=float(mf)
          vc=float(df)
          op=mf/vc
          np=op*dff
          np=str(np)
          dy=str(dff)
          hjam="𝐐𝐮𝐚𝐧𝐭𝐢𝐭𝐲:  {}".format(xdd)
          hjamn="𝐐𝐮𝐚𝐧𝐭𝐢𝐭𝐲:  {}".format(dy)
          cc=cc.replace(hjam,hjamn)
          cc=cc.replace(fgh,np+"$")
          keyboard =[[InlineKeyboardButton("➕ Quantity", callback_data="160"),InlineKeyboardButton("➖ Quantity", callback_data="161")],[InlineKeyboardButton("➕ Add to cart", callback_data="16"),InlineKeyboardButton("Back to Products", callback_data="200")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup) 
    elif a== '161s': #quantity--
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        dv=update.callback_query.message.photo[-1].file_id
        cf=update.callback_query.message.message_id
        d=cc
        df=cc.split("𝐂𝐚𝐧𝐭𝐢𝐝𝐚𝐝: ")
        df=df[1]
        df=df.strip()
        xd=df
        df=float(df)
        if df>1:
          d=cc
          df=cc.split("𝐂𝐚𝐧𝐭𝐢𝐝𝐚𝐝: ")
          df=df[1]
          df=df.strip()
          xdd=df
          dff=float(df)
          df=dff
          dff=dff-1
          mf=cc.split("𝐏𝐫𝐞𝐜𝐢𝐨:   ")
          mf=mf[1]
          mf=mf.split("𝐃𝐞𝐬𝐜𝐫𝐢𝐩𝐜𝐢ó𝐧:  ")
          mf=mf[0]
          mf=mf.strip()
          fgh=mf
          mf=float(mf)
          vc=float(df)
          op=mf/vc
          np=op*dff
          np=str(np)
          dy=str(dff)
          hjam="𝐂𝐚𝐧𝐭𝐢𝐝𝐚𝐝: {}".format(xdd)
          hjamn="𝐂𝐚𝐧𝐭𝐢𝐝𝐚𝐝: {}".format(dy)
          cc=cc.replace(hjam,hjamn)
          cc=cc.replace(fgh,np+"$")
          keyboard =[[InlineKeyboardButton("➕ Añadir a la cesta", callback_data="16s"),InlineKeyboardButton("🔙 Cancelar", callback_data="spanish")],[InlineKeyboardButton("➕ Cantidad", callback_data="160s"),InlineKeyboardButton("➖ Cantidad", callback_data="161s")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)     
    elif a== '16':
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        uo=cc
        dv=update.callback_query.message.photo[-1].file_id
        fg=cc
        cf=update.callback_query.message.message_id
        df=cc.split("𝐏𝐫𝐢𝐜𝐞:   ")
        df=df[1]
        df=df.split("𝐃𝐞𝐬𝐜𝐫𝐢𝐩𝐭𝐢𝐨𝐧:  ")
        df=df[0]
        df=df.strip()
        dd=cc.split("𝐍𝐚𝐦𝐞:  ")
        dd=dd[1]
        dd=dd.split("𝐈𝐃:  ")
        dd=dd[0]
        dd=dd.strip()
        mm=cc.split("𝐐𝐮𝐚𝐧𝐭𝐢𝐭𝐲:  ")
        mm=mm[1]
        mm=mm.strip()
        cc=cc.split("𝐈𝐃:  ")
        cc=cc[1]
        cc=cc.split("𝐏𝐫𝐢𝐜𝐞:   ")
        cc=cc[0]
        cc=cc.strip()             
        print(mm)
        pric=df
        print(pric)
        nam=dd
        pid=cc
        conn = sqlite3.connect('cart.db') 
        conn.execute("INSERT INTO COMPANY (ID,name,productID,price,quantity) \
        VALUES ('{}', '{}','{}', '{}', '{}')".format(str(update.effective_user.id),nam,pid,pric,mm))
        conn.commit()
        connection = sqlite3.connect("shop.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT stock FROM COMPANY where productID= '{}'".format(pid) )
        for names in cursor:
          inv=float(names[0])
        bn=inv-float(mm)
        bn=str(bn)
        conn = sqlite3.connect("shop.db")  
        conn.execute("UPDATE COMPANY set stock = '{}' where productID = '{}'".format(bn,pid))
        conn.commit()
        conn.close()
        keyboard =[[InlineKeyboardButton("➖ Remove", callback_data="255"),InlineKeyboardButton("🛒 Cart", callback_data="22")],[InlineKeyboardButton("Continue Shopping", callback_data="200")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=fg,reply_markup=reply_markup)
        conn.close()
        return BUTTON
    elif a== '16s':
        print("SSSSSSSSSS")
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        uo=cc
        dv=update.callback_query.message.photo[-1].file_id
        fg=cc
        cf=update.callback_query.message.message_id
        df=cc.split("𝐏𝐫𝐞𝐜𝐢𝐨:   ")
        df=df[1]
        df=df.split("𝐃𝐞𝐬𝐜𝐫𝐢𝐩𝐜𝐢ó𝐧:  ")
        df=df[0]
        df=df.strip()
        dd=cc.split("𝐍𝐨𝐦𝐛𝐫𝐞: ")
        dd=dd[1]
        dd=dd.split("𝐈𝐃:  ")
        dd=dd[0]
        dd=dd.strip()
        mm=cc.split("𝐂𝐚𝐧𝐭𝐢𝐝𝐚𝐝: ")
        mm=mm[1]
        mm=mm.strip()
        cc=cc.split("𝐈𝐃:  ")
        cc=cc[1]
        cc=cc.split("𝐏𝐫𝐞𝐜𝐢𝐨:   ")
        cc=cc[0]
        cc=cc.strip()             
        print(mm)
        pric=df
        print(pric)
        nam=dd
        pid=cc
        print(nam)
        conn = sqlite3.connect('cart.db') 
        conn.execute("INSERT INTO COMPANY (ID,name,productID,price,quantity) \
        VALUES ('{}', '{}','{}', '{}', '{}')".format(str(update.effective_user.id),nam,pid,pric,mm))
        conn.commit()
        connection = sqlite3.connect("shop.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT stock FROM COMPANY where productID= '{}'".format(pid) )
        for names in cursor:
          inv=float(names[0])
        bn=inv-float(mm)
        bn=str(bn)
        conn = sqlite3.connect("shop.db")  
        conn.execute("UPDATE COMPANY set stock = '{}' where productID = '{}'".format(bn,pid))
        conn.commit()
        conn.close()
        keyboard =[[InlineKeyboardButton("➖ Remover", callback_data="255s"),InlineKeyboardButton("🛒 Carro", callback_data="22s")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=fg,reply_markup=reply_markup)
        conn.close()
        return BUTTON

    elif a== '16f':
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        uo=cc
        dv=update.callback_query.message.photo[-1].file_id
        fg=cc
        cf=update.callback_query.message.message_id
        df=cc.split("𝐏𝐫𝐢𝐱:   ")
        df=df[1]
        df=df.split("𝐋𝐚 𝐝𝐞𝐬𝐜𝐫𝐢𝐩𝐭𝐢𝐨𝐧:  ")
        df=df[0]
        df=df.strip()
        dd=cc.split("𝐍𝐨𝐦:  ")
        dd=dd[1]
        dd=dd.split("𝐈𝐃:  ")
        dd=dd[0]
        dd=dd.strip()
        mm=cc.split("𝐐𝐮𝐚𝐧𝐭𝐢𝐭é:  ")
        mm=mm[1]
        mm=mm.strip()
        cc=cc.split("𝐈𝐃:  ")
        cc=cc[1]
        cc=cc.split("𝐏𝐫𝐢𝐱:   ")
        cc=cc[0]
        cc=cc.strip()             
        print(mm)
        pric=df
        print(pric)
        nam=dd
        pid=cc
        conn = sqlite3.connect('cart.db') 
        conn.execute("INSERT INTO COMPANY (ID,name,productID,price,quantity) \
        VALUES ('{}', '{}','{}', '{}', '{}')".format(str(update.effective_user.id),nam,pid,pric,mm))
        conn.commit()
        connection = sqlite3.connect("shop.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT stock FROM COMPANY where productID= '{}'".format(pid) )
        for names in cursor:
          inv=float(names[0])
        bn=inv-float(mm)
        bn=str(bn)
        conn = sqlite3.connect("shop.db")  
        conn.execute("UPDATE COMPANY set stock = '{}' where productID = '{}'".format(bn,pid))
        conn.commit()
        conn.close()
        keyboard =[[InlineKeyboardButton("➖ Retirer", callback_data="255f"),InlineKeyboardButton("🛒 Chariot", callback_data="22f")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=fg,reply_markup=reply_markup)
        conn.close()
        return BUTTON
    elif a== '22':     
      query.answer()
      conn = sqlite3.connect('cart.db')
      cursor = conn.execute("SELECT name,Price,quantity from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
      conn.commit()
      global xru
      xru="𝐘𝐨𝐮𝐫 𝐂𝐚𝐫𝐭:\n\n"
      for row in cursor: 

 
        m="   "+row[0]+" x "+row[2] + "   =   "+row[1]+"\n"
      
        xru=xru+m
      cursor = conn.execute("SELECT SUM(Price),SUM(quantity) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
      global ap
      ap=""
      global nj
      nj=""
      for row in cursor:
          ap=str(row[0])
      keyboard =[[InlineKeyboardButton("✅ Checkout", callback_data="25"),InlineKeyboardButton("❌ Cart", callback_data="29")],[InlineKeyboardButton("Return to Shop", callback_data="200")]]   
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text=xru+"\nTotal"+" "+ap+"$" ,reply_markup=reply_markup)    
        
      return BUTTON
    elif a== '22s':     
        query.answer()
        conn = sqlite3.connect('cart.db')
        cursor = conn.execute("SELECT name,Price from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
        conn.commit()
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('cart.db')
            cursor = conn.execute("SELECT name,Price,quantity from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            conn.commit()
            global xs
            xs="Aquí está tu carrito:\nDetalles\n"
            for row in cursor: 
                m="   "+row[0]+" x "+row[2] + "   =   "+row[1]+"\n"   
                xs=xs+m
            cursor = conn.execute("SELECT SUM(Price),SUM(Price) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            global aps
            aps=""
            global njs
            njs=""
            for row in cursor:
                cer="\n𝐓𝐨𝐭𝐚𝐥 ="+str(row[0])+"$"
            keyboard =[[InlineKeyboardButton("✅ Verificar", callback_data="25s"),InlineKeyboardButton("❌ Carro", callback_data="29s")],[InlineKeyboardButton("🔙 Cancelar", callback_data="spanish")]]   
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
            context.bot.send_message(chat_id=update.effective_user.id,text=xs+cer ,reply_markup=reply_markup)                    
            return BUTTON
        else:
            keyboard = [[InlineKeyboardButton("🔙 Menú principal", callback_data='spanish')]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            context.bot.send_message(chat_id=update.effective_user.id,text="El carrito esta vacío",reply_markup=reply_markup)                       
            return BUTTON
    elif a== '25':
            query.answer()
            conn = sqlite3.connect('cart.db')
            cursor = conn.execute("SELECT COUNT(Price),SUM(Price) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            conn.commit()
            global summ
            summ=" "
            global cou
            cou=" "
            for row in cursor: 
                summ=float(row[1])
                cou=row[0]
            connection = sqlite3.connect("wallet.db")  
            cursor = connection.cursor()  
            cursor.execute("SELECT balance FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
            for names in cursor:
              inv=float(names[0])
            if summ>inv:
              keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
              context.bot.send_message(chat_id=update.effective_user.id,text="You dont have enough balance for this purchase",reply_markup=reply_markup)
              return BUTTON

            else:
              keyboard =[[InlineKeyboardButton("Pay from Wallet", callback_data="995")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
              context.bot.send_message(chat_id=update.effective_user.id,text="Payment Confirmation\nTotal products= "+str(cou)+"\nTotal amount= "+str(summ)+"$",reply_markup=reply_markup)
              return BUTTON
    elif a== '25s':
            query.answer()
            conn = sqlite3.connect('cart.db')
            cursor = conn.execute("SELECT COUNT(Price),SUM(Price) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            conn.commit()
            global summs
            summs=" "
            global cous
            cous=" "
            for row in cursor: 
                summs=float(row[1])
                cous=row[0]
            connection = sqlite3.connect("wallet.db")  
            cursor = connection.cursor()  
            cursor.execute("SELECT balance FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
            for names in cursor:
              inv=float(names[0])
            if summs>inv:
              keyboard =[[InlineKeyboardButton("🔙 atrás", callback_data="spanish")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
              context.bot.send_message(chat_id=update.effective_user.id,text="No tienes saldo suficiente para esta compra",reply_markup=reply_markup)
              return BUTTON

            else:
              keyboard =[[InlineKeyboardButton("Pagar desde Monedero", callback_data="995s")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
              context.bot.send_message(chat_id=update.effective_user.id,text="Confirmación de pago\nTotal de productos= "+str(cous)+"\nImporte total= "+str(summs)+"$",reply_markup=reply_markup)
              return BUTTON
    elif a== '995s':
        keyboard =[[InlineKeyboardButton("🌎Menú principal", callback_data="spanish")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="¡Proporcione una dirección de correo electrónico para que podamos notificarle cuando se lleve a cabo el sorteo o para notificarle si gana!",reply_markup=reply_markup)
        return TRADES
    elif a== '25f':
            query.answer()
            conn = sqlite3.connect('cart.db')
            cursor = conn.execute("SELECT COUNT(Price),SUM(Price) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            conn.commit()
            global summf
            summf=" "
            global couf
            couf=" "
            for row in cursor: 
                summf=float(row[1])
                couf=row[0]
            connection = sqlite3.connect("wallet.db")  
            cursor = connection.cursor()  
            cursor.execute("SELECT balance FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
            for names in cursor:
              inv=float(names[0])
            if summf>inv:
              keyboard =[[InlineKeyboardButton("🔙 Menu principal", callback_data="french")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
              context.bot.send_message(chat_id=update.effective_user.id,text="Vous n'avez pas assez de solde pour cet achat",reply_markup=reply_markup)
              return BUTTON

            else:
              keyboard =[[InlineKeyboardButton("Payer depuis Wallet", callback_data="995f")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
              context.bot.send_message(chat_id=update.effective_user.id,text="Sélectionnez l'option Paiement\nProduits totaux= "+str(couf)+"\nMontant total= "+str(summf)+"$",reply_markup=reply_markup)
              return BUTTON
    elif a== '995':
        keyboard =[[InlineKeyboardButton("Return to Cart", callback_data="22")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Please provide an email address so that we can notify you when the draw is taking place or to notify you if you win!",reply_markup=reply_markup)
        return TRADE
    elif a== '995f':
              context.bot.send_message(chat_id=update.effective_user.id,text="merci de bien vouloir indiquer votre adresse")
              return TRADEF
    elif a== '255':
  
        cc=update.callback_query.message.caption
        fg=cc
        cf=update.callback_query.message.message_id
        cc=cc.split("𝐈𝐃:  ")
        cc=cc[1]
        cc=cc.split("𝐏𝐫𝐢𝐜𝐞:   ")
        cc=cc[0]
        cc=cc.strip()
        pid=cc
        conn = sqlite3.connect('cart.db')
        cursor = conn.execute("DELETE  from COMPANY where productID='{}'".format(pid))
        conn.commit()
        keyboard =[[InlineKeyboardButton("➕ Add to cart", callback_data="16"),InlineKeyboardButton("Return to Shop", callback_data="200")],[InlineKeyboardButton("➕ Quantity", callback_data="160"),InlineKeyboardButton("➖ Quantity", callback_data="161")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=fg,reply_markup=reply_markup)
        return BUTTON
    elif a== '255s':
  
        cc=update.callback_query.message.caption
        fg=cc
        cf=update.callback_query.message.message_id
        cc=cc.split("𝐈𝐃:  ")
        cc=cc[1]
        cc=cc.split("𝐏𝐫𝐞𝐜𝐢𝐨:   ")
        cc=cc[0]
        cc=cc.strip()
        pid=cc
        conn = sqlite3.connect('cart.db')
        cursor = conn.execute("DELETE  from COMPANY where productID='{}'".format(pid))
        conn.commit()
        keyboard =[[InlineKeyboardButton("➕ Añadir a la cesta", callback_data="16s"),InlineKeyboardButton("🔙 Cancelar", callback_data="spanish")],[InlineKeyboardButton("➕ Cantidad", callback_data="160s"),InlineKeyboardButton("➖ Cantidad", callback_data="161s")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=fg,reply_markup=reply_markup)
        return BUTTON
    elif a== '255f': 
        cc=update.callback_query.message.caption
        fg=cc
        cf=update.callback_query.message.message_id

        conn = sqlite3.connect('cart.db')
        cursor = conn.execute("DELETE  from COMPANY where ID='{}'".format(update.effective_user.id))
        conn.commit()
        keyboard =[[InlineKeyboardButton("➕ Add", callback_data="16f"),InlineKeyboardButton("🔙Menu principal", callback_data="french")],[InlineKeyboardButton("➕ Quantité", callback_data="160f"),InlineKeyboardButton("➖ Quantité", callback_data="161f")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=fg,reply_markup=reply_markup)
        return BUTTON
    elif a== '22f':     
      query.answer()
      conn = sqlite3.connect('cart.db')
      cursor = conn.execute("SELECT name,Price,quantity from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
      conn.commit()
      global xruf
      xruf="Voici votre panier:\n\n"
      for row in cursor: 

 
        m=row[2]+" x "+row[0]+ "  =   "+row[1]+"\n"
      
        xruf=xruf+m
      cursor = conn.execute("SELECT SUM(Price),SUM(quantity) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
      global apf
      apf=""
      global njf
      njf=""
      for row in cursor:
          apf=str(row[0])
      keyboard =[[InlineKeyboardButton("✅ Vérifier", callback_data="25f"),InlineKeyboardButton("❌ Chariot", callback_data="29f")],[InlineKeyboardButton("🔙 Annuler", callback_data="french")]]   
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text=xruf+"\nTotal"+" "+apf+"$" ,reply_markup=reply_markup)    
        
      return BUTTON
    elif a== '200':
        keyf=[]
        connection = sqlite3.connect("cata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT cat FROM COMPANY")
        keyf=[]
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[0])]
            keyf.append(c)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Select Main Category",reply_markup=reply_markup) 
 
        return PDB
    elif a== '29':
        keyf=[]
        connection = sqlite3.connect("cart.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT name,productID FROM COMPANY where ID='{}'".format(update.effective_user.id))
        keyf=[]
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[1])]
            keyf.append(c)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Select Product to delete",reply_markup=reply_markup) 
        return JOP

    elif a== '29s':
        keyf=[]
        connection = sqlite3.connect("cart.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT name,productID FROM COMPANY where ID='{}'".format(update.effective_user.id))
        keyf=[]
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[1])]
            keyf.append(c)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Seleccionar producto para eliminar",reply_markup=reply_markup) 
        return JOPS
    elif a== '29f':
        keyf=[]
        connection = sqlite3.connect("cart.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT name,productID FROM COMPANY where ID='{}'".format(update.effective_user.id))
        keyf=[]
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[1])]
            keyf.append(c)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Sélectionnez le produit à supprimer",reply_markup=reply_markup) 
        return JOPF
    elif a=='1bal':
            connection = sqlite3.connect("deposit.db")  
            cursor = connection.cursor()  
            cursor.execute("SELECT name,address FROM COMPANY")
            keyf=[]
            for row in cursor:
                c=[InlineKeyboardButton("{}".format(row[0]), callback_data=row[1])]
                keyf.append(c)
            b=[InlineKeyboardButton("🔙Back", callback_data='🔙Back')] 
            keyf.append(b)
            connection.close() 
            reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
            context.bot.send_message(chat_id=update.effective_user.id,text="Choose the Currency you Would Like to Use to Fund Your Wallet",reply_markup=reply_markup) 
            return BYE
    elif a=='1bals':
            connection = sqlite3.connect("deposit.db")  
            cursor = connection.cursor()  
            cursor.execute("SELECT name,address FROM COMPANY")
            keyf=[]
            for row in cursor:
                c=[InlineKeyboardButton("{}".format(row[0]), callback_data=row[1])]
                keyf.append(c)
            b=[InlineKeyboardButton("🔙atrás", callback_data='🔙 atrás')] 
            keyf.append(b)
            reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
            context.bot.send_message(chat_id=update.effective_user.id,text="Seleccione Monedero para obtener detalles",reply_markup=reply_markup) 
            return BYES

    elif a=='1balf':
        connection = sqlite3.connect("deposit.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT name,address FROM COMPANY")
        keyf=[]
        jkl=' '
        for row in cursor:
          m="Porte monnaie: {}\n Adresse: {}\n\n".format(row[0],row[1])
          jkl=m+jkl
        keyboard =[[InlineKeyboardButton("J'ai payé", callback_data='999zf'),InlineKeyboardButton("🌎Menu principal", callback_data="french")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Déposez le montant sur n'importe quelle adresse indiquée ci-dessous, puis soumettez une capture d'écran de la preuve. MERCI\n\n"+jkl,reply_markup=reply_markup)
        return BUTTON
    elif a== '999z':
        query.answer()

        context.bot.send_message(chat_id=update.effective_user.id,text="Thank You for Supporting Us. Please Send a Screenshot of Proof of Payment and we will Update Your Wallet as Soon as the Funds Have Arrived") 
                          
        return PROFT
    elif a== '999zs':
        query.answer()
        keyboard = [[InlineKeyboardButton("🔙 Main Menu", callback_data='spanish')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.bot.send_message(chat_id=update.effective_user.id,text="Enviar captura de pantalla del comprobante de pago. Gracias",reply_markup=reply_markup) 
                          
        return PROFTS
    elif a== '999zf':
        query.answer()
        keyboard = [[InlineKeyboardButton("🔙 Menu principal", callback_data='french')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.bot.send_message(chat_id=update.effective_user.id,text="Envoyer une capture d'écran de la preuve de paiement. Merci",reply_markup=reply_markup)                       
        return PROFTF
    elif a== '1abb':
        cc=update.callback_query.message.caption
        
        uo=cc
        dv=update.callback_query.message.photo[-1].file_id
        fg=cc
        cf=update.callback_query.message.message_id
        df=cc.split("ID:")
        df=df[1]
        df=df.split("Status:")
        df=df[0]
        df=df.strip()
        print(df)
        global shaf
        shaf=df
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Send amount to add in WalletID: {}'.format(shaf),reply_markup=reply_markup)
        return ADDB 
    elif a== '1acc':
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        uo=cc
        dv=update.callback_query.message.photo[-1].file_id
        fg=cc
        cf=update.callback_query.message.message_id
        df=cc.split("WalletID: ")
        df=df[1]
        df=df.strip()

        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Payment Proof Rejected for WalletID" {}'.format(shaf),reply_markup=reply_markup)
        context.bot.send_message(chat_id=update.effective_user.id,text="Your payment proof has been rejected by admin. If you have anu query contact Admin. Thanks")
        return BUTTON   
    elif a== '99':
      query.answer()  
      key=[["Pending","History"],["Cancel"]]
      reply_markup = ReplyKeyboardMarkup(key,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text="Select the Order Type:",reply_markup=reply_markup)
      return HAM 
    elif a=='120a':
        c=update.callback_query.message.message_id
        cc=update.callback_query.message.text
        d=cc
        df=cc.split("UserID:  ")
        df=df[1]
        df=df.strip()
        df=df.split("Withdrawl Amount:  ")
        df=df[0]
        df=df.strip()

        dfg=cc.split("Withdrawl Amount:  ")
        dfg=dfg[1]
        dfg=dfg.strip()
        dfg=dfg.split("Wallet Address:   ")
        dfg=dfg[0]
        dfg=dfg.strip()
        summ=float(dfg)
        dd=cc.split("Withdrawl ID:  ")
        dd=dd[1]
        dd=dd.strip()
        hhhh="Your withdrawl request is being processed".format(df) 
        context.bot.send_message(chat_id=df,text=hhhh)  
        connection = sqlite3.connect("wallet.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT balance,twd FROM COMPANY where id= {}".format(int(df)) )
        for names in cursor:
          inv=float(names[0])
          inv1=float(names[1])
        bn=inv-summ
        bn1=inv1+summ
        bn=str(bn)
        conn = sqlite3.connect("wallet.db")  
        conn.execute("UPDATE COMPANY set balance = '{}',twd= '{}' where ID = {}".format(bn,bn1,int(df)))
        conn.commit()
        conn.close()
        conn = sqlite3.connect('withdraw.db')
        cursor = conn.execute("UPDATE COMPANY set status='Accept' where rid='{}'".format(dd))
        conn.commit() 
        context.bot.edit_message_text(chat_id=update.effective_user.id,message_id=c,text="{}\n ACCEPTED".format(d))
        return BUTTON    
    elif a=='120':
        c=update.callback_query.message.message_id
        cc=update.callback_query.message.text
        d=cc
        df=cc.split("OrderID: ")
        df=df[1]
        df=df.strip()
        df=df.split("Order Status: ")
        df=df[0]
        df=df.strip()
        dd=cc.split("UserID:  ")
        dd=dd[1]
        dd=dd.split("Email:  ")
        dd=dd[0]
        dd=dd.strip()
        global dkop
        dkop=dd
        hhhh="Your Order has Been Completed\n𝐎𝐫𝐝𝐞𝐫 𝐧𝐮𝐦𝐛𝐞𝐫# {}".format(df) 
        context.bot.send_message(chat_id=dd,text=hhhh)  
        conn = sqlite3.connect('orders.db')
        cursor = conn.execute("UPDATE COMPANY set status='Accept' where oid='{}'".format(df))
        conn.commit() 
        context.bot.edit_message_text(chat_id=update.effective_user.id,message_id=c,text="{}\n ACCEPTED".format(d))
        return BUTTON
    elif a=='130':
        c=update.callback_query.message.message_id
        cc=update.callback_query.message.text
        d=cc
        df=cc.split("OrderID: ")
        df=df[1]
        df=df.strip()
        df=df.split("Order Status: ")
        df=df[0]
        df=df.strip()
        dd=cc.split("UserID:  ")
        dd=dd[1]
        dd=dd.split("Email:  ")
        dd=dd[0]
        dd=dd.strip()
        hhhh="Your order rejected by the admin\n𝐎𝐫𝐝𝐞𝐫 𝐧𝐮𝐦𝐛𝐞𝐫# {}".format(df)  
        conn = sqlite3.connect('orders.db')
        cursor = conn.execute("UPDATE COMPANY set status='Rejected' where oid='{}'".format(df))
        conn.commit() 
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=dd,text=hhhh)   
        context.bot.edit_message_text(chat_id=update.effective_user.id,message_id=c,text="{}\n Rejected".format(d),reply_markup=reply_markup)
        return BUTTON
    elif a=='130a':
        c=update.callback_query.message.message_id
        cc=update.callback_query.message.text
        d=cc
        df=cc.split("UserID:  ")
        df=df[1]
        df=df.strip()
        df=df.split("Withdrawl Amount:  ")
        df=df[0]
        df=df.strip()
        dfg=cc.split("Withdrawl Amount:  ")
        dfg=dfg[1]
        dfg=dfg.strip()
        dfg=dfg.split("Wallet Address:   ")
        dfg=dfg[0]
        dfg=dfg.strip()
        summ=float(dfg)
        dd=cc.split("Withdrawl ID:  ")
        dd=dd[1]
        dd=dd.strip()
        hhhh="Your Withdrawl request rejected by Admin"
        conn = sqlite3.connect('withdraw.db')
        cursor = conn.execute("UPDATE COMPANY set status='Rejected' where rid='{}'".format(dd))
        conn.commit() 
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=df,text=hhhh)   
        context.bot.edit_message_text(chat_id=update.effective_user.id,message_id=c,text="{}\n Rejected".format(d),reply_markup=reply_markup)
        return BUTTON
    elif a=='100':
      query.answer()  
      keyboard =[[InlineKeyboardButton("Products", callback_data="pak") ,InlineKeyboardButton("Wallet", callback_data="wallet")],  
                  [InlineKeyboardButton("Communications", callback_data="comm")]]

      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to the Admin Dashboard" ,reply_markup=reply_markup)
      return BUTTON
    elif a== '912':
        keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Send Welcome Message',reply_markup=reply_markup)
        return SH
    elif a== '916':
        keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text='Send Announcement Message',reply_markup=reply_markup)
        return JE

def bye(update,context):
    query = update.callback_query
    msg=query.data
    print(msg)
    conn = sqlite3.connect('deposit.db')
    cursor = conn.execute("SELECT name,address FROM COMPANY where address='{}'".format(msg))
    conn.commit()
    for row in cursor:
      m="Wallet: {}\n Address: {}\n\n".format(row[0],row[1])
      jkl=row[1]
    img = qrcode.make(row[1])
    type(img)   # qrcode.image.pil.PilImage
    img.save("qr.png")
    keyboard =[[InlineKeyboardButton("I Have Paid", callback_data='999z')]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_photo(chat_id=update.effective_user.id,photo=open('qr.png','rb'),caption="Scan to get Address")
    context.bot.send_message(chat_id=update.effective_user.id,text=jkl)
    context.bot.send_message(chat_id=update.effective_user.id,text="To Load Your Wallet Send the Correct Currency to the Wallet Above, Scan the Wallet QR Code or Copy the Address Above. Once You Have Made Your Payment Click “I Have Paid”. **Warning** Make Sure you are Sending the Correct Currency to the Correct Wallet Address. This Will be a Loss that we Cannot Recover for you!",reply_markup=reply_markup)
    return BUTTON
def byes(update,context):
    query = update.callback_query
    msg=query.data
    connection = sqlite3.connect("deposit.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT name,address FROM COMPANY where address='{}'".format(str(msg)))
    keyf=[]
    jkl=' '
    for row in cursor:
      m="Cartera: {}\n Dirección: {}\n\n".format(row[0],row[1])
      jkl=row[1]
    img = qrcode.make(row[1])
    type(img)   # qrcode.image.pil.PilImage
    img.save("qr.png")
    keyboard =[[InlineKeyboardButton("He pagado", callback_data='999zs'),InlineKeyboardButton("🌎Menú principal", callback_data="spanish")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_photo(chat_id=update.effective_user.id,photo=open('qr.png','rb'),caption="Escanear para obtener la dirección")
    context.bot.send_message(chat_id=update.effective_user.id,text=jkl)
    context.bot.send_message(chat_id=update.effective_user.id,text="Detalles de la dirección de la billetera",reply_markup=reply_markup)
    return BUTTON
def wwrr(update,context):
    query = update.callback_query
    msg=query.data
    print(msg)
    if msg=='MAINN':
      keyboard =[[InlineKeyboardButton("Products", callback_data="pak") ,InlineKeyboardButton("Wallet", callback_data="wallet")],  
                  [InlineKeyboardButton("🛒 Orders", callback_data="99"),InlineKeyboardButton("🔊 Announcement", callback_data="916")]]

      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to the Admin Dashboard" ,reply_markup=reply_markup)
      return BUTTON
    else:
      conn = sqlite3.connect('withdraw.db')
      cursor = conn.execute("SELECT ID,balance,status,wid,email,rid from COMPANY where rid ='{}'".format(msg))
      conn.commit()
      for row in cursor: 
          g="\nUserID:  "+str(row[0])+"\nWithdrawl Amount:  "+str(row[3])+"\nWallet Address:   "+str(row[4])+'\nCurrency:  '+row[6]+"\nWithdrawl ID:  "+str(row[5])+"\n\n"
          keyboard =[[InlineKeyboardButton("Accept", callback_data="120a"),InlineKeyboardButton("Reject", callback_data="130a")],[InlineKeyboardButton("Return to Shop", callback_data="100")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
          context.bot.send_message(chat_id=update.effective_user.id,text=str(row[4]))
          context.bot.send_message(chat_id=update.effective_user.id,text="Withdrawl Request:\n"+g,reply_markup=reply_markup)
          return BUTTON
def byef(update,context):
    query = update.callback_query
    msg=query.data
    connection = sqlite3.connect("deposit.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT name,address FROM COMPANY where address='{}'".format(msg))
    keyf=[]
    jkl=' '
    for row in cursor:
      m="Wallet: {}\n Address: {}\n\n".format(row[0],row[1])
      jkl=row[1]
    img = qrcode.make(row[1])
    type(img)  # qrcode.image.pil.PilImage
    img.save("qr.png")
    keyboard =[[InlineKeyboardButton("I have Paid", callback_data='999z'),InlineKeyboardButton("🌎Main Menu", callback_data="200")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_photo(chat_id=update.effective_user.id,photo=open('qr.png','rb'),caption="Scan to get Address")
    context.bot.send_message(chat_id=update.effective_user.id,text=jkl)
    context.bot.send_message(chat_id=update.effective_user.id,text="Wallet Address Details",reply_markup=reply_markup)
    return BUTTON
def penord(update,context):
    query = update.callback_query
    msg=query.data
    print(msg)
    if msg=="🔙Back":
      key=[["⏳ - PENDING","History"],["Cancel"]]
      reply_markup = ReplyKeyboardMarkup(key,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text="Select the Order Type:",reply_markup=reply_markup)
      return HAM 
    else:
        conn = sqlite3.connect('orders.db')
        cursor = conn.execute("SELECT ID,price,oid,status,details,oid,date,address,name,time from COMPANY where oid ='{}'".format(msg))
        conn.commit()
        for row in cursor: 
            g="\nDate:  "+row[6]+"\nTime: "+row[9]+"\nUserID:  "+row[0]+"\nEmail:  "+row[7]+"\nProduct Detail: "+row[4]+"\nPrice: "+row[1]+"\nOrderID: "+row[5]+"\nOrder Status: "+row[3]+"\n\n"
            keyboard =[[InlineKeyboardButton("Accept", callback_data="120"),InlineKeyboardButton("Reject", callback_data="130")],[InlineKeyboardButton("🔙 Cancel", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Pending order:\n"+g,reply_markup=reply_markup)
            return BUTTON
def dbs(update,context):
  msg=update.message.text
  try:
    msg=int(msg)
    conn = sqlite3.connect('wallet.db')
    cursor = conn.execute("SELECT balance from COMPANY where ID= {}   ".format(int(update.effective_user.id))) 
    conn.commit()
    for row in cursor:  
        qoop=float(row[0]) 
    if float(msg)>=qoop:
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text="Sorry You cannot withdraw this amount. Try with less amount. Thanks",reply_markup=reply_markup)
      return DBS
    else:
      conn = sqlite3.connect("wallet.db")  
      conn.execute("UPDATE COMPANY set link = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
      conn.commit()
      conn.close()
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text="In What Currency Would you Like to be Paid? **Example** ETH, BTC, USDT",reply_markup=reply_markup)   
      return CUITQ
  except:
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text="Send in digit only",reply_markup=reply_markup)   
      return DBS
def dbss(update,context):
  msg=update.message.text
  try:
    msg=int(msg)
    conn = sqlite3.connect('wallet.db')
    cursor = conn.execute("SELECT balance from COMPANY where ID= {}   ".format(int(update.effective_user.id))) 
    conn.commit()
    for row in cursor:  
        qoop=float(row[0]) 
    if float(msg)>=qoop:

      context.bot.send_message(chat_id=update.effective_user.id,text="Lo sentimos, no puede retirar esta cantidad. Prueba con menos cantidad. Gracias")
      return DBSS
    else:
      conn = sqlite3.connect("wallet.db")  
      conn.execute("UPDATE COMPANY set link = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
      conn.commit()
      conn.close()
      context.bot.send_message(chat_id=update.effective_user.id,text="¿En qué moneda le gustaría que le paguen? **Ejemplo** ETH, BTC, USDT")   
      return CUITQS
  except:
      context.bot.send_message(chat_id=update.effective_user.id,text="Enviar solo en dígitos")   
      return DBSS
def cuitq(update,context):
      msg=update.message.text
      conn = sqlite3.connect("wallet.db")  
      conn.execute("UPDATE COMPANY set curr = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
      conn.commit()
      conn.close()
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text="Provide your wallet address to receive this amount",reply_markup=reply_markup)   
      return CUIT
def cuitqf(update,context):
      msg=update.message.text
      conn = sqlite3.connect("wallet.db")  
      conn.execute("UPDATE COMPANY set curr = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
      conn.commit()
      conn.close()
      context.bot.send_message(chat_id=update.effective_user.id,text="Indiquez votre adresse de portefeuille pour recevoir ce montant")   
      return CUITF
def cuitqs(update,context):
      msg=update.message.text
      conn = sqlite3.connect("wallet.db")  
      conn.execute("UPDATE COMPANY set curr = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
      conn.commit()
      conn.close()
      context.bot.send_message(chat_id=update.effective_user.id,text="Proporcione la dirección de su billetera para recibir esta cantidad")   
      return CUITS
def dbsf(update,context):
  msg=update.message.text
  try:
    msg=int(msg)
    conn = sqlite3.connect('wallet.db')
    cursor = conn.execute("SELECT balance from COMPANY where ID= {}   ".format(int(update.effective_user.id))) 
    conn.commit()
    for row in cursor:  
        qoop=float(row[0]) 
    if float(msg)>=qoop:
      context.bot.send_message(chat_id=update.effective_user.id,text="Désolé, vous ne pouvez pas retirer ce montant. Essayez avec moins de quantité. Merci")
      return DBSF
    else:
      conn = sqlite3.connect("wallet.db")  
      conn.execute("UPDATE COMPANY set link = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
      conn.commit()
      conn.close()
      context.bot.send_message(chat_id=update.effective_user.id,text="Dans quelle devise souhaitez-vous être payé ? **Exemple** ETH, BTC, USDT")   
      return CUITQF
  except:
      context.bot.send_message(chat_id=update.effective_user.id,text="Envoyer en chiffres uniquement")   
      return DBSF
def cuit(update,context):
    msg=update.message.text
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set code = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    context.bot.send_message(chat_id=update.effective_user.id,text="Provide your email address")   
    return CUI
def cuits(update,context):
    msg=update.message.text
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set code = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    context.bot.send_message(chat_id=update.effective_user.id,text="Proporcione su dirección de correo electrónico")   
    return CUIS#DBSF,CUIF,CUITF
def cuitf(update,context):
    msg=update.message.text
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set code = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("🔙 Menu principal", callback_data="french")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Indiquez votre adresse e-mail",reply_markup=reply_markup)   
    return CUIF

def cui(update,context):
    msg=update.message.text
    connection = sqlite3.connect("wallet.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT link,balance,code,curr FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
    for row in cursor:
      print(row[0])
      today = date.today()
      d1 = today.strftime("%d/%m/%Y")
      now = datetime.now()

      current_time = now.strftime("%H:%M:%S")
      yu= random.randint (0,999999)
      conn = sqlite3.connect('withdraw.db')
      conn.execute("INSERT INTO COMPANY (ID,balance,status,wid,email,rid,date,time,address,currency) \
                          VALUES ('{}', '{}','{}', '{}','{}','{}', '{}','{}','{}','{}')".format(str(update.effective_user.id),row[1],"Pending",row[0],msg,yu,d1,current_time,row[2],row[3]))
      conn.commit()
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Your Request has Been Received and Will be Processed as Quickly as Possible. We will Contact you with Payment Confirmation. Thank you for Supporting Us!" ,reply_markup=reply_markup)
      return BUTTON
def cuis(update,context):
    msg=update.message.text
    connection = sqlite3.connect("wallet.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT link,balance,code,curr FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
    for row in cursor:
      print(row[0])
      today = date.today()
      d1 = today.strftime("%d/%m/%Y")
      now = datetime.now()

      current_time = now.strftime("%H:%M:%S")
      yu= random.randint (0,999999)
      conn = sqlite3.connect('withdraw.db')
      conn.execute("INSERT INTO COMPANY (ID,balance,status,wid,email,rid,date,time,address,currency) \
                          VALUES ('{}', '{}','{}','{}', '{}','{}','{}', '{}','{}','{}')".format(str(update.effective_user.id),row[1],"Pending",row[0],msg,yu,d1,current_time,row[2],row[3]))
      conn.commit()
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Su solicitud ha sido recibida y será procesada lo más rápido posible. Nos pondremos en contacto con usted con la confirmación de pago. ¡Gracias por apoyarnos!" ,reply_markup=reply_markup)
      return BUTTON 
def cuif(update,context):
    msg=update.message.text
    connection = sqlite3.connect("wallet.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT link,balance,code,curr FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
    for row in cursor:
      print(row[0])
      today = date.today()
      d1 = today.strftime("%d/%m/%Y")
      now = datetime.now()

      current_time = now.strftime("%H:%M:%S")
      yu= random.randint (0,999999)
      conn = sqlite3.connect('withdraw.db')
      conn.execute("INSERT INTO COMPANY (ID,balance,status,wid,email,rid,date,time,address,currency) \
                          VALUES ('{}', '{}','{}', '{}','{}','{}', '{}','{}','{}','{}')".format(str(update.effective_user.id),row[1],"Pending",row[0],msg,yu,d1,current_time,row[2],row[3]))
      conn.commit()
      keyboard =[[InlineKeyboardButton("🔙 Menu principal", callback_data="french")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Votre demande a été reçue et sera traitée le plus rapidement possible. Nous vous contacterons avec la confirmation de paiement. Merci de nous soutenir !" ,reply_markup=reply_markup)
      return BUTTON     
def ham(update,context):
  msg=update.message.text
  if msg=="Pending":
            conn = sqlite3.connect('orders.db')
            cursor = conn.execute("SELECT oid,name from COMPANY where status= '{}' ".format("pending"))
            conn.commit()
            keyf=[]
            for row in cursor:
                c=[InlineKeyboardButton("⏳ "+"{},{}".format(row[0],row[1]), callback_data=row[0])]
                keyf.append(c)
            b=[InlineKeyboardButton("🔙Back", callback_data='🔙Back')] 
            keyf.append(b)
            reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
            context.bot.send_message(chat_id=update.effective_user.id,text="Select Order to check detail",reply_markup=reply_markup) 
            return PENORD
  elif msg=="History":
        style = xlwt.easyxf('font: bold 1, color black;')
        conn = sqlite3.connect('orders.db') 
        cursor = conn.execute("SELECT ID,price from COMPANY ")
        conn.commit() 
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            f = open("orders.txt", "w",encoding="utf-8")
            conn = sqlite3.connect('orders.db') 
            cursor = conn.execute("SELECT ID,price,oid,status,details,oid,date,address,name,time from COMPANY ")
            conn.commit()
            xa="Orders\n"
            i=0
            j=0
            workbook = xlwt.Workbook()
            sheet = workbook.add_sheet("Order History")
            sheet.write(i, 1, "Order_Date",  style)
            sheet.write(i, 2, "Order_Time",  style)
            sheet.write(i, 3, "User_Name", style)
            sheet.write(i, 4, "User_ID",  style)
            sheet.write(i, 5, "Email", style)
            sheet.write(i, 6,"Product_name",  style)
            sheet.write(i, 7, "Price",  style)
            sheet.write(i, 8, "Order_ID", style)
            sheet.write(i, 9, "Order_Status",  style)
            for row in cursor:
                    i=i+1
                    j=j+1
                    inv=row[0]
                    vgy=row[1]
                    xdrt=row[2]
                    sheet.write(i, 1, row[6], style)
                    sheet.write(i, 2, row[9], style)
                    sheet.write(i, 3, row[8],  style)
                    sheet.write(i, 4, row[0],  style)
                    sheet.write(i, 5, row[7],  style)
                    sheet.write(i, 6, row[4],  style)
                    sheet.write(i, 7, row[1], style)
                    sheet.write(i, 8, row[5], style)
                    sheet.write(i, 9, row[3], style)
                    
            workbook.save("orders.xls")
            keyboard =[[InlineKeyboardButton("Main Menu", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
            context.bot.send_document(chat_id=update.effective_user.id,document=open('orders.xls', 'rb'),reply_markup=reply_markup)
        else:
                keyboard =[[InlineKeyboardButton("❌", callback_data="100")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
                context.bot.send_message(chat_id=update.effective_user.id,text="You have no orders",reply_markup=reply_markup)
                return BUTTON
def fika(update,context):
    msg=a=update.message.text
    global ashr
    ashr= msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Send Answer of FAQ",reply_markup=reply_markup)  
    return FIKAA
def fikaa(update,context):
    a=update.message.text
    conn = sqlite3.connect('faqs.db')
    conn.execute("INSERT INTO COMPANY (ques,ans) \
            VALUES ('{}','{}')".format(ashr,a)) 
    conn.commit() 
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="FAQ added",reply_markup=reply_markup)  
    return BUTTON
def fikaf(update,context):
    a=msg=update.message.text
    global ashrf
    ashrf= msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Send Answer of FAQ in French",reply_markup=reply_markup)  
    return FIKAAF
def fikaaf(update,context):
    msg=update.message.text
    conn = sqlite3.connect('faqsf.db')
    conn.execute("INSERT INTO COMPANY (ques,ans) \
            VALUES ('{}','{}')".format(ashrf,a)) 
    conn.commit() 
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="FAQ added",reply_markup=reply_markup)  
    return BUTTON
def fikas(update,context):
    msg=a=update.message.text
    global ashrs
    ashrs= msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Send Answer of FAQ in Spanish",reply_markup=reply_markup)  
    return FIKAAS
def fikaas(update,context):
    msg=update.message.text
    conn = sqlite3.connect('faqss.db')
    conn.execute("INSERT INTO COMPANY (ques,ans) \
            VALUES ('{}','{}')".format(ashrs,a)) 
    conn.commit() 
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="FAQ Added",reply_markup=reply_markup)  
    return BUTTON
def trade(update,context):
    msg=update.message.text
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set email = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("Return to Cart", callback_data="22")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Please provide your mobile number if you would like us to contact you by text or type “skip” to decline texting",reply_markup=reply_markup)
    return YUO
def trades(update,context):
    msg=update.message.text
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set email = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("🌎Menú principal", callback_data="spanish")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Proporcione su número de teléfono móvil si desea que nos comuniquemos con usted por mensaje de texto o escriba omitir",reply_markup=reply_markup)
    return YUOS
def yuo(update,context):
    msg=update.message.text
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set exp = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("Return to Cart", callback_data="200")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Would you like us to contact you another way? Please send the details now or type “skip” to decline additional contact methods",reply_markup=reply_markup)   
    return ADDRESS
def yuos(update,context):
    msg=update.message.text
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set exp = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("🌎Menú principal", callback_data="spanish")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="¿Quieres que contactemos contigo de otra forma? Envíe los detalles ahora o escriba omitir para rechazar métodos de contacto adicionales",reply_markup=reply_markup)   
    return ADDRESSS
def tradef(update,context):
    msg=update.message.text
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set naam = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    context.bot.send_message(chat_id=update.effective_user.id,text="Envoyer Votre numéro de contact (facultatif) ou envoyer Passer")
    return YUOF
def yuof(update,context):
    msg=update.message.text
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set exp = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="200")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Souhaitez-vous que nous vous contactions d'une autre manière ? Veuillez envoyer les détails maintenant ou tapez ignorer pour refuser d'autres méthodes de contact",reply_markup=reply_markup)   
    return ADDRESSF
def address(update,context):
    msg=str(update.message.text)
    bnm=update.message.from_user
    ms=bnm.first_name
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("SELECT COUNT(Price),SUM(Price) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
    conn.commit()
    global summ
    summ=" "
    global cou
    cou=" "
    for row in cursor: 
      summ=float(row[1])
      cou=row[0]
    connection = sqlite3.connect("wallet.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT balance,naam,email,exp,name FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
    for names in cursor:
      inv=float(names[0])
      rdus=names[1]
      rdusd=names[2]
      rdrt=names[3]
    bn=inv-summ
    bn=str(bn)
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set balance = '{}',email='{}' where ID = {}".format(bn,msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    naak=''
    wqp=''
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("SELECT ID,name, quantity,Price,productID from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
    conn.commit()
    for row in cursor:
        v= random.randint (0,999999)
        today = date.today()
        dat = today.strftime("%d/%m/%Y")
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        deepe=row[3]
        sib=row[4]
        wqp=row[2]
        uop=row[1]+' * '+row[2]+ '='+'$'+row[3]+'\n'    
        connection = sqlite3.connect("shop.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT stock FROM COMPANY where productID= '{}'".format(sib) )
        for names in cursor:
          inv=float(names[0])
        bn=inv-float(wqp)
        bn=str(bn)
        conn = sqlite3.connect("shop.db")  
        conn.execute("UPDATE COMPANY set stock = '{}' where productID = '{}'".format(bn,sib))
        conn.commit()
        conn.close()
        naak=naak+uop
    print(naak)
    conn = sqlite3.connect('orders.db')
    conn.execute("INSERT INTO COMPANY (ID,name,oid,address,date,details,status,opt,contact,time,price) \
            VALUES ('{}', '{}', '{}', '{}','{}','{}','{}','{}','{}','{}','{}')".format(str(update.effective_user.id),ms,str(v),rdusd,dat,naak,"pending",rdrt,msg,current_time,summ)) 
    conn.commit()
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("DELETE  from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
    conn.commit()
    keyboard =[[InlineKeyboardButton("Return Home", callback_data="200")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="You order has been placed you will recieve conformation message shortly",reply_markup=reply_markup) 
    context.bot.send_message(chat_id=5791144147,text="You have a new order, check order section")
    return BUTTON
def addresss(update,context):
    msg=str(update.message.text)
    bnm=update.message.from_user
    ms=bnm.first_name
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("SELECT COUNT(Price),SUM(Price) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
    conn.commit()
    global summs
    summs=" "
    global cous
    cous=" "
    for row in cursor: 
      summs=float(row[1])
      cous=row[0]
    connection = sqlite3.connect("wallet.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT balance,naam,email,exp,name FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
    for names in cursor:
      inv=float(names[0])
      rdus=names[1]
      rdusd=names[2]
      rdrt=names[3]
    bn=inv-summs
    bn=str(bn)
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set balance = '{}',email='{}' where ID = {}".format(bn,msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    naak=''
    wqp=''
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("SELECT ID,name, quantity,Price,productID from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
    conn.commit()
    for row in cursor:
        v= random.randint (0,999999)
        today = date.today()
        dat = today.strftime("%d/%m/%Y")
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        deepe=row[3]
        sib=row[4]
        wqp=row[2]
        uop=row[1]+' * '+row[2]+ '='+'$'+row[3]+'\n'    
        connection = sqlite3.connect("shop.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT stock FROM COMPANY where productID= '{}'".format(sib) )
        for names in cursor:
          inv=float(names[0])
        bn=inv-float(wqp)
        bn=str(bn)
        conn = sqlite3.connect("shop.db")  
        conn.execute("UPDATE COMPANY set stock = '{}' where productID = '{}'".format(bn,sib))
        conn.commit()
        conn.close()
        naak=naak+uop
    print(naak)
    conn = sqlite3.connect('orders.db')
    conn.execute("INSERT INTO COMPANY (ID,name,oid,address,date,details,status,opt,contact,time,price) \
            VALUES ('{}', '{}', '{}', '{}','{}','{}','{}','{}','{}','{}','{}')".format(str(update.effective_user.id),ms,str(v),rdusd,dat,naak,"pending",rdrt,msg,current_time,summs)) 
    conn.commit()
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("DELETE  from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
    conn.commit()
    keyboard =[[InlineKeyboardButton("Volver a casa", callback_data="spanish")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Su pedido ha sido realizado, recibirá un mensaje de confirmación en breve.",reply_markup=reply_markup) 
    context.bot.send_message(chat_id=5791144147,text="You have a new order, check order section")
    return BUTTON
def je(update,context):
        msg=update.message.text
        conn = sqlite3.connect('users.db')
        cursor = conn.execute("SELECT ID from COMPANY")
        conn.commit()
        for row in cursor:
            aa=row[0]
            try:
                context.bot.send_message(chat_id=aa,text=msg)
            except:
                pass
        keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Message Sent",reply_markup=reply_markup)
        return BUTTON
def addressf(update,context):
    msg=str(update.message.text)
    bnm=update.message.from_user
    ms=bnm.first_name
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("SELECT COUNT(Price),SUM(Price) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
    conn.commit()
    global summf
    summf=" "
    global couf
    couf=" "
    for row in cursor: 
      summf=float(row[1])
      couf=row[0]
    connection = sqlite3.connect("wallet.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT balance,naam,email,exp,name FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
    for names in cursor:
      inv=float(names[0])
      rdus=names[1]
      rdusd=names[2]
      rdrt=names[3]
    bn=inv-summf
    bn=str(bn)
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set balance = '{}',email='{}' where ID = {}".format(bn,msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    naak=''
    wqp=''
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("SELECT ID,name, quantity,Price,productID from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
    conn.commit()
    for row in cursor:
        v= random.randint (0,999999)
        today = date.today()
        dat = today.strftime("%d/%m/%Y")
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        deepe=row[3]
        sib=row[4]
        wqp=row[2]
        uop=row[1]+' * '+row[2]+ '='+'$'+row[3]+'\n'    
        connection = sqlite3.connect("shop.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT stock FROM COMPANY where productID= '{}'".format(sib) )
        for names in cursor:
          inv=float(names[0])
        bn=inv-float(wqp)
        bn=str(bn)
        conn = sqlite3.connect("shop.db")  
        conn.execute("UPDATE COMPANY set stock = '{}' where productID = '{}'".format(bn,sib))
        conn.commit()
        conn.close()
        naak=naak+uop
    print(naak)
    conn = sqlite3.connect('orders.db')
    conn.execute("INSERT INTO COMPANY (ID,name,oid,address,date,details,status,opt,contact,time,price) \
            VALUES ('{}', '{}', '{}', '{}','{}','{}','{}','{}','{}','{}','{}')".format(str(update.effective_user.id),ms,str(v),rdusd,dat,naak,"pending",rdrt,msg,current_time,summf)) 
    conn.commit()
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("DELETE  from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
    conn.commit()
    keyboard =[[InlineKeyboardButton("Rentrer à la maison", callback_data="200")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Votre commande a été passée, vous recevrez un message de confirmation sous peu",reply_markup=reply_markup) 
    context.bot.send_message(chat_id=5791144147,text="You have a new order, check order section")
    return BUTTON
def addb(update,context):
    msg=update.message.text 
    try:
      msg=int(msg)
      conn = sqlite3.connect('wallet.db')
      cursor = conn.execute("SELECT balance,ttr from COMPANY where ID={} ".format(int(shaf))) 
      for row in cursor:           
            nft=float(row[0])+float(msg)
            nftr=float(row[1])+float(msg)
            conn.execute("UPDATE COMPANY set balance = '{}',ttr = '{}' where ID = {}".format(nft,nftr,int(shaf)))
            conn.commit()  
            conn = sqlite3.connect("dep.db")  
            conn.execute("UPDATE COMPANY set status = '{}' where ID = '{}'".format("Accepted",str(shaf)))
            conn.commit()
            conn.close()
            keyboard = [[InlineKeyboardButton("🌎Main Menu", callback_data='100')]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Thank you\n{}$ added to walletID: {}".format(msg,str(shaf)),reply_markup=reply_markup) 
            context.bot.send_message(chat_id=str(shaf),text="Thank you\n{}$ added to your walletID: {}".format(msg,str(shaf)))
    except:
      context.bot.send_message(chat_id=update.effective_user.id,text="Send in digits only without symbol")
      return addb 
def proft(update,context):
    msg=update.message.photo[-1].file_id
    conn = sqlite3.connect('dep.db')
    conn.execute("INSERT INTO COMPANY (ID,amount,status) \
            VALUES ('{}', '{}', '{}')".format(str(update.effective_user.id),msg,'pending')) 
    conn.commit() 
    context.bot.send_message(chat_id=update.effective_user.id,text='Thank you for Your Purchase, we will Review and Confirm as Soon as Possible')
    keyboard =[[InlineKeyboardButton("Check Deposit", callback_data="cdp")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=5791144147,text='Customer submitted payment proof. Check the deposit',reply_markup=reply_markup)
    return BUTTON 
def profts(update,context):
    msg=update.message.photo[-1].file_id

    keyboard =[[InlineKeyboardButton("🔙 atrás", callback_data="spanish")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Gracias. El administrador aceptará/rechazará lo antes posible',reply_markup=reply_markup)

    keyboard =[[InlineKeyboardButton("Accept", callback_data="1abb"),InlineKeyboardButton("Reject", callback_data="1acc")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_photo(chat_id= 5791144147,photo=msg,caption='Payment Proof\n\nWalletID: {}'.format(str(update.effective_user.id)),reply_markup=reply_markup)
    return BUTTON
def proftf(update,context):
    msg=update.message.photo[-1].file_id

    keyboard =[[InlineKeyboardButton("Menu principal", callback_data="french")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Merci. L'administrateur acceptera/rejetera dès que possible",reply_markup=reply_markup)

    keyboard =[[InlineKeyboardButton("Accept", callback_data="1abb"),InlineKeyboardButton("Reject", callback_data="1acc")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_photo(chat_id= 5791144147,photo=msg,caption='Payment Proof\n\nWalletID: {}'.format(str(update.effective_user.id)),reply_markup=reply_markup)
    return BUTTON
def uol(update,context):
    msg=a=update.message.text
    conn = sqlite3.connect('wallet.db')
    conn.execute("UPDATE COMPANY set wname = '{}' where id='{}'".format(msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("Return Home", callback_data="200")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Your Wallet Name has Been Updated',reply_markup=reply_markup)
    return BUTTON
def halp(update,context):
    msg=a=update.message.text
    conn = sqlite3.connect('help.db')
    conn.execute("UPDATE COMPANY set contact = '{}' where id='{}'".format(msg,'123'))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Thank You. Help/Support Updated',reply_markup=reply_markup)
    return BUTTON
def sh(update,context):
    msg=a=update.message.text
    conn = sqlite3.connect('wel.db')
    conn.execute("UPDATE COMPANY set msg = '{}' ".format(msg))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Thank you. Welcome Message Updated',reply_markup=reply_markup)
    return BUTTON
def depo(update,context):
    msg=update.message.text
    global jaun
    jaun=msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Send Wallet Address',reply_markup=reply_markup)
    return DEPOT 
def depot(update,context):
    msg=update.message.text
    conn = sqlite3.connect('deposit.db')
    conn.execute("INSERT INTO COMPANY (name,address) \
            VALUES ('{}', '{}')".format(jaun,msg)) 
    conn.commit() 
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Wallet Added',reply_markup=reply_markup)
    return BUTTON    

def wall(update,context):
    msg=update.message.text
    try:
        ahse=float(msg)
        a=ahse
        create_transaction_params = {
            'amount' : int(a),
            'currency1' : 'USD',
            'currency2' : 'BTC',
            'buyer_email': 'blockpartytrading@gmail.com'
                }
        client = CryptoPayments(API_KEY, API_SECRET, IPN_URL)
        transaction = client.createTransaction(create_transaction_params)
        if transaction['error'] == 'ok':
                am= transaction['amount'] 
                ad=transaction['address']
                url=transaction['checkout_url']
        post = {
            'txid' : transaction['txn_id'],}
        post=post['txid']
        keyboard = [[InlineKeyboardButton("I have Paid", callback_data='999z')],[InlineKeyboardButton("Make a Payment", url=url)],
                    [InlineKeyboardButton("Return to Shop", callback_data='12')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        user=int(update.effective_user.id)
        conn = sqlite3.connect('wallet')
        conn.execute("UPDATE COMPANY set Trx = '{}',exp='{}' where ID = {}".format(post,int(msg),user))
        conn.commit()
        conn.close()
        context.bot.send_message(chat_id=update.effective_user.id,text="Please send {} to {}\n\n""Or click on 'Make a Payment' button\n\n""After making payment wait for 30 minutes to wait for the confirmation of transcation  and then click on 'I have Paid' button".format(am,ad),reply_markup=reply_markup)
        return BUTTON
    except:
        context.bot.send_message(chat_id=update.effective_user.id,text='Invalid price! send in digits only')
def jos(update,context):
    msg=update.message.text
    print(msg)#keyf=[["📝️ Productos","🛒 Carro"],["💳 Cartera","🛍 Mis ordenes"],[" 🙋 FAQ ","❓ Servicio de asistencia"]]
    if msg=="📝️ Products":
        keyf=[]
        connection = sqlite3.connect("cata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT cat FROM COMPANY")
        keyf=[]
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[0])]
            keyf.append(c)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Select Main Category",reply_markup=reply_markup) 
 
        return PDB
    elif msg=="📝️ Productos":
        keyf=[]
        connection = sqlite3.connect("cata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT cat FROM COMPANY")
        keyf=[]
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[0])]
            keyf.append(c)
        b=[InlineKeyboardButton("🌎Menú principal", callback_data='🌎Menú principal')]
        keyf.append(b)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Seleccionar categoría principal",reply_markup=reply_markup) 
 
        return PDBS
    elif msg=="📝️ Des produits":
        keyf=[]
        connection = sqlite3.connect("cata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT cat FROM COMPANY")
        keyf=[]
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[0])]
            keyf.append(c)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Sélectionnez la catégorie principale",reply_markup=reply_markup) 
        return PDBF
    elif msg=="🛒 Cart":
        conn = sqlite3.connect('cart.db')
        cursor = conn.execute("SELECT name,Price from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
        conn.commit()
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('cart.db')
            cursor = conn.execute("SELECT name,Price,quantity from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            conn.commit()
            global x
            x="𝐘𝐨𝐮𝐫 𝐂𝐚𝐫𝐭:\n𝐃𝐞𝐭𝐚𝐢𝐥𝐬\n"
            for row in cursor: 
                m="   "+row[0]+" x "+row[2] + "   =   "+row[1]+"\n"   
                x=x+m 
            cursor = conn.execute("SELECT SUM(Price),SUM(Price) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            global ap
            ap=""
            global nj
            nj=""
            for row in cursor:
                cer="\n𝐓𝐨𝐭𝐚𝐥 ="+str(row[0])+"$"
            keyboard =[[InlineKeyboardButton("✅ Checkout", callback_data="25"),InlineKeyboardButton("❌ Cart", callback_data="29")],[InlineKeyboardButton("Return to Shop", callback_data="200")]]   
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
            context.bot.send_message(chat_id=update.effective_user.id,text=x+cer ,reply_markup=reply_markup)                    
            return BUTTON
        else:
            context.bot.send_message(chat_id=update.effective_user.id,text="Your Cart is Empty, Find Something Great Below!")
            keyf=[]
            connection = sqlite3.connect("cata.db")  
            cursor = connection.cursor()  
            cursor.execute("SELECT cat FROM COMPANY")
            keyf=[]
            for name in cursor:
                c=[InlineKeyboardButton(name[0], callback_data=name[0])]
                keyf.append(c)
            reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
            context.bot.send_message(chat_id=update.effective_user.id,text="Select Main Category",reply_markup=reply_markup) 
    
            return PDB

    elif msg=="🛒 Carro":
        conn = sqlite3.connect('cart.db')
        cursor = conn.execute("SELECT name,Price from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
        conn.commit()
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('cart.db')
            cursor = conn.execute("SELECT name,Price,quantity from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            conn.commit()
            global xs
            xs="Aquí está tu carrito:\nDetalles\n"
            for row in cursor: 
                m="   "+row[0]+" x "+row[2] + "   =   "+row[1]+"\n"   
                xs=xs+m
            cursor = conn.execute("SELECT SUM(Price),SUM(Price) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            global aps
            aps=""
            global njs
            njs=""
            for row in cursor:
                cer="\n𝐓𝐨𝐭𝐚𝐥 ="+str(row[0])+"$"
            keyboard =[[InlineKeyboardButton("✅ Verificar", callback_data="25s"),InlineKeyboardButton("❌ Carro", callback_data="29s")],[InlineKeyboardButton("🔙 Cancelar", callback_data="spanish")]]   
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
            context.bot.send_message(chat_id=update.effective_user.id,text=xs+cer ,reply_markup=reply_markup)                    
            return BUTTON
        else:
            context.bot.send_message(chat_id=update.effective_user.id,text="El carrito esta vacío")                       
            return BUTTON
    elif msg=="🛒 Chariot":
        conn = sqlite3.connect('cart.db')
        cursor = conn.execute("SELECT name,Price from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
        conn.commit()
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('cart.db')
            cursor = conn.execute("SELECT name,Price,quantity from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            conn.commit()
            global xf
            xf="𝐕𝐨𝐢𝐜𝐢 𝐯𝐨𝐭𝐫𝐞 𝐩𝐚𝐧𝐢𝐞𝐫:\n𝐃é𝐭𝐚𝐢𝐥𝐬\n"
            for row in cursor: 
                m="   "+row[0]+" x "+row[2] + "   =   "+row[1]+"\n"     
                xf=xf+m
            cursor = conn.execute("SELECT SUM(Price),SUM(Price) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            global apf
            apf=""
            global njf
            njf=""
            for row in cursor:
                cer="\n𝐓𝐨𝐭𝐚𝐥 ="+str(row[0])+"$"
            keyboard =[[InlineKeyboardButton("✅ Vérifier", callback_data="25f"),InlineKeyboardButton("❌ Chariot", callback_data="29f")],[InlineKeyboardButton("🔙 Menu principal", callback_data="french")]]   
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
            context.bot.send_message(chat_id=update.effective_user.id,text=xf+cer ,reply_markup=reply_markup)                    
            return BUTTON

        else:
            context.bot.send_message(chat_id=update.effective_user.id,text="Le panier est vide")                       
            return BUTTON
    elif msg=="💳 Wallet":
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT balance,name,ID,wname from COMPANY where ID= {}   ".format(int(update.effective_user.id))) 
        conn.commit()
        for row in cursor:  
            qoop=row[0]   
        if float(qoop)==0:
            toop="𝐖𝐚𝐥𝐥𝐞𝐭 𝐈𝐃:  {}\n\n𝐖𝐚𝐥𝐥𝐞𝐭 𝐍𝐚𝐦𝐞:  {}\n𝐁𝐚𝐥𝐚𝐧𝐜𝐞: {}$".format(row[2],row[3],qoop)
            keyboard = [[InlineKeyboardButton("➕ Add Funds", callback_data='1bal'),InlineKeyboardButton("💸 Withdraw Funds", callback_data='1with')],[InlineKeyboardButton("Name Your Wallet", callback_data='90p')]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            context.bot.send_message(chat_id=update.effective_user.id,text=toop+"\n\nYou need credits in your wallet to make purchases.\nClick ➕ 𝐀𝐝𝐝 𝐅𝐮𝐧𝐝𝐬 to finance your wallet",reply_markup=reply_markup)                       
            return BUTTON
        else:
            toop="𝐖𝐚𝐥𝐥𝐞𝐭 𝐈𝐃:  {}\n\n𝐖𝐚𝐥𝐥𝐞𝐭 𝐍𝐚𝐦𝐞:  {}\n𝐁𝐚𝐥𝐚𝐧𝐜𝐞: {}$".format(row[2],row[3],qoop)
            keyboard = [[InlineKeyboardButton("➕ Add Funds", callback_data='1bal'),InlineKeyboardButton("💸 Withdraw Funds", callback_data='1with')],[InlineKeyboardButton("Name Your Wallet", callback_data='90p')]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            context.bot.send_message(chat_id=update.effective_user.id,text=toop,reply_markup=reply_markup)                       
            return BUTTON
    elif msg=="💳 Cartera":
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT balance,name,ID,wname from COMPANY where ID= {}   ".format(int(update.effective_user.id))) 
        conn.commit()
        for row in cursor:  
            qoop=row[0]   
        toop="𝐈𝐃:  {}\n\n𝐍𝐨𝐦𝐛𝐫𝐞 𝐝𝐞 𝐥𝐚 𝐛𝐢𝐥𝐥𝐞𝐭𝐞𝐫𝐚:  {}\n𝐁𝐚𝐥𝐚𝐧𝐜𝐞: {}$".format(row[2],row[3],qoop)
        keyboard = [[InlineKeyboardButton("➕ Añadir fondos", callback_data='1bals'),InlineKeyboardButton("💸 Retirar", callback_data='1withs')]
           ,[InlineKeyboardButton("Crear nombre de billetera", callback_data='90ps')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.bot.send_message(chat_id=update.effective_user.id,text=toop+"\n\nNecesita créditos en su billetera para realizar compras.\nHaga clic en ➕ Añadir fondos para financiar su billetera",reply_markup=reply_markup)                       
        return BUTTON
    elif msg=="💳 Porte monnaie":
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT balance,name,ID from COMPANY where ID= {}   ".format(int(update.effective_user.id))) 
        conn.commit()
        for row in cursor:  
            qoop=row[0]   
        toop="𝐈𝐃 𝐝𝐞 𝐩𝐨𝐫𝐭𝐞𝐟𝐞𝐮𝐢𝐥𝐥𝐞:  {}\n\n𝐍𝐨𝐦:  {}\n𝐒𝐨𝐥𝐝𝐞: {}$".format(row[2],row[1],qoop)
        keyboard = [[InlineKeyboardButton("➕ Solde", callback_data='1balf'),InlineKeyboardButton("💸 Retrait", callback_data='1withf')]
                    ,[InlineKeyboardButton("Créer le nom du portefeuille", callback_data='90pf'),InlineKeyboardButton("🔙 Menu principal", callback_data='french')]]
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.bot.send_message(chat_id=update.effective_user.id,text=toop+"\n\nCliquez sur ➕ Solde pour ajouter un montant dans votre portefeuille",reply_markup=reply_markup)                       
        return BUTTON
    elif msg=="❓ Help/Support":
        connection = sqlite3.connect("help.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT contact FROM COMPANY")
        keyf=[]
        for row in cursor:
          aaa=row[0]
          context.bot.send_message(chat_id=update.effective_user.id,text=aaa)
          return BUTTON
    elif msg=="❓ Support d'aide":
        connection = sqlite3.connect("help.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT contact FROM COMPANY")
        keyf=[]
        for row in cursor:
          aaa=row[0]
          context.bot.send_message(chat_id=update.effective_user.id,text=aaa)
          return BUTTON
    elif msg=="❓ Servicio de asistencia":
        connection = sqlite3.connect("help.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT contact FROM COMPANY")
        keyf=[]
        for row in cursor:
          aaa=row[0]
          context.bot.send_message(chat_id=update.effective_user.id,text=aaa)
          return BUTTON
    elif msg=="🙋 FAQ":
        connection = sqlite3.connect("faqs.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT ques,ans FROM COMPANY")
        keyf=[]
        jkl=' '
        for row in cursor:
          m="{}\n{}\n\n".format(row[0],row[1])
          jkl=m+jkl
        context.bot.send_message(chat_id=update.effective_user.id,text=jkl)
        return BUTTON
    elif msg=="🙋 FAQs":
        connection = sqlite3.connect("faqs.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT ques,ans FROM COMPANY")
        keyf=[]
        jkl=' '
        for row in cursor:
          m="{}\n{}\n\n".format(row[0],row[1])
          jkl=m+jkl
        context.bot.send_message(chat_id=update.effective_user.id,text=jkl)
        return BUTTON
    elif msg==" 🙋 FAQ ":
        connection = sqlite3.connect("faqs.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT ques,ans FROM COMPANY")
        keyf=[]
        jkl=' '
        for row in cursor:
          m="{}\n{}\n\n".format(row[0],row[1])
          jkl=m+jkl
        context.bot.send_message(chat_id=update.effective_user.id,text=jkl)
        return BUTTON

    elif msg=="🛍 My Orders":
        connection = sqlite3.connect("orders.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT oid FROM COMPANY where ID= '{}' ".format(str(update.effective_user.id)))
        jobs = cursor.fetchall()
        if len(jobs) !=0:
                    conn = sqlite3.connect('orders.db')
                    cursor = conn.execute("SELECT oid,status from COMPANY where ID= '{}' ".format(str(update.effective_user.id)))
                    conn.commit()
                    keyf=[]
                    for row in cursor:
                        if row[1]=="pending":
                            k="⏳"
                        elif row[1]=="Accept":
                            k="✅"
                        elif row[1]=="Delivered":
                            k="🚚"
                        elif row[1]=="Rejected":
                            k="❌"
                        c=[InlineKeyboardButton(k+row[0], callback_data=row[0])]
                        keyf.append(c)
                    reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
                    context.bot.send_message(chat_id=update.effective_user.id,text="Select Order to check detail",reply_markup=reply_markup) 
                    return DDFF
        else:
          context.bot.send_message(chat_id=update.effective_user.id,text="You Have Not Made Any Orders")
          return BUTTON
    elif msg=="🛍 Mis ordenes":
        connection = sqlite3.connect("orders.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT oid FROM COMPANY where ID= '{}' ".format(str(update.effective_user.id)))
        jobs = cursor.fetchall()
        if len(jobs) !=0:
                    conn = sqlite3.connect('orders.db')
                    cursor = conn.execute("SELECT oid,status,pnam from COMPANY where ID= '{}' ".format(str(update.effective_user.id)))
                    conn.commit()
                    keyf=[]
                    for row in cursor:
                        if row[1]=="pending":
                            k="⏳"
                        elif row[1]=="Accept":
                            k="✅"
                        elif row[1]=="Delivered":
                            k="🚚"
                        elif row[1]=="Rejected":
                            k="❌"
                        c=[InlineKeyboardButton(k+row[0], callback_data=row[0])]
                        keyf.append(c)
                    b=[InlineKeyboardButton("🌎Menú principal", callback_data='🌎Menú principal')]
                    keyf.append(b)
                    reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
                    context.bot.send_message(chat_id=update.effective_user.id,text="Seleccione Orden para verificar el detalle",reply_markup=reply_markup) 
                    return DDFFS
        else:
          context.bot.send_message(chat_id=update.effective_user.id,text="No has hecho ningún pedido")
          return BUTTON

    elif msg=="🛍 Mes commandes":
      connection = sqlite3.connect("orders.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT oid FROM COMPANY where ID= '{}' ".format(str(update.effective_user.id)))
      jobs = cursor.fetchall()
      if len(jobs) !=0:
                  conn = sqlite3.connect('orders.db')
                  cursor = conn.execute("SELECT oid,status,pnam from COMPANY where ID= '{}' ".format(str(update.effective_user.id)))
                  conn.commit()
                  keyf=[]
                  for row in cursor:
                      if row[1]=="pending":
                          k="⏳"
                      elif row[1]=="Accept":
                          k="✅"
                      elif row[1]=="Delivered":
                          k="🚚"
                      elif row[1]=="Rejected":
                          k="❌"
                      c=[InlineKeyboardButton(k+row[0], callback_data=row[0])]
                      keyf.append(c)
                  b=[InlineKeyboardButton("🌎Menu principal", callback_data='🌎Menu principal')]
                  keyf.append(b)
                  reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
                  context.bot.send_message(chat_id=update.effective_user.id,text="Sélectionnez Commander pour vérifier les détails",reply_markup=reply_markup) 
                  return DDFFF
      else:
        context.bot.send_message(chat_id=update.effective_user.id,text="Vous n'avez passé aucune commande")
        return BUTTON
def jop(update,context):
    query = update.callback_query
    msg=query.data
    #3conn = sqlite3.connect('cart.db')
    #cursor = conn.execute("SELECT quantity from COMPANY where ProductID='{}'".format(msg)) 
    #for row in cursor: 
    # hh=row[0]
    #conn = sqlite3.connect('shop.db')
    #cursor = conn.execute("SELECT max from COMPANY where ProductID='{} '".format(msg))
    #for row in cursor:           
    #      nft=int(row[0])+int(hh)
    #      conn.execute("UPDATE COMPANY set max = '{}' where ProductID = '{}'".format(nft,msg))
     #     conn.commit() 
    connection = sqlite3.connect("cart.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT quantity FROM COMPANY where productID= '{}'".format(msg) )
    for names in cursor:
      ted=float(names[0])
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("DELETE  from COMPANY where ProductID='{}' AND ID='{}'".format(msg,update.effective_user.id))
    conn.commit()
    conn.close()
    connection = sqlite3.connect("shop.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT stock FROM COMPANY where productID= '{}'".format(msg) )
    for names in cursor:
      inv=float(names[0])
    bn=inv+float(ted)
    bn=str(bn)
    conn = sqlite3.connect("shop.db")  
    conn.execute("UPDATE COMPANY set stock = '{}' where productID = '{}'".format(bn,msg))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("🔙 Main Menu", callback_data="200")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Product deleted from cart",reply_markup=reply_markup) 
    return BUTTON
def jops(update,context):
    query = update.callback_query
    msg=query.data
    connection = sqlite3.connect("cart.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT quantity FROM COMPANY where productID= '{}'".format(msg) )
    for names in cursor:
      ted=float(names[0])
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("DELETE  from COMPANY where ProductID='{}' AND ID='{}'".format(msg,update.effective_user.id))
    conn.commit()
    conn.close()
    connection = sqlite3.connect("shop.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT stock FROM COMPANY where productID= '{}'".format(msg) )
    for names in cursor:
      inv=float(names[0])
    bn=inv+float(ted)
    bn=str(bn)
    conn = sqlite3.connect("shop.db")  
    conn.execute("UPDATE COMPANY set stock = '{}' where productID = '{}'".format(bn,msg))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("🔙 Main Menu", callback_data="spanish")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Producto eliminado del carrito",reply_markup=reply_markup) 
    return BUTTON
def jopf(update,context):
    query = update.callback_query
    msg=query.data
    connection = sqlite3.connect("cart.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT quantity FROM COMPANY where productID= '{}'".format(msg) )
    for names in cursor:
      ted=float(names[0])
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("DELETE  from COMPANY where ProductID='{}' AND ID='{}'".format(msg,update.effective_user.id))
    conn.commit()
    conn.close()
    connection = sqlite3.connect("shop.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT stock FROM COMPANY where productID= '{}'".format(msg) )
    for names in cursor:
      inv=float(names[0])
    bn=inv+float(ted)
    bn=str(bn)
    conn = sqlite3.connect("shop.db")  
    conn.execute("UPDATE COMPANY set stock = '{}' where productID = '{}'".format(bn,msg))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("🔙 Main Menu", callback_data="french")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Produit supprimé du panier",reply_markup=reply_markup) 
    return BUTTON
def ddff(update,context):
    query = update.callback_query
    msg=query.data
    print(msg)
    conn = sqlite3.connect('orders.db')
    cursor = conn.execute("SELECT ID,price,details,name,oid,date,status from COMPANY where oid ='{}'".format(msg))
    conn.commit()
    for row in cursor:
        g="𝐎𝐫𝐝𝐞𝐫 𝐝𝐞𝐭𝐚𝐢𝐥𝐬: \n\n"+row[2]+"𝐎𝐫𝐝𝐞𝐫 𝐧𝐮𝐦𝐛𝐞𝐫: "+row[4]+"\n𝐎𝐫𝐝𝐞𝐫 𝐃𝐚𝐭𝐞: "+row[5]+"\n𝐏𝐚𝐲𝐦𝐞𝐧𝐭: "+row[1]+"$"+"\n"
        keyboard =[[InlineKeyboardButton("🌎Back to Orders", callback_data="2190")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text=g,reply_markup=reply_markup)
        return BUTTON
def ddffs(update,context):
    query = update.callback_query
    msg=query.data
    print(msg)
    conn = sqlite3.connect('orders.db')
    cursor = conn.execute("SELECT ID,price,name,name,oid,date,status,pnam from COMPANY where oid ='{}'".format(msg))
    conn.commit()
    for row in cursor:
        g="𝐃𝐞𝐭𝐚𝐥𝐥𝐞𝐬 𝐝𝐞𝐥 𝐩𝐞𝐝𝐢𝐝𝐨: \n\n"+"𝐍ú𝐦𝐞𝐫𝐨 𝐝𝐞 𝐨𝐫𝐝𝐞: "+row[4]+"\n𝐏𝐫𝐨𝐝𝐮𝐜𝐭𝐨: "+row[7]+"\n𝐅𝐞𝐜𝐡𝐚 𝐝𝐞 𝐨𝐫𝐝𝐞𝐧: "+row[5]+"\n𝐏𝐚𝐲𝐦𝐞𝐧𝐭: "+row[1]+"$"+"\n"
        keyboard =[[InlineKeyboardButton("🌎Menú principal", callback_data="spanish")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text=g,reply_markup=reply_markup)
        return BUTTON
def ddfff(update,context):
    query = update.callback_query
    msg=query.data
    print(msg)
    if msg=="🌎Menu principal":
      keyf=[["📝️ Des produits","🛒 Chariot"],["💳 Porte monnaie","🛍 Mes commandes"],["🙋 FAQs","❓ Support d'aide"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Bienvenue sur le tableau de bord utilisateur",reply_markup=reply_markup)
    else:
      conn = sqlite3.connect('orders.db')
      cursor = conn.execute("SELECT ID,price,name,name,oid,date,status,pnam from COMPANY where oid ='{}'".format(msg))
      conn.commit()
      for row in cursor:
          g="Détails de la commande \n\n"+"Numéro de commande: "+row[4]+"\nProduit: "+row[7]+"\nDate de commande: "+row[5]+"\nPaiement: "+row[1]+"$"+"\n"
          keyboard =[[InlineKeyboardButton("🌎Menu principal", callback_data="french")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
          context.bot.send_message(chat_id=update.effective_user.id,text=g,reply_markup=reply_markup)
          return BUTTON
def pdb(update,context):
    query = update.callback_query
    msg=query.data
    print(msg)    
    c=update.callback_query.message.message_id
    context.bot.delete_message(chat_id=update.effective_user.id,
                        message_id=c)
    if msg=="🌎Main Menu":
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Products","🛒 Cart"],["💳 Wallet","🛍 My Orders"],["🙋 FAQ","❓ Help/Support"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    elif msg=="200":
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Products","🛒 Cart"],["💳 Wallet","🛍 My Orders"],["🙋 FAQ","❓ Help/Support"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    else:
        keyf=[]
        conn = sqlite3.connect('wallet.db')
        cursor=conn.execute("UPDATE COMPANY set ct = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
        conn.commit()
        connection = sqlite3.connect("subcata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT subcat FROM COMPANY where cat='{}'".format(str(msg)))
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[0])]
            keyf.append(c)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Select Sub-Category",reply_markup=reply_markup)           
        return PDVM 

def pdbs(update,context):
    query = update.callback_query
    msg=query.data
    print(msg)    
    c=update.callback_query.message.message_id
    context.bot.delete_message(chat_id=update.effective_user.id,
                        message_id=c)
    if msg=="🌎Menú principal":
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Productos","🛒 Carro"],["💳 Cartera","🛍 Mis ordenes"],[" 🙋 FAQ ","❓ Servicio de asistencia"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    elif msg=="spanish":
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Productos","🛒 Carro"],["💳 Cartera","🛍 Mis ordenes"],[" 🙋 FAQ ","❓ Servicio de asistencia"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    else:
        keyf=[]
        conn = sqlite3.connect('wallet.db')
        cursor=conn.execute("UPDATE COMPANY set ct = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
        conn.commit()
        connection = sqlite3.connect("subcata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT subcat FROM COMPANY where cat='{}'".format(str(msg)))
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[0])]
            keyf.append(c)
        b=[InlineKeyboardButton("🌎Menú principal", callback_data='🌎Menú principal')]
        keyf.append(b)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Seleccionar subcategoría",reply_markup=reply_markup)           
        return PDVMS

def pdbf(update,context):
    query = update.callback_query
    msg=query.data
    print(msg)    
    c=update.callback_query.message.message_id
    context.bot.delete_message(chat_id=update.effective_user.id,
                        message_id=c)
    if msg=="🌎Menu principal":
      keyf=[["📝️ Des produits","🛒 Chariot"],["💳 Porte monnaie","🛍 Mes commandes"],["🙋 FAQs","❓ Support d'aide"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Bienvenue sur le tableau de bord utilisateur",reply_markup=reply_markup)
    elif msg=="200":
      keyf=[["📝️ Des produits","🛒 Chariot"],["💳 Porte monnaie","🛍 Mes commandes"],["🙋 FAQs","❓ Support d'aide"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Bienvenue sur le tableau de bord utilisateur",reply_markup=reply_markup)
    else:
        keyf=[]
        conn = sqlite3.connect('wallet.db')
        cursor=conn.execute("UPDATE COMPANY set ct = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
        conn.commit()
        connection = sqlite3.connect("subcata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT subcat FROM COMPANY where cat='{}'".format(str(msg)))
        for name in cursor:
            c=[InlineKeyboardButton(name[0], callback_data=name[0])]
            keyf.append(c)
        b=[InlineKeyboardButton("🌎Menu principal", callback_data='🌎Menu principal')]
        keyf.append(b)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Sélectionnez la sous-catégorie",reply_markup=reply_markup)           
        return PDVMF
def pdvm(update,context):
    query = update.callback_query
    msg=query.data
    if msg=="🌎Main Menu":
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Products","🛒 Cart"],["💳 Wallet","🛍 My Orders"],["🙋 FAQ","❓ Help/Support"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    elif msg=="200":
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Products","🛒 Cart"],["💳 Wallet","🛍 My Orders"],["🙋 FAQ","❓ Help/Support"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    else:
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT ct from COMPANY where ID= {}   ".format(int(update.effective_user.id))) 
        conn.commit()
        for row in cursor:  
            qoop=row[0]
            cursor=conn.execute("UPDATE COMPANY set sct = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
            conn.commit()
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("SELECT name from COMPANY where category= '{}'   AND subcat='{}' ".format(qoop,msg))                              
        keyf=[]
        for name in cursor:
                c=[InlineKeyboardButton(name[0], callback_data=name[0])]
                keyf.append(c)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Select a Product Below",reply_markup=reply_markup)            
        return SHAKA

def pdvms(update,context):
    query = update.callback_query
    msg=query.data
    if msg=="🌎Menú principal":
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Productos","🛒 Carro"],["💳 Cartera","🛍 Mis ordenes"],[" 🙋 FAQ ","❓ Servicio de asistencia"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    elif msg=="spanish":
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Productos","🛒 Carro"],["💳 Cartera","🛍 Mis ordenes"],[" 🙋 FAQ ","❓ Servicio de asistencia"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    else:
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT ct from COMPANY where ID= {}   ".format(int(update.effective_user.id))) 
        conn.commit()
        for row in cursor:  
            qoop=row[0]
            cursor=conn.execute("UPDATE COMPANY set sct = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
            conn.commit()
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("SELECT name from COMPANY where category= '{}'   AND subcat='{}' ".format(qoop,msg))                              
        keyf=[]
        for name in cursor:
                c=[InlineKeyboardButton(name[0], callback_data=name[0])]
                keyf.append(c)
        b=[InlineKeyboardButton("🌎Menú principal", callback_data='🌎Menú principal')]
        keyf.append(b)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Seleccione el nombre del producto para ver",reply_markup=reply_markup)            
        return SHAKAS
def pdvmf(update,context):
    query = update.callback_query
    msg=query.data
    if msg=="🌎Menu principal":
      keyf=[["📝️ Des produits","🛒 Chariot"],["💳 Porte monnaie","🛍 Mes commandes"],["🙋 FAQs","❓ Support d'aide"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Bienvenue sur le tableau de bord utilisateur",reply_markup=reply_markup)
    elif msg=="200":
      keyf=[["📝️ Des produits","🛒 Chariot"],["💳 Porte monnaie","🛍 Mes commandes"],["🙋 FAQs","❓ Support d'aide"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Bienvenue sur le tableau de bord utilisateur",reply_markup=reply_markup)
    else:
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT ct from COMPANY where ID= {}   ".format(int(update.effective_user.id))) 
        conn.commit()
        for row in cursor:  
            qoop=row[0]
            cursor=conn.execute("UPDATE COMPANY set sct = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
            conn.commit()
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("SELECT name from COMPANY where category= '{}'   AND subcat='{}' ".format(qoop,msg))                              
        keyf=[]
        for name in cursor:
                c=[InlineKeyboardButton(name[0], callback_data=name[0])]
                keyf.append(c)
        b=[InlineKeyboardButton("🌎Menu principal", callback_data='🌎Menu principal')]
        keyf.append(b)
        reply_markup = InlineKeyboardMarkup(keyf,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Sélectionnez le nom du produit à afficher",reply_markup=reply_markup)            
        return SHAKAF
def shaka(update,context):
    query = update.callback_query
    msg=query.data
    if msg=="🌎Main Menu":
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Products","🛒 Cart"],["💳 Wallet","🛍 My Orders"],["🙋 FAQ","❓ Help/Support"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    elif msg=="200":
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Products","🛒 Cart"],["💳 Wallet","🛍 My Orders"],["🙋 FAQ","❓ Help/Support"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    else:
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT ct,sct from COMPANY where ID={} ".format(str(update.effective_user.id)))
        conn.commit()
        for names in cursor:
            hjj=names[0] 
            rpl=names[1]
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("SELECT name,description,price,productID,photoid from COMPANY where category= '{}'  AND subcat='{}' AND name='{}' ".format(hjj,rpl,msg))                              
        keyf=[]
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('shop.db')
            cursor = conn.execute("SELECT name,description,price,productID,photoid,stock from COMPANY where category= '{}'  AND subcat='{}' AND name='{}'".format(hjj,rpl,msg)) 
            for row in cursor:       
                tyu=row[5]
                if tyu=='0':
                  tyu="Out of Stock"
                else:
                    tyu=str(tyu)         
                m="𝐍𝐚𝐦𝐞:  "+row[0]+"\n𝐈𝐃:  "+row[3]+"\n𝐏𝐫𝐢𝐜𝐞:   "+row[2]+"$"+"\n𝐃𝐞𝐬𝐜𝐫𝐢𝐩𝐭𝐢𝐨𝐧:  "+row[1]+"\n𝐒𝐭𝐨𝐜𝐤 𝐚𝐯𝐚𝐢𝐥𝐢𝐛𝐢𝐥𝐭𝐲:  "+tyu+"\n𝐐𝐮𝐚𝐧𝐭𝐢𝐭𝐲:  1"
                keyboard =[[InlineKeyboardButton("➕ Quantity", callback_data="160"),InlineKeyboardButton("➖ Quantity", callback_data="161")],[InlineKeyboardButton("➕ Add to cart", callback_data="16"),InlineKeyboardButton("Back to Products", callback_data="200")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
                context.bot.send_photo(chat_id=update.effective_user.id,photo=row[4],caption=m,reply_markup=reply_markup)
                return BUTTON   
        else:
                    context.bot.send_message(chat_id=update.effective_user.id,text='This Category has no Items')
                    return BUTTON 
def shakaf(update,context):
    query = update.callback_query
    msg=query.data
    if msg=="🌎Menu principal":
      keyf=[["📝️ Des produits","🛒 Chariot"],["💳 Porte monnaie","🛍 Mes commandes"],["🙋 FAQs","❓ Support d'aide"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Bienvenue sur le tableau de bord utilisateur",reply_markup=reply_markup)
    elif msg=="200":
      keyf=[["📝️ Des produits","🛒 Chariot"],["💳 Porte monnaie","🛍 Mes commandes"],["🙋 FAQs","❓ Support d'aide"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Bienvenue sur le tableau de bord utilisateur",reply_markup=reply_markup)
    else:
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT ct,sct from COMPANY where ID={} ".format(str(update.effective_user.id)))
        conn.commit()
        for names in cursor:
            hjj=names[0] 
            rpl=names[1]
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("SELECT name,description,price,productID,photoid from COMPANY where category= '{}'  AND subcat='{}' AND name='{}' ".format(hjj,rpl,msg))                              
        keyf=[]
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('shop.db')
            cursor = conn.execute("SELECT name,description,price,productID,photoid,stock from COMPANY where category= '{}'  AND subcat='{}' AND name='{}'".format(hjj,rpl,msg)) 
            
            for row in cursor:       
                tyu=row[5]
                if tyu=='0':
                  tyu="En rupture de stock"
                else:
                    tyu=str(tyu)                              
                m="𝐍𝐨𝐦:  "+row[0]+"\n𝐈𝐃:  "+row[3]+"\n𝐏𝐫𝐢𝐱:   "+row[2]+"$"+"\n𝐋𝐚 𝐝𝐞𝐬𝐜𝐫𝐢𝐩𝐭𝐢𝐨𝐧:  "+row[1] +"\n𝐃𝐢𝐬𝐩𝐨𝐧𝐢𝐛𝐢𝐥𝐢𝐭é 𝐝𝐞𝐬 𝐬𝐭𝐨𝐜𝐤𝐬:  "+tyu+"\n𝐐𝐮𝐚𝐧𝐭𝐢𝐭é:  1"
                keyboard =[[InlineKeyboardButton("➕ Ajouter au panier", callback_data="16f"),InlineKeyboardButton("🔙Menu principal", callback_data="french")],[InlineKeyboardButton("➕ Quantité", callback_data="160f"),InlineKeyboardButton("➖ Quantité", callback_data="161f")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
                context.bot.send_photo(chat_id=update.effective_user.id,photo=row[4],caption=m,reply_markup=reply_markup)
                return BUTTON   
        else:
                    keyboard =[[InlineKeyboardButton("🌎Menu principal", callback_data="french")]]
                    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
                    context.bot.send_message(chat_id=update.effective_user.id,text='This Category has no Items',reply_markup=reply_markup)
                    return BUTTON 
def shakas(update,context):
    query = update.callback_query
    msg=query.data
    if msg=="🌎Menú principal":
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Productos","🛒 Carro"],["💳 Cartera","🛍 Mis ordenes"],[" 🙋 FAQ ","❓ Servicio de asistencia"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    elif msg=="Menú principal":
      connection = sqlite3.connect("wel.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT msg FROM COMPANY")
      for name in cursor:
        welcome=str(name[0])
      keyf=[["📝️ Productos","🛒 Carro"],["💳 Cartera","🛍 Mis ordenes"],[" 🙋 FAQ ","❓ Servicio de asistencia"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=welcome,reply_markup=reply_markup)
    else:
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT ct,sct from COMPANY where ID={} ".format(str(update.effective_user.id)))
        conn.commit()
        for names in cursor:
            hjj=names[0] 
            rpl=names[1]
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("SELECT name,description,price,productID,photoid from COMPANY where category= '{}'  AND subcat='{}' AND name='{}' ".format(hjj,rpl,msg))                              
        keyf=[]
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('shop.db')
            cursor = conn.execute("SELECT name,description,price,productID,photoid,stock from COMPANY where category= '{}'  AND subcat='{}' AND name='{}'".format(hjj,rpl,msg)) 
            for row in cursor:       
                tyu=row[5]
                if tyu=='0':
                  tyu="Agotado"
                else:
                    tyu=str(tyu)         
                m="𝐍𝐨𝐦𝐛𝐫𝐞:  "+row[0]+"\n𝐈𝐃:  "+row[3]+"\n𝐏𝐫𝐞𝐜𝐢𝐨:   "+row[2]+"$"+"\n𝐃𝐞𝐬𝐜𝐫𝐢𝐩𝐜𝐢ó𝐧:  "+row[1]+"\n𝐃𝐢𝐬𝐩𝐨𝐧𝐢𝐛𝐢𝐥𝐢𝐝𝐚𝐝 𝐝𝐞 𝐞𝐱𝐢𝐬𝐭𝐞𝐧𝐜𝐢𝐚𝐬:  "+tyu+"\n𝐂𝐚𝐧𝐭𝐢𝐝𝐚𝐝:  1"
                keyboard =[[InlineKeyboardButton("➕ Añadir a la cesta", callback_data="16s"),InlineKeyboardButton("🔙 Menú principal", callback_data="spanish")],[InlineKeyboardButton("➕ Cantidad", callback_data="160s"),InlineKeyboardButton("➖ Cantidad", callback_data="161s")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
                context.bot.send_photo(chat_id=update.effective_user.id,photo=row[4],caption=m,reply_markup=reply_markup)
                return BUTTON   
        else:
                    keyboard =[[InlineKeyboardButton("🌎Main Menu", callback_data="200")]]
                    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
                    context.bot.send_message(chat_id=update.effective_user.id,text='This Category has no Items',reply_markup=reply_markup)
                    return BUTTON 
def asub(update,context):
    query = update.callback_query
    msg=query.data
    print(msg)
    global rtop
    rtop=msg 
    if msg=="🌎Main Menu":
      keyboard =[[InlineKeyboardButton("Products", callback_data="pak")],  
                  [InlineKeyboardButton("🛒 Orders", callback_data="99"),InlineKeyboardButton("🔊 Announcement", callback_data="916")],
                  [InlineKeyboardButton("Wallet", callback_data="wallet")]]

      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to the Admin Dashboard" ,reply_markup=reply_markup)
      return BUTTON 
    else:
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text="Send sub category",reply_markup=reply_markup)
      return ASUBB
def dfaq(update,context):
    query = update.callback_query
    msg=query.data
    if msg=='100':
      keyboard =[[InlineKeyboardButton("Products", callback_data="pak") ,InlineKeyboardButton("Wallet", callback_data="wallet")],  
                  [InlineKeyboardButton("Communications", callback_data="comm")]]

      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to the Admin Dashboard" ,reply_markup=reply_markup)
      return BUTTON 
    else:
      conn = sqlite3.connect('faqs.db')
      cursor = conn.execute("DELETE from COMPANY where ques='{}'".format(msg))
      conn.commit()
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text="FAQ Deleted",reply_markup=reply_markup)
      return BUTTON
def ddepo(update,context):
    query = update.callback_query
    msg=query.data
    if msg=='🌎Main Menu':
      keyboard =[[InlineKeyboardButton("Products", callback_data="pak") ,InlineKeyboardButton("Wallet", callback_data="wallet")],  
                  [InlineKeyboardButton("Communications", callback_data="comm")]]

      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to the Admin Dashboard" ,reply_markup=reply_markup)
      return BUTTON
    else:
      conn = sqlite3.connect('deposit.db')
      cursor = conn.execute("DELETE from COMPANY where address='{}'".format(msg))
      conn.commit()
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text="Address Deleted",reply_markup=reply_markup)
      return BUTTON
def asubb(update,context):
    a=update.message.text
    conn = sqlite3.connect('subcata.db')
    conn.execute("INSERT INTO COMPANY (cat,subcat) \
            VALUES ('{}','{}')".format(rtop,a)) 
    conn.commit()
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Sub-Category Added Successfully",reply_markup=reply_markup)
    return BUTTON
def bel(update,context):
    query = update.callback_query
    msg=query.data
    print(msg)
    if msg=='🌎Main Menu':
      keyboard =[[InlineKeyboardButton("Products", callback_data="pak") ,InlineKeyboardButton("Wallet", callback_data="wallet")],  
                  [InlineKeyboardButton("🛒 Orders", callback_data="99"),InlineKeyboardButton("🔊 Announcement", callback_data="916")]]

      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to the Admin Dashboard" ,reply_markup=reply_markup)
      return BUTTON
    else:
      conn = sqlite3.connect('cata.db')
      cursor = conn.execute("DELETE  from COMPANY where cat='{}'".format(msg))
      conn.commit()
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text="Category Deleted",reply_markup=reply_markup)
      return BUTTON
def bela(update,context):
    query = update.callback_query
    msg=query.data
    if msg=='🌎Main Menu':
      keyboard =[[InlineKeyboardButton("Products", callback_data="pak") ,InlineKeyboardButton("Wallet", callback_data="wallet")],  
                  [InlineKeyboardButton("🛒 Orders", callback_data="99"),InlineKeyboardButton("🔊 Announcement", callback_data="916")]]

      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to the Admin Dashboard" ,reply_markup=reply_markup)
      return BUTTON
    else:
      conn = sqlite3.connect('subcata.db')
      cursor = conn.execute("DELETE  from COMPANY where subcat='{}'".format(msg))
      conn.commit()
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text="Category Deleted",reply_markup=reply_markup)
      return BUTTON
def ch(update,context):
    a=update.message.text
    conn = sqlite3.connect('cata.db')
    conn.execute("INSERT INTO COMPANY (cat) \
            VALUES ('{}')".format(a)) 
    conn.commit()
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Category Added Successfully",reply_markup=reply_markup)
    return BUTTON
def delete(update,context):
    msg=update.message.text
    print(msg)
    connection = sqlite3.connect("shop.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT productid FROM COMPANY where productid='{}'".format(msg))
    jobs = cursor.fetchall()
    print(jobs)
    if len(jobs) !=0:
        cursor = connection.cursor() 
        cursor.execute("DELETE  from COMPANY where productID='{}'".format(msg))
        connection.commit()
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Product Deleted",reply_markup=reply_markup)
        return BUTTON
    else:
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="INVALID product ID",reply_markup=reply_markup)
        return DELETE
def pstock(update,context):
    msg=update.message.text
    global ueue
    ueue=msg
    connection = sqlite3.connect("shop.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT productid FROM COMPANY where productid='{}'".format(msg))
    jobs = cursor.fetchall()
    if len(jobs) !=0:

        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Send new Stock Number",reply_markup=reply_markup)
        return PSTOCKA
    else:
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
        context.bot.send_message(chat_id=update.effective_user.id,text="Invalid Product ID",reply_markup=reply_markup)
        return BUTTON
def pstocka(update,context):
    msg=update.message.text
    conn = sqlite3.connect("shop.db")  
    conn.execute("UPDATE COMPANY set stock = '{}' where productID = '{}'".format(msg,ueue))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Stock Updated",reply_markup=reply_markup)
    return BUTTON
#ADD PRODUCT
def ab(update,context):
    msg=update.message.text
    global nm
    nm =msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Send Product Image',reply_markup=reply_markup)
    return PHOTO
def photo(update,context):
    msg=update.message.photo[-1].file_id
    global pha
    pha =msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Send Description of Product',reply_markup=reply_markup)
    return DES
def des(update,context):
    msg=update.message.text
    global ds
    ds =msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Send the Price of Your Product',reply_markup=reply_markup)
    return PR 
def pr(update,context):
    msg=update.message.text
    global pri
    pri =msg
    try:
      pri=float(pri)
      keyf=[]
      connection = sqlite3.connect("cata.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT cat FROM COMPANY")
      for name in cursor:
        name=name[0]
        name=str(name)
        cv=[]
        cv.append(name)
        keyf.append(cv)
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Select a Category",reply_markup=reply_markup)  
      return PIDA 
    except:
      context.bot.send_message(chat_id=update.effective_user.id,text='Invalid price! send in digits only')
      return PR 
def pida(update,context):
    msg=update.message.text
    global cat
    cat =msg
    keyf=[]
    connection = sqlite3.connect("subcata.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT subcat FROM COMPANY where cat='{}'".format(str(msg)))
    for name in cursor:
        name=name[0]
        name=str(name)
        cv=[]
        cv.append(name)
        keyf.append(cv)
    reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Select a Sub-Category",reply_markup=reply_markup)  
    return PID
def pid(update,context):
    msg=update.message.text             
    global sctr
    sctr =msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Send Special Product ID',reply_markup=reply_markup)
    return TYPP
def typp(update,context):
    msg=update.message.text             
    global specialid
    specialid =msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text="Send the Max Quantity a User Can Buy",reply_markup=reply_markup)
    return MADDA
def madda(update,context):
    msg=update.message.text             
    global ssttoo
    ssttoo =msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
    context.bot.send_message(chat_id=update.effective_user.id,text='Send Stock Quantity as a Number',reply_markup=reply_markup)
    return MADD
def madd(update,context):
    msg=update.message.text
    print(nm)
    try:
      msg=int(msg)
      conn = sqlite3.connect('shop.db')
      conn.execute("INSERT INTO COMPANY (name,description,price,productID,photoid,category,subcat,stock,max) \
                          VALUES ('{}', '{}','{}', '{}','{}','{}','{}','{}','{}')".format(nm,ds,pri,specialid,pha,cat,sctr,msg,ssttoo))
      conn.commit()
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Product Added Successfully" ,reply_markup=reply_markup)
      return BUTTON
    except:
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True )
      context.bot.send_message(chat_id=update.effective_user.id,text='ERROR\nSend total quantities of Product in stock, Should be a number',reply_markup=reply_markup)
      return MADD
def cancel(update, context):
    user = update.message.from_user
    return ConversationHandler.END
def main():                        
  updater = Updater("5784876602:AAHmGUcBvgqD_S3tI-p9CEvZKEeiq1l6oI0", use_context=True)
  dp = updater.dispatcher
  conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start),MessageHandler(Filters.regex("^(❓ Servicio de asistencia| 🙋 FAQ |🛍 Mis ordenes|💳 Cartera|📝️ Productos|🛒 Carro|📝️ Products|🛒 Cart|💳 Wallet|💳 Add Credit|🛍 My Orders|❓ Help/Support|🔙 Main Menu|📝️ Des produits|🙋 FAQ|🛒 Chariot|🙋 FAQs|❓ Support d'aide|🛍 Mes commandes|💳 Porte monnaie)$"), jos)],

        states={
            
            BUTTON: [CallbackQueryHandler(button)],#ASUB,ASUBB
            AB: [MessageHandler(Filters.text, ab),CallbackQueryHandler(button)],
            PHOTO: [MessageHandler(Filters.photo, photo),CallbackQueryHandler(button)],
            DES: [MessageHandler(Filters.text, des),CallbackQueryHandler(button)],
            PR: [MessageHandler(Filters.text, pr),CallbackQueryHandler(button)],
            PIDA: [MessageHandler(Filters.text, pida),CallbackQueryHandler(button)],
            PID: [MessageHandler(Filters.text, pid),CallbackQueryHandler(button)],
            TYPP: [MessageHandler(Filters.text, typp),CallbackQueryHandler(button)],
            MADD: [MessageHandler(Filters.text, madd),CallbackQueryHandler(button)],
            MADDA: [MessageHandler(Filters.text, madda),CallbackQueryHandler(button)],
            DELETE: [MessageHandler(Filters.text, delete),CallbackQueryHandler(button)],
            YUO: [MessageHandler(Filters.text, yuo),CallbackQueryHandler(button)],
            YUOS: [MessageHandler(Filters.text, yuos),CallbackQueryHandler(button)],
            YUOF: [MessageHandler(Filters.text, yuof),CallbackQueryHandler(button)],
            CH: [MessageHandler(Filters.text, ch),CallbackQueryHandler(button)],
            BEL: [CallbackQueryHandler(bel)],
            DFAQ: [CallbackQueryHandler(dfaq)],
            BELA: [CallbackQueryHandler(bela)],
            ASUB: [CallbackQueryHandler(asub)],
            BYES: [CallbackQueryHandler(byes)],
            DDFFS: [CallbackQueryHandler(ddffs)],
            JOP: [CallbackQueryHandler(jop)],
            JOPS: [CallbackQueryHandler(jops)],
            JOPF: [CallbackQueryHandler(jopf)],
            ASUBB: [MessageHandler(Filters.text, asubb),CallbackQueryHandler(button)],
            FIKA: [MessageHandler(Filters.text, fika),CallbackQueryHandler(button)],
            FIKAA: [MessageHandler(Filters.text, fikaa),CallbackQueryHandler(button)],
            PDVM: [CallbackQueryHandler(pdvm)],
            SHAKA: [CallbackQueryHandler(shaka)],
            PDVMF: [CallbackQueryHandler(pdvmf)],
            PDVM: [CallbackQueryHandler(pdvm)],
            SHAKAS: [CallbackQueryHandler(shakas)],
            PDVMS: [CallbackQueryHandler(pdvms)],
            SHAKAF: [CallbackQueryHandler(shakaf)],
            PDBS: [CallbackQueryHandler(pdbs)],
            PDB: [CallbackQueryHandler(pdb)],
            PDBF: [CallbackQueryHandler(pdbf)],
            DDEPO: [CallbackQueryHandler(ddepo)],
            PROFT: [MessageHandler(Filters.photo, proft),CallbackQueryHandler(button)],
            PROFTS: [MessageHandler(Filters.photo, profts),CallbackQueryHandler(button)],
            PROFTF: [MessageHandler(Filters.photo, proftf),CallbackQueryHandler(button)],
            WALL: [MessageHandler(Filters.text, wall),CallbackQueryHandler(button)],
            UOL: [MessageHandler(Filters.text, uol),CallbackQueryHandler(button)],
            ADDB: [MessageHandler(Filters.text, addb),CallbackQueryHandler(button)],
            TRADE: [MessageHandler(Filters.text, trade),CallbackQueryHandler(button)],
            TRADEF: [MessageHandler(Filters.text, tradef),CallbackQueryHandler(button)],
            TRADES: [MessageHandler(Filters.text, trades),CallbackQueryHandler(button)],
            HAM: [MessageHandler(Filters.text, ham),CallbackQueryHandler(button)],
            ADDRESS: [MessageHandler(Filters.text, address),CallbackQueryHandler(button)],
            ADDRESSS: [MessageHandler(Filters.text, addresss),CallbackQueryHandler(button)],
            ADDRESSF: [MessageHandler(Filters.text, addressf),CallbackQueryHandler(button)],
            FIKAF: [MessageHandler(Filters.text, fikaf),CallbackQueryHandler(button)],
            FIKAAF: [MessageHandler(Filters.text, fikaaf),CallbackQueryHandler(button)],
            DDFF: [CallbackQueryHandler(ddff)],
            DDFFF: [CallbackQueryHandler(ddfff)],
            BYE: [CallbackQueryHandler(bye)],
            WWRR: [CallbackQueryHandler(wwrr)],
            PENORD: [CallbackQueryHandler(penord)],
            FIKAS: [MessageHandler(Filters.text, fikas),CallbackQueryHandler(button)],
            FIKAAS: [MessageHandler(Filters.text, fikaas),CallbackQueryHandler(button)],
            HALP: [MessageHandler(Filters.text, halp),CallbackQueryHandler(button)],
            DEPO: [MessageHandler(Filters.text, depo),CallbackQueryHandler(button)],
            JE: [MessageHandler(Filters.text, je),CallbackQueryHandler(button)],
            SH: [MessageHandler(Filters.text, sh),CallbackQueryHandler(button)],
            DEPOT: [MessageHandler(Filters.text, depot),CallbackQueryHandler(button)],
            CUI: [MessageHandler(Filters.text, cui),CallbackQueryHandler(button)],
            DBS: [MessageHandler(Filters.text, dbs),CallbackQueryHandler(button)],
            DBSS: [MessageHandler(Filters.text, dbss),CallbackQueryHandler(button)],
            CUIT: [MessageHandler(Filters.text, cuit),CallbackQueryHandler(button)],

            DBSF: [MessageHandler(Filters.text, dbsf),CallbackQueryHandler(button)],
            CUIF: [MessageHandler(Filters.text, cuif),CallbackQueryHandler(button)],
            CUIS: [MessageHandler(Filters.text, cuis),CallbackQueryHandler(button)],
            CUITF: [MessageHandler(Filters.text, cuitf),CallbackQueryHandler(button)],
            CUITS: [MessageHandler(Filters.text, cuits),CallbackQueryHandler(button)],
            CUITQ: [MessageHandler(Filters.text, cuitq),CallbackQueryHandler(button)],
            CUITQF: [MessageHandler(Filters.text, cuitqf),CallbackQueryHandler(button)],
            CUITQS: [MessageHandler(Filters.text, cuitqs),CallbackQueryHandler(button)],
            PSTOCK: [MessageHandler(Filters.text, pstock),CallbackQueryHandler(button)],
            PSTOCKA: [MessageHandler(Filters.text, pstocka),CallbackQueryHandler(button)],
       
        }, 
        fallbacks=[CommandHandler('cancel', cancel)],
        allow_reentry=True
    )
  dp.add_handler(conv_handler)
  updater.start_polling()
  updater.idle()
    
if __name__ == '__main__':
    main()




