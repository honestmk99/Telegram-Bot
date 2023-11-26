import telegram.ext
import telegram.ext 
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, ConversationHandler,CallbackQueryHandler
from telegram.ext import (
    Updater,
    CommandHandler,
    MessageHandler,
    Filters,
    PreCheckoutQueryHandler,
    ShippingQueryHandler,
)
from telegram.error import TelegramError
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram import (ReplyKeyboardMarkup, ReplyKeyboardRemove)
import datetime
import sqlite3
import random
import requests
from datetime import date
from datetime import datetime, timedelta
import datetime
import sqlite3 as sql
import os
import smtplib, ssl
import json
import random
import string
from sqlite3 import Error
from telegram import KeyboardButton 
from telegram import LabeledPrice, ShippingOption
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram import (ReplyKeyboardMarkup, ReplyKeyboardRemove)
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
import re
from coinbase_commerce.client import Client
from datetime import datetime
from datetime import date
import xlwt
from openpyxl import load_workbook
client = Client(api_key="e613b63f-2f88-4570-9596-5af712b4f6e0")
wc=["WELCOME TO SCAMILY VALUES‼️🥷🏽\nHere at SCAMILY VALUES you will find EVERYTHING you need to succeed‼️\n\n🤝Here you can automatically buy what you need🤫\n🤝If you have any problems - our support is online 24/7!\n🤝Good Luck‼️"]
hel=["admin will update soon"]
fa=["admin will update soon"]

BUTTON,AB,DES,ABAD,DESAD,CATEGORY,CATEGORYAD,PR,PRAD,PD,PID,CART,DELETE,PHO,PHOAD,MODIFY,PIDAD,CH,SH,WA,SHO,JE,ADDRESS,STOCK,PHOSTOCK,DESSTOCK,PIDSTOCK,PRSTOCK,CATEGORYSTOCK,PDSTOCK,HAM,AMAN,RMAN,AMO,TRC,CRP,FCAT,TYPP,MODF,MODFF,AKEY,COU,COD,DEL,FAQ,C1,C2,C3,C4,ADFILE,ADFILEE,QWER,TRADE= range(53) 
man=[]
def start(update, context):
  userg = update.message.from_user
  try:
    usaf=userg.username
  except:
    usaf=userg.first_name
  user=update.effective_user.id
  print(update.effective_user.id)
  user = update.message.from_user
  usa=str(update.effective_user.id)
  user=int(update.effective_user.id)
  print(user)
  ref_id = update.message.text
  ref_id=ref_id.split()
  if len(ref_id) > 1:
          asf=str(ref_id[1]).strip()
          connection = sqlite3.connect("users.db")  
          cursor = connection.cursor()  
          cursor.execute("SELECT id FROM COMPANY where id= {}".format(int(user))) 
          jobs = cursor.fetchall()
          if len(jobs) ==0:
              cursor.execute("INSERT INTO COMPANY (ID) \
                VALUES ({})".format(int(user)))
              connection.commit()
              connection.close()
              connection = sqlite3.connect("wallet.db")  
              cursor = connection.cursor() 
              cursor.execute("INSERT INTO COMPANY (ID,balance,link,code,amount,name,refby,tref,tradview,email) \
                VALUES ({}, '{}','{}','{}','{}','{}','{}','{}','{}','{}')".format(int(user),"0","0","0","0",usaf,asf,"0","0","0"))
              connection.commit()
              connection.close()
              conn = sqlite3.connect('oo.db')
              conn.execute("INSERT INTO COMPANY (ID,pid,pname,type,amount) \
                      VALUES ('{}', '{}','{}', '{}','{}')".format(str(update.effective_user.id),"0","0","0","0")) 
              conn.commit()
              conn = sqlite3.connect('wallet.db')
              cursor=conn.execute("SELECT tref FROM COMPANY where ID={}".format(int(asf))) 
              conn.commit()
              for row in cursor:
                  gh=int(row[0])
                  gh=gh+1
                  print(gh)
                  
                  conn.execute("UPDATE COMPANY set tref = '{}' where ID = {}".format(gh,int(asf)))
                  conn.commit()
          keyf=[["📝️ Products","🛒 Cart"],["💳 Wallet","💳 Add Credit"],["🛍 My orders","❓ Help/Support"],["🙋 FAQ","🤝 Affiliate Program"]]
          reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
          context.bot.send_message(chat_id=update.effective_user.id,text="User Panel",reply_markup=reply_markup)


  elif usa == "1926801217" or usa == "1394902938" or usa =="5515451493" or usa in man:#5515451493
    keyboard =[[InlineKeyboardButton("➕ Product", callback_data="1"),InlineKeyboardButton("❌ Product", callback_data="2"),
                InlineKeyboardButton("⚙️ Modify Product", callback_data="32")],[InlineKeyboardButton("⚙️ Delete Category", callback_data="del"),
                InlineKeyboardButton("⚙️ ADD Category", callback_data="90")],[InlineKeyboardButton("🛒 Orders", callback_data="99"),
                InlineKeyboardButton("🛒 ADD File", callback_data="1file"),InlineKeyboardButton("Managers List", callback_data="lman")],
                [InlineKeyboardButton("🙋‍♂️ Welome message", callback_data="912"),InlineKeyboardButton("❓ Set Help / Support", callback_data="913"),
                InlineKeyboardButton("Edit FAQ", callback_data="faq")],[InlineKeyboardButton("🔊 Announcement", callback_data="916"),
                InlineKeyboardButton("Referral Data", callback_data="refd")],[InlineKeyboardButton("Modify the Product's file", callback_data="3344"),
                InlineKeyboardButton("User's Data", callback_data="tdata")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to admin Dashboard" ,reply_markup=reply_markup)
    return BUTTON
  else:
      connection = sqlite3.connect("users.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT id FROM COMPANY where id= {}".format(int(user))) 
      jobs = cursor.fetchall()
      if len(jobs) ==0:
          cursor.execute("INSERT INTO COMPANY (ID) \
            VALUES ({})".format(int(user)))
          connection.commit()
          connection.close()
          connection = sqlite3.connect("wallet.db")  
          cursor = connection.cursor() 
          cursor.execute("INSERT INTO COMPANY (ID,balance,link,code,amount,name,refby,tref,tradview,email) \
            VALUES ({}, '{}','{}','{}','{}','{}','{}','{}','{}','{}')".format(int(user),"0","0","0","0",usaf,"0","0","0","0"))
          connection.commit()
          connection.close()
          conn = sqlite3.connect('oo.db')
          conn.execute("INSERT INTO COMPANY (ID,pid,pname,type,amount) \
                  VALUES ('{}', '{}','{}', '{}','{}')".format(str(update.effective_user.id),"0","0","0","0")) 
          conn.commit()
      keyf=[["📝️ Products","🛒 Cart"],["💳 Wallet","💳 Add Credit"],["🛍 My orders","❓ Help/Support"],["🙋 FAQ","🤝 Affiliate Program"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="User DashBoard",reply_markup=reply_markup)
def button(update,context):
    h=update.effective_user.id
    query = update.callback_query
    a=query.data
    if a=="3344":
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text='Send productID of product to modify its file',reply_markup=reply_markup)
        return MODF
    elif a=="1file":
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text='Send productID of product to add file',reply_markup=reply_markup)
        return ADFILE
    elif a=="3678":
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text='Send photo to upadate',reply_markup=reply_markup)
        return AKEY
    elif a== 'tdata':
        style = xlwt.easyxf('font: bold 1, color black;')
        conn = sqlite3.connect('wallet.db') 
        cursor = conn.execute("SELECT ID,balance,email,tradview from COMPANY")
        conn.commit() 
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('wallet.db') 
            cursor = conn.execute("SELECT ID,balance,email,tradview  from COMPANY ")
            conn.commit()
            xa="Trading View\n"
            i=0
            j=0
            workbook = xlwt.Workbook()
            sheet = workbook.add_sheet("Referrals Data")
            sheet.write(i, 1, "username", style)
            sheet.write(i, 2, "Email", style)
            sheet.write(i, 3, "Balance",  style)

            for row in cursor:
                if row[3]!="0":
                    i=i+1
                    j=j+1
                    inv=row[0]
                    vgy=row[1]
                    xdrt=row[2]
                    sheet.write(i, 1, row[3], style)
                    sheet.write(i, 2, row[2], style)
                    sheet.write(i, 3, row[1],  style)


            workbook.save("td.xls")
            keyboard =[[InlineKeyboardButton("Main Menu", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_document(chat_id=update.effective_user.id,document=open('td.xls', 'rb'),reply_markup=reply_markup)
    elif a== 'refd':
        style = xlwt.easyxf('font: bold 1, color black;')
        conn = sqlite3.connect('wallet.db') 
        cursor = conn.execute("SELECT ID,balance,link,code,amount,name,refby,tref from COMPANY where tref>=1")
        conn.commit() 
        jobs = cursor.fetchall()
        if len(jobs) !=0:
            conn = sqlite3.connect('wallet.db') 
            cursor = conn.execute("SELECT ID,balance,link,code,amount,name,refby,tref from COMPANY where tref>=1")
            conn.commit()
            xa="Referrals\n"
            i=0
            j=0
            workbook = xlwt.Workbook()
            sheet = workbook.add_sheet("Referrals Data")
            sheet.write(i, 1, "User_ID", style)
            sheet.write(i, 2, "user_name", style)
            sheet.write(i, 3, "Total_Referrals",  style)
            sheet.write(i, 4,"Balance",  style)
            for row in cursor:
                    i=i+1
                    j=j+1
                    inv=row[0]
                    vgy=row[1]
                    xdrt=row[2]
                    sheet.write(i, 1, row[0], style)
                    sheet.write(i, 2, row[5], style)
                    sheet.write(i, 3, row[7],  style)
                    sheet.write(i, 4, row[1],  style)

            workbook.save("ref.xls")
            keyboard =[[InlineKeyboardButton("Main Menu", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_document(chat_id=update.effective_user.id,document=open('ref.xls', 'rb'),reply_markup=reply_markup)
    elif a== '440':
        keyf=[]
        connection = sqlite3.connect("cata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT cat FROM COMPANY")
        for name in cursor:
          name=name[0]
          name=str(name)
          cv=[]
          cv.append(name)
          keyf.append(cv)
        reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Select the category",reply_markup=reply_markup)  
        return PDSTOCK
        return AB

    elif a=="912":
        context.bot.send_message(chat_id=update.effective_user.id,text="Send Welcome message")  
        
        return SH

    elif a=="9956":
        connection = sqlite3.connect("wallet.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT code,link FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
        for names in cursor:
          inv=names[0]
          inv1=names[1]
        if inv=="0":
          charge = client.charge.create(name='Telegram botShop',
                            description='Pay to add credit to your wallet',
                            pricing_type='no_price',
                            local_price={ 
                        })
          linka=charge["hosted_url"]
          coda=charge["id"]
          conn = sqlite3.connect("wallet.db")  
          conn.execute("UPDATE COMPANY set link = '{}', code='{}' where ID = {}".format(linka,coda,int(update.effective_user.id)))
          conn.commit()
          conn.close()
          keyboard =[[InlineKeyboardButton("Make a Payment",url=linka)],[InlineKeyboardButton("I have paid", callback_data="927")],[InlineKeyboardButton("🔙 Back", callback_data="200")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
          context.bot.send_message(chat_id=update.effective_user.id,text="Click on below button to make payment and after making payment wait for few minutes then come back and click on i have paid.",reply_markup=reply_markup)
          return BUTTON 
        else:
          aa=client.charge.retrieve(inv)
          b=aa['timeline']
          a=0
          for names in b:
            ax=names['status']
            if ax=="NEW":
              a=4
          for names in b:
            ax=names['status']
            if ax=="EXPIRED":
              a=3
          for names in b:
            ax=names['status']
            if ax=="PENDING":
              a=2
          for names in b:
            ax=names['status']
            if ax=="COMPLETED":
              a=1
          if a==1:
            cv=aa['payments']
            for names in cv:
              vb=names['value']['local']['amount']
              vb=float(vb)
              vb=round(vb,5)
              print(vb)
              connection = sqlite3.connect("wallet.db")  
              cursor = connection.cursor()  
              cursor.execute("SELECT balance FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
              for names in cursor:
                inv=float(names[0])
              bn=vb+inv
              bn=str(bn)
              conn = sqlite3.connect("wallet.db")  
              conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(bn,int(update.effective_user.id)))
              conn.commit()
              conn.close()
              conn = sqlite3.connect("wallet.db")  
              conn.execute("UPDATE COMPANY set link = '0', code='0' where ID = {}".format(int(update.effective_user.id)))
              conn.commit()
              conn.close()
              keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
              context.bot.send_message(chat_id=update.effective_user.id,text="£{} added to your wallet".format(vb),reply_markup=reply_markup)
              return BUTTON
          elif a==2:
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")],[InlineKeyboardButton("Check Again", callback_data="927")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Your previous transcation is pending please wait for its completion",reply_markup=reply_markup)
            return BUTTON
          elif a==3:
              charge = client.charge.create(name='Telegram botShop',
                            description='Pay to add credit to your wallet',
                            pricing_type='no_price',
                            local_price={ 
                            })
              linka=charge["hosted_url"]
              coda=charge["id"]
              conn = sqlite3.connect("wallet.db")  
              conn.execute("UPDATE COMPANY set link = '{}', code='{}' where ID = {}".format(linka,coda,int(update.effective_user.id)))
              conn.commit()
              conn.close()
              keyboard =[[InlineKeyboardButton("Make a Payment",url=linka)],[InlineKeyboardButton("I have paid", callback_data="927")],[InlineKeyboardButton("🔙 Back", callback_data="200")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
              context.bot.send_message(chat_id=update.effective_user.id,text="Click on below button to make payment and after making payment wait for few minutes then come back and click on i have paid.",reply_markup=reply_markup)
              return BUTTON

          else:
            connection = sqlite3.connect("wallet.db")  
            cursor = connection.cursor()  
            cursor.execute("SELECT link FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
            for names in cursor:
              inv1=names[0]
            keyboard =[[InlineKeyboardButton("Make a Payment",url=inv1)],[InlineKeyboardButton("I have paid", callback_data="927")],[InlineKeyboardButton("🔙 Back", callback_data="200")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Click on below button to make payment and after making payment wait for few minutes then come back and click on i have paid.",reply_markup=reply_markup)
            return BUTTON
    elif a=="913":
        context.bot.send_message(chat_id=update.effective_user.id,text="Send Help/Support message")  
        
        return SHO
    elif a=="faq":
        context.bot.send_message(chat_id=update.effective_user.id,text="Send FAQ message")  
        
        return FAQ
    elif a=="aman"or a=="rman" or a=="lman":
      if a=="aman":
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text='Send Id of the user to add as manager',reply_markup=reply_markup)
        return AMAN
      elif a=="rman":
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text='Send Id of the user to romve as manager',reply_markup=reply_markup)
        return RMAN
      else:
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        a="Following is List of managers:\n\n"
        b=1
        for names in man:
          a=a+str(b)+") "+names+"\n\n"
          b=b+1
        context.bot.send_message(chat_id=update.effective_user.id,text=a,reply_markup=reply_markup)
        return RMAN
        
    elif a=="90":      
        context.bot.send_message(chat_id=update.effective_user.id,text="Send all catagories seperated by comma like bag,shirts,jeans")  
        
        return CH
    elif a== '914':
 
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text=hel[0],reply_markup=reply_markup)
            return BUTTON
    elif a== '1':
 
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send name of product',reply_markup=reply_markup)
            return AB
    elif a== '916':

            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send announcement message',reply_markup=reply_markup)
            return JE
    elif a== '67':
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send name of product for Pre-Order',reply_markup=reply_markup)

            return STOCK
    elif a=="920":
          print(update)

          connection = sqlite3.connect("wallet.db")  
          cursor = connection.cursor()  
          cursor.execute("SELECT balance,name FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
          for names in cursor:
            inv=names[0]
            gsr=names[1]
            inv=float(inv)
            inv=round(inv,5)
            inv=str(inv)
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Wallet ID:  "+str(update.effective_user.id)+"\n\nName:  "+gsr+"\nBalance:  "+"$"+inv,reply_markup=reply_markup)
    elif a== '32':

            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send ProductID to modify the product',reply_markup=reply_markup)
            return MODIFY
    elif a=="back":
      c=update.callback_query.message.message_id
      cc=update.callback_query.message.text
      d=cc
      vbm=d.split("Orders Page")
      vbm=vbm[1]
      vbm=vbm.split(":-")
      vbm=vbm[0]
      vbm=vbm.strip()
      opno=vbm
      vbm=int(vbm)
      vbm=vbm-1
      npno=vbm
      vbm=vbm*10
      conn = sqlite3.connect('orders.db')
      cursor = conn.execute("SELECT ID from COMPANY")
      jobs = cursor.fetchall()
      n=len(jobs)
      n=n-vbm
      a="Orders Page {}:-\n\n".format(npno)
      cursor = conn.execute("SELECT name,quantity,price,date,pnam,address from COMPANY LIMIT 10 OFFSET {}".format(n))
      conn.commit()
      for row in cursor:
        g="\nProduct Name:  "+row[0]+"\nQuantity: "+row[1]+"\nCustomer Name:  "+row[4]+"\nProduct Price:  "+row[2]+ "\nBuyer's Address:  "+row[5]+"\nDate: "+row[3]+"\n\n"
        a=a+g
      if npno>1:
        keyboard =[[InlineKeyboardButton("Next page", callback_data="page"),InlineKeyboardButton("Previous page", callback_data="back")],[InlineKeyboardButton("🔙 Main Menu", callback_data="100")]]
      else:
        keyboard =[[InlineKeyboardButton("Next page", callback_data="page")],[InlineKeyboardButton("🔙 Main Menu", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.edit_message_text(chat_id=update.effective_user.id,message_id=c,text=a,reply_markup=reply_markup)
      return BUTTON
    elif a=="page":
      c=update.callback_query.message.message_id
      cc=update.callback_query.message.text
      d=cc
      vbm=d.split("Orders Page")
      vbm=vbm[1]
      vbm=vbm.split(":-")
      vbm=vbm[0]
      vbm=vbm.strip()
      opno=vbm
      vbm=int(vbm)
      vbm=vbm+1
      npno=vbm
      vbm=vbm*10
      conn = sqlite3.connect('orders.db')
      cursor = conn.execute("SELECT ID from COMPANY")
      jobs = cursor.fetchall()
      n=len(jobs)
      n=n-vbm
      a="Orders Page {}:-\n\n".format(npno)
      cursor = conn.execute("SELECT name,quantity,price,date,pnam,address from COMPANY LIMIT 10 OFFSET {}".format(n))
      conn.commit()
      for row in cursor:
        g="\nProduct Name:  "+row[0]+"\nQuantity: "+row[1]+"\nCustomer Name:  "+row[4]+"\nProduct Price:  "+row[2]+ "\nBuyer's Address:  "+row[5]+"\nDate: "+row[3]+"\n\n"
        a=a+g
        
      if n>0:
        keyboard =[[InlineKeyboardButton("Next page", callback_data="page"),InlineKeyboardButton("Previous page", callback_data="back")],[InlineKeyboardButton("🔙 Main Menu", callback_data="100")]]
      else:
        keyboard =[[InlineKeyboardButton("Previous page", callback_data="back")],[InlineKeyboardButton("🔙 Main Menu", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.edit_message_text(chat_id=update.effective_user.id,message_id=c,text=a,reply_markup=reply_markup)
      return BUTTON

    elif a=="uspage":
      c=update.callback_query.message.message_id
      cc=update.callback_query.message.text
      d=cc
      vbm=d.split("Orders Page")
      vbm=vbm[1]
      vbm=vbm.split(":-")
      vbm=vbm[0]
      vbm=vbm.strip()
      opno=vbm
      vbm=int(vbm)
      vbm=vbm+1
      npno=vbm
      vbm=vbm*10
      conn = sqlite3.connect('orders.db')
      cursor = conn.execute("SELECT ID from COMPANY where ID= {}".format(int(update.effective_user.id)))
      jobs = cursor.fetchall()
      n=len(jobs)
      n=n-vbm
      a="Orders Page {}:-\n\n".format(npno)
      cursor = conn.execute("SELECT name,quantity,price,date,status,productID from COMPANY where ID='{}' LIMIT 10 OFFSET {}".format(int(update.effective_user.id),n))
      conn.commit()
      for row in cursor:
        g="\nProduct Name:  "+row[0]+"\nQuantity: "+row[1]+"\nProduct Price:  "+row[2]+"\nProduct ID:  "+row[5]+"\nDate: "+row[3]+"\nStatus: "+row[4]+"\n\n"
        a=a+g
      if n>0:
        keyboard =[[InlineKeyboardButton("Next page", callback_data="uspage"),InlineKeyboardButton("Previous page", callback_data="usback")],[InlineKeyboardButton("🔙 Main Menu", callback_data="100")]]
      else:
        keyboard =[[InlineKeyboardButton("Previous page", callback_data="usback")],[InlineKeyboardButton("🔙 Main Menu", callback_data="200")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.edit_message_text(chat_id=update.effective_user.id,message_id=c,text=a,reply_markup=reply_markup)
      return BUTTON
    
    elif a=="usback":
      c=update.callback_query.message.message_id
      cc=update.callback_query.message.text
      d=cc
      vbm=d.split("Orders Page")
      vbm=vbm[1]
      vbm=vbm.split(":-")
      vbm=vbm[0]
      vbm=vbm.strip()
      opno=vbm
      vbm=int(vbm)
      vbm=vbm-1
      npno=vbm
      vbm=vbm*10
      conn = sqlite3.connect('orders.db')
      cursor = conn.execute("SELECT ID from COMPANY")
      jobs = cursor.fetchall()
      n=len(jobs)
      n=n-vbm
      print(n)
      a="Orders Page {}:-\n\n".format(npno)
      cursor = conn.execute("SELECT name,quantity,price,date,status,productID from COMPANY where ID='{}' LIMIT 10 OFFSET {}".format(int(update.effective_user.id),n))
      conn.commit()
      for row in cursor:
        g="\nProduct Name:  "+row[0]+"\nQuantity: "+row[1]+"\nProduct Price:  "+row[2]+"\nProduct ID:  "+row[5]+"\nDate: "+row[3]+"\nStatus: "+row[4]+"\n\n"
        a=a+g
      if npno>1:
        keyboard =[[InlineKeyboardButton("Next page", callback_data="uspage"),InlineKeyboardButton("Previous page", callback_data="usback")],[InlineKeyboardButton("🔙 Main Menu", callback_data="100")]]
      else:
        keyboard =[[InlineKeyboardButton("Next page", callback_data="uspage")],[InlineKeyboardButton("🔙 Main Menu", callback_data="200")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.edit_message_text(chat_id=update.effective_user.id,message_id=c,text=a,reply_markup=reply_markup)
      return BUTTON
    elif a=="trc":
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text='Send wallet ID of user to make transcation',reply_markup=reply_markup)
      return TRC

    elif a== '33':
        keyf=[]
        connection = sqlite3.connect("cata.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT cat FROM COMPANY")
        for name in cursor:
          name=name[0]
          name=str(name)
          cv=[]
          cv.append(name)
          keyf.append(cv)
        reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Select the category",reply_markup=reply_markup)   
        return PD 
    elif a=="919":
        connection = sqlite3.connect("wallet.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT code,link FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
        for names in cursor:
          inv=names[0]
          inv1=names[1]
        if inv=="0":
          charge = client.charge.create(name='Telegram botShop',
                            description='Pay to add credit to your wallet',
                            pricing_type='no_price',
                            local_price={ 
                        })
          linka=charge["hosted_url"]
          coda=charge["id"]
          conn = sqlite3.connect("wallet.db")  
          conn.execute("UPDATE COMPANY set link = '{}', code='{}' where ID = {}".format(linka,coda,int(update.effective_user.id)))
          conn.commit()
          conn.close()
          keyboard =[[InlineKeyboardButton("Make a Payment",url=linka)],[InlineKeyboardButton("I have paid", callback_data="927")],[InlineKeyboardButton("🔙 Back", callback_data="200")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
          context.bot.send_message(chat_id=update.effective_user.id,text="Click on below button to make payment and after making payment wait for few minutes then come back and click on i have paid.",reply_markup=reply_markup)
          return BUTTON 
        else:
          aa=client.charge.retrieve(inv)
          b=aa['timeline']
          a=0
          for names in b:
            ax=names['status']
            if ax=="NEW":
              a=4
          for names in b:
            ax=names['status']
            if ax=="EXPIRED":
              a=3
          for names in b:
            ax=names['status']
            if ax=="PENDING":
              a=2
          for names in b:
            ax=names['status']
            if ax=="COMPLETED":
              a=1
          if a==1:
            cv=aa['payments']
            for names in cv:
              vb=names['value']['local']['amount']
              vb=float(vb)
              vb=round(vb,5)
              print(vb)
              connection = sqlite3.connect("wallet.db")  
              cursor = connection.cursor()  
              cursor.execute("SELECT balance FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
              for names in cursor:
                inv=float(names[0])
              bn=vb+inv
              bn=str(bn)
              conn = sqlite3.connect("wallet.db")  
              conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(bn,int(update.effective_user.id)))
              conn.commit()
              conn.close()
              conn = sqlite3.connect("wallet.db")  
              conn.execute("UPDATE COMPANY set link = '0', code='0' where ID = {}".format(int(update.effective_user.id)))
              conn.commit()
              conn.close()
              keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
              context.bot.send_message(chat_id=update.effective_user.id,text="${} is added to your wallet".format(vb),reply_markup=reply_markup)
              return BUTTON
          elif a==2:
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")],[InlineKeyboardButton("Check Again", callback_data="927")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Your previous transcation is pending please wait for its completion",reply_markup=reply_markup)
            return BUTTON
          elif a==3:
              charge = client.charge.create(name='Telegram botShop',
                            description='Pay to add credit to your wallet',
                            pricing_type='no_price',
                            local_price={ 
                            })
              linka=charge["hosted_url"]
              coda=charge["id"]
              conn = sqlite3.connect("wallet.db")  
              conn.execute("UPDATE COMPANY set link = '{}', code='{}' where ID = {}".format(linka,coda,int(update.effective_user.id)))
              conn.commit()
              conn.close()
              keyboard =[[InlineKeyboardButton("Make a Payment",url=linka)],[InlineKeyboardButton("I have paid", callback_data="927")],[InlineKeyboardButton("🔙 Back", callback_data="200")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
              context.bot.send_message(chat_id=update.effective_user.id,text="Click on below button to make payment and after making payment wait for few minutes then come back and click on i have paid.",reply_markup=reply_markup)
              return BUTTON

          else:
            connection = sqlite3.connect("wallet.db")  
            cursor = connection.cursor()  
            cursor.execute("SELECT link FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
            for names in cursor:
             inv1=names[0]
            keyboard =[[InlineKeyboardButton("Make a Payment",url=inv1)],[InlineKeyboardButton("I have paid", callback_data="927")],[InlineKeyboardButton("🔙 Back", callback_data="200")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Click on below button to make payment and after making payment wait for few minutes then come back and click on i have paid.",reply_markup=reply_markup)
            return BUTTON
                    
    elif a== '927':
      connection = sqlite3.connect("wallet.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT code FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
      for names in cursor:
        inv=names[0]
      aa=client.charge.retrieve(inv)
      b=aa['timeline']
      a=0
      for names in b:
        ax=names['status']
        if ax=="NEW":
          a=4
      for names in b:
        ax=names['status']
        if ax=="EXPIRED":
          a=3
      for names in b:
        ax=names['status']
        if ax=="PENDING":
          a=2
      for names in b:
        ax=names['status']
        if ax=="COMPLETED":
          a=1
    
      if a==1:
        cv=aa['payments']
        for names in cv:
          vb=names['value']['local']['amount']
          vb=float(vb)        
          conn = sqlite3.connect("wallet.db") 
          cursor=conn.execute("SELECT balance,refby,name FROM COMPANY where id= {}".format(int(update.effective_user.id)) )       
          for names in cursor:
            inv=float(names[0])
            reffby=names[1]
            thu=names[2]
            cursor.execute("SELECT balance  FROM COMPANY where id= {}".format(int(reffby)) )
            for names in cursor:
              bala=float(names[0])
              balan=0.05*vb
              bala=bala+balan
              conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(bala,int(reffby)))
              conn.commit()
          bn=vb+inv
          bn=str(bn)
          conn = sqlite3.connect("wallet.db")  
          conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(bn,int(update.effective_user.id)))
          conn.commit()
          conn.close()
          conn = sqlite3.connect("wallet.db")  
          conn.execute("UPDATE COMPANY set link = '0', code='0' where ID = {}".format(int(update.effective_user.id)))
          conn.commit()
          conn.close()
          keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
          context.bot.send_message(chat_id=update.effective_user.id,text="${} is added to your wallet".format(vb),reply_markup=reply_markup)
        port = 465  # For SSL
        smtp_server = "smtp.gmail.com"
        sender_email = "astre.trading2@gmail.com"  # Enter your address
        receiver_email = "astre.trading@gmail.com"  # Enter receiver address
        password = "juggrzxgdmcojadj"

        eml="Wallet Notification{} deposited {}$".format(thu,vb)
        message = """\
        Subject: Trading View

        """+vb
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, message)
        return BUTTON
      elif a==2:
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")],[InlineKeyboardButton("Check Again", callback_data="927")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Your Transcation is detected please wait for confirmation and then click Check Again button",reply_markup=reply_markup)
        return BUTTON
      elif a==3:
          keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
          context.bot.send_message(chat_id=update.effective_user.id,text="Transcation Expired!",reply_markup=reply_markup)
          return BUTTON
      else:
        
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")],[InlineKeyboardButton("Check Again", callback_data="927")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="No Transcation is detected yet",reply_markup=reply_markup)
        return BUTTON

    elif a== '255':
  
        cc=update.callback_query.message.caption
        fg=cc
        cf=update.callback_query.message.message_id
        cc=cc.split("Product ID:")
        cc=cc[1]
        cc=cc.split("Price:")
        cc=cc[0]
        cc=cc.strip()
        pid=cc
        conn = sqlite3.connect('cart.db')
        cursor = conn.execute("DELETE  from COMPANY where productID='{}'".format(pid))
        conn.commit()
        keyboard =[[InlineKeyboardButton("➕ Add", callback_data="16"),InlineKeyboardButton("🔙 Cancel", callback_data="200")],[InlineKeyboardButton("➕ Quantity", callback_data="160"),InlineKeyboardButton("➖ Quantity", callback_data="161")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=fg,reply_markup=reply_markup)
        return BUTTON

    elif a== '16':
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        uo=cc
        dv=update.callback_query.message.photo[-1].file_id
        fg=cc
        cf=update.callback_query.message.message_id
        df=cc.split("Price:")
        df=df[1]

        df=df.split("Descripton:")
        df=df[0]
        df=df.strip()
        dd=cc.split("Product name:  ")
        dd=dd[1]
        dd=dd.split("Product ID:")
        dd=dd[0]
        dd=dd.strip()
        mm=cc.split("Quantity:")
        mm=mm[1]
        mm=mm.strip()
        cc=cc.split("Product ID:")
        cc=cc[1]
        cc=cc.split("Price:")
        cc=cc[0]
        cc=cc.strip()
        qp=uo.split("Product Type:  ")
        qp=qp[1]
        qp=qp.split("Quantity:")
        qp=qp[0]
        qp=qp.strip()              
        print(mm)
        pric=df
        print(pric)
        nam=dd
        pid=cc
        if qp=="Physical":
          conn = sqlite3.connect('cart.db') 
          conn.execute("INSERT INTO COMPANY (ID,name,productID,price,quantity) \
            VALUES ('{}', '{}','{}', '{}', '{}')".format(str(update.effective_user.id),nam,pid,pric,mm))
          conn.commit()
          keyboard =[[InlineKeyboardButton("➖ Remove", callback_data="255"),InlineKeyboardButton("🛒 Cart", callback_data="22")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=fg,reply_markup=reply_markup)
          conn.close()
          return BUTTON
        else:
          conn = sqlite3.connect("oo.db")  
          conn.execute("UPDATE COMPANY set amount = '{}', pid='{}',pname='{}',type='{}' where ID = {}".format(pric,pid,nam,qp,int(update.effective_user.id)))
          conn.commit()
          
          gs="Product ID:  "+pid+"\nTotal Amount:  "+pric+"$"+"\nProduct Type:  "+qp+"\nQuantity:  "+mm+"\n\n"
          keyboard =[[InlineKeyboardButton("Checkout", callback_data="2801")],[InlineKeyboardButton("🔙 Back", callback_data="200")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
          context.bot.send_message(chat_id=update.effective_user.id,text=gs,reply_markup=reply_markup)
          return BUTTON
    elif a=='2801':
            query.answer()
            conn = sqlite3.connect('oo.db')
            cursor = conn.execute("SELECT amount from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            conn.commit()
            for row in cursor: 
                ghhf=row[0]
                summl=float(ghhf)
            connection = sqlite3.connect("wallet.db")  
            cursor = connection.cursor()  
            cursor.execute("SELECT balance FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
            for names in cursor:
              inv=float(names[0])
            if summl>inv:
              keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
              context.bot.send_message(chat_id=update.effective_user.id,text="You dont have enough balance for this purchase",reply_markup=reply_markup)
              return BUTTON
            else:
              keyboard =[[InlineKeyboardButton("Pay from Wallet", callback_data="2902")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
              context.bot.send_message(chat_id=update.effective_user.id,text="Select the Payment option\nTotal amount= "+str(summl)+"$",reply_markup=reply_markup)
              return BUTTON
    elif a=="2902":
            query.answer()
            conn = sqlite3.connect('oo.db')
            cursor = conn.execute("SELECT pid,amount,pname,type from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            conn.commit()
            for row in cursor:
                x=row[0]
                tu=float(row[1])
                print(tu)
                namee=row[2]
                ttype=row[3]
 
                conn = sqlite3.connect('wallet.db')
                cursor = conn.execute("SELECT balance from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
                conn.commit()
                for row in cursor:
                    inv=float(row[0])
                    if inv<tu:
                     context.bot.send_message(chat_id=update.effective_user.id,text="Your Balance is not enough")
                    else:
                        balll=inv-tu
                        conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(balll,int(update.effective_user.id)))
                        conn.commit()
                        conn = sqlite3.connect('shop.db')
                        cursor = conn.execute("SELECT file from COMPANY where productID= '{}'".format(str(x)))
                        conn.commit()
                        for row in cursor:
                            fala=row[0]
                            if row[0]=='0':
                                conn = sqlite3.connect('file.db')
                                cursor = conn.execute("SELECT file,fid from COMPANY where ID= '{}' ".format(str(x)))
                                conn.commit()  
                                jobs = cursor.fetchall()
                                jobs=len(jobs)
                                if jobs>=2:
                                    conn = sqlite3.connect('file.db')
                                    cursor = conn.execute("SELECT file,fid from COMPANY where ID= '{}' order by RANDOM() LIMIT 1".format(str(x)))
                                    conn.commit()
                                    for row in cursor:
                                        vv=row[0]
                                        
                                        v= random.randint (0,999999)
                                        dat=datetime.today().strftime('%Y-%m-%d')
                                        conn.execute("DELETE from COMPANY where fid = '{}'".format(row[1]))
                                        conn.commit()
                                        conn = sqlite3.connect('Qorder.db')
                                        conn.execute("INSERT INTO COMPANY (ID,price,oid,type,date,pnam,productID) \
                                                VALUES ('{}', '{}','{}', '{}', '{}', '{}', '{}')".format(str(update.effective_user.id),tu,v,ttype,dat,namee,x)) 
                                        conn.commit()
                                        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
                                        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                                        context.bot.send_document(chat_id=update.effective_user.id,document=vv,caption="Here is your delivery\n",reply_markup=reply_markup)
                                        return BUTTON                                
                                elif jobs==1: 
                                    conn = sqlite3.connect('file.db')
                                    cursor = conn.execute("SELECT file,fid from COMPANY where ID= '{}' order by RANDOM() LIMIT 1".format(str(x)))
                                    conn.commit() 
                                    for row in cursor:
                                        vv=row[0]
                                        context.bot.send_message(chat_id=1394902938,text="Update file for product where productID is:  "+x)
                                        
                                        v= random.randint (0,999999)
                                        dat=datetime.today().strftime('%Y-%m-%d')
                                        conn.execute("DELETE from COMPANY where fid = '{}'".format(row[1]))
                                        conn.commit()
                                        conn = sqlite3.connect('Qorder.db')
                                        conn.execute("INSERT INTO COMPANY (ID,price,oid,type,date,pnam,productID) \
                                                VALUES ('{}', '{}','{}', '{}', '{}', '{}', '{}')".format(str(update.effective_user.id),tu,v,ttype,dat,namee,x)) 
                                        conn.commit()
                                        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
                                        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                                        context.bot.send_document(chat_id=update.effective_user.id,document=vv,caption="Here is your delivery\n",reply_markup=reply_markup)
                                        return BUTTON
                                else:
                                  conn = sqlite3.connect('wallet.db')
                                  cursor = conn.execute("SELECT balance from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
                                  conn.commit()
                                  for row in cursor:
                                      inv=float(row[0])
                                      balll=inv+tu
                                      conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(balll,int(update.effective_user.id)))
                                      conn.commit() 
                                  keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
                                  reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                                  context.bot.send_message(chat_id=update.effective_user.id,text="Sorry product in not available right now",reply_markup=reply_markup)
                                  return BUTTON
                            else:
                                cursor = conn.execute("UPDATE COMPANY set file='0' where productID= '{}'".format(x))
                                conn.commit()
                                
                                v= random.randint (0,999999)
                                dat=datetime.today().strftime('%Y-%m-%d')
                                conn = sqlite3.connect('Qorder.db')
                                conn.execute("INSERT INTO COMPANY (ID,price,oid,type,date,pnam,productID) \
                                        VALUES ('{}', '{}','{}', '{}', '{}', '{}', '{}')".format(str(update.effective_user.id),tu,v,ttype,dat,namee,x)) 
                                conn.commit()
                                keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
                                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                                context.bot.send_document(chat_id=update.effective_user.id,document=fala,caption="Here is your delivery\n",reply_markup=reply_markup)
                                return BUTTON
            #   except:
            #           keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
            #           reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            #           context.bot.send_message(chat_id=update.effective_user.id,text="Sorry product in not available right now",reply_markup=reply_markup)
            #           return BUTTON
       
    elif a=='2':
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text='Send ProductID to delete',reply_markup=reply_markup)
      return DELETE

    elif a=='100':
      keyboard =[[InlineKeyboardButton("➕ Product", callback_data="1"),InlineKeyboardButton("❌ Product", callback_data="2"),
                  InlineKeyboardButton("⚙️ Modify Product", callback_data="32")],[InlineKeyboardButton("⚙️ Delete Category", callback_data="del"),
                  InlineKeyboardButton("⚙️ ADD Category", callback_data="90")],[InlineKeyboardButton("🛒 Orders", callback_data="99"),
                  InlineKeyboardButton("🛒 ADD File", callback_data="1file"),InlineKeyboardButton("Managers List", callback_data="lman")],
                  [InlineKeyboardButton("🙋‍♂️ Welome message", callback_data="912"),InlineKeyboardButton("❓ Set Help / Support", callback_data="913"),
                  InlineKeyboardButton("Edit FAQ", callback_data="faq")],[InlineKeyboardButton("🔊 Announcement", callback_data="916"),
                  InlineKeyboardButton("Referral Data", callback_data="refd")],[InlineKeyboardButton("Modify the Product's file", callback_data="3344"),
                  InlineKeyboardButton("Change Photo", callback_data="3678")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to admin Dashboard" ,reply_markup=reply_markup)
    elif a== '22':     
      query.answer()
      conn = sqlite3.connect('cart.db')
      cursor = conn.execute("SELECT name,Price,quantity from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
      conn.commit()
      global xru
      xru="Here is your Cart:\n\n"
      for row in cursor: 

 
        m=row[2]+" x "+row[0]+ "     "+row[1]+"\n"
      
        xru=xru+m
      cursor = conn.execute("SELECT SUM(Price),SUM(quantity) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
      global ap
      ap=""
      global nj
      nj=""
      for row in cursor:
          ap=str(row[0])
      keyboard =[[InlineKeyboardButton("✅ Checkout", callback_data="25"),InlineKeyboardButton("❌ Cart", callback_data="29")],[InlineKeyboardButton("🔙 Cancel", callback_data="200")]]   
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=xru+"\nTotal"+" "+ap+"$" ,reply_markup=reply_markup)    
        
      return BUTTON
    
    elif a== '25':
            query.answer()
            conn = sqlite3.connect('cart.db')
            cursor = conn.execute("SELECT COUNT(Price),SUM(Price) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
            conn.commit()
            global summ
            summ=" "
            global cou
            cou=" "
            for row in cursor: 
                summ=float(row[1])
                cou=row[0]
            connection = sqlite3.connect("wallet.db")  
            cursor = connection.cursor()  
            cursor.execute("SELECT balance FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
            for names in cursor:
              inv=float(names[0])
            if summ>inv:
              keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
              context.bot.send_message(chat_id=update.effective_user.id,text="You dont have enough balance for this purchase",reply_markup=reply_markup)
              return BUTTON

            else:
              keyboard =[[InlineKeyboardButton("Pay from Wallet", callback_data="995")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
              context.bot.send_message(chat_id=update.effective_user.id,text="Select the Payment option\nTotal products= "+str(cou)+"\nTotal amount= "+str(summ)+"$",reply_markup=reply_markup)
              return BUTTON
    elif a== '995':
              context.bot.send_message(chat_id=update.effective_user.id,text="kindly provide your Tradingview username")
              return TRADE

    elif a== '200':
      query.answer()
      keyf=[["📝️ Products","🛒 Cart"],["💳 Wallet","💳 Add Credit"],["🛍 My orders","❓ Help/Support"],["🙋 FAQ","🤝 Affiliate Program"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Main Menu",reply_markup=reply_markup)
      
    elif a== 'cash': 
      context.bot.send_message(chat_id=update.effective_user.id,text="Send amount you want to add in your wallet")
      return COU
      

    elif a== '29':

      query.answer()  
      c=update.callback_query.message.message_id
      conn = sqlite3.connect('cart.db')
      cursor = conn.execute("SELECT ID from COMPANY")
      conn.commit()
      for names in cursor:
              cursor = conn.execute("DELETE  from COMPANY where ID={}".format(int(update.effective_user.id)))
              conn.commit() 
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.edit_message_text(chat_id=update.effective_user.id,message_id=c,text="Here is your cart:\n\n",reply_markup=reply_markup)
 

      return BUTTON

    elif a== '99':
      query.answer()  
      key=[["Pending Orders","Digital Orders"],["Order History"],["Cancel"]]
      reply_markup = ReplyKeyboardMarkup(key,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Select the Order Type:",reply_markup=reply_markup)
      return HAM   
      
    elif a== '130': #Reject
        c=update.callback_query.message.message_id
        cc=update.callback_query.message.text
        d=cc
        df=cc.split("Order ID:")
        df=df[1]
        df=df.split("User ID:")
        df=df[0]
        df=df.strip()
        dd=cc.split("User ID:")
        dd=dd[1]
        dd=dd.split("Product Price:")
        dd=dd[0]
        dd=dd.strip()
        cv=cc.split("Product Price:")
        cv=cv[1]
        cv=cv.split("Buyer's Address: ")
        cv=cv[0]
        cv=cv.strip()
        cv=cv.split("$")
        cv=cv[0]
        cv=cv.strip()
        cv=float(cv)
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("SELECT balance from COMPANY where ID={}".format(int(dd)))
        conn.commit()
        for row in cursor:
            bal=float(row[0])
        baln=bal+cv
        baln=str(baln)
        conn = sqlite3.connect('wallet.db')
        cursor = conn.execute("UPDATE COMPANY set balance='{}' where ID={}".format(baln,int(dd)))
        conn.commit() 
        print(dd)   
        conn = sqlite3.connect('orders.db')
        cursor = conn.execute("UPDATE COMPANY set status='Rejected' where oid='{}'".format(df))
        conn.commit()
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=dd,text="Your order has been rejected by Seller")
       
        context.bot.edit_message_text(chat_id=update.effective_user.id,message_id=c,text="{}\n Status: Rejected".format(d),reply_markup=reply_markup)       
        conn = sqlite3.connect('usorders.db')
        cursor = conn.execute("UPDATE COMPANY set status='Rejected' where oid='{}'".format(df))
        conn.commit()  
    elif a== '161': #quantity--
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        dv=update.callback_query.message.photo[-1].file_id
        cf=update.callback_query.message.message_id
        d=cc
        df=cc.split("Quantity:")
        df=df[1]
        df=df.strip()
        xd=df
        df=float(df)
        if df>1:
          d=cc
          df=cc.split("Quantity:")
          df=df[1]
          df=df.strip()
          xdd=df
          dff=float(df)
          df=dff
          dff=dff-1
          mf=cc.split("Price:")
          mf=mf[1]
          mf=mf.split("Descripton:")
          mf=mf[0]
          mf=mf.strip()
          fgh=mf
          mf=float(mf)
          vc=float(df)
          op=mf/vc
          np=op*dff
          np=str(np)
          dy=str(dff)

          hjam="Quantity:  {}".format(xdd)
          hjamn="Quantity:  {}".format(dy)
          cc=cc.replace(hjam,hjamn)
          cc=cc.replace(fgh,np+"$")
          keyboard =[[InlineKeyboardButton("➕ Add", callback_data="16"),InlineKeyboardButton("🔙 Cancel", callback_data="200")],[InlineKeyboardButton("+ Quantity", callback_data="160"),InlineKeyboardButton("- Quantity", callback_data="161")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
          context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)     
    elif a== '160': #quantity++
        cc=update.callback_query.message.caption
        cc=cc.replace("$","")
        dv=update.callback_query.message.photo[-1].file_id
        cf=update.callback_query.message.message_id
        d=cc
        df=cc.split("Quantity:")
        df=df[1]
        df=df.strip()
        xdd=df
        dff=float(df)
        df=dff
        dff=dff+1
        mf=cc.split("Price:")
        mf=mf[1]
        mf=mf.split("Descripton:")
        mf=mf[0]
        mf=mf.strip()
        fgh=mf
        mf=float(mf)
        vc=float(df)
        op=mf/vc
        np=op*dff
        np=str(np)
        dy=str(dff)
        hjam="Quantity:  {}".format(xdd)
        hjamn="Quantity:  {}".format(dy)
        cc=cc.replace(hjam,hjamn)
        cc=cc.replace(fgh,np+"$")
        keyboard =[[InlineKeyboardButton("➕ Add", callback_data="16"),InlineKeyboardButton("🔙 Cancel", callback_data="200")],[InlineKeyboardButton("+ Quantity", callback_data="160"),InlineKeyboardButton("- Quantity", callback_data="161")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.edit_message_caption(chat_id=update.effective_user.id,message_id=cf,caption=cc,reply_markup=reply_markup)              
    elif a== '120': #Accept
        c=update.callback_query.message.message_id
        cc=update.callback_query.message.text
        d=cc
        df=cc.split("Order ID:")
        df=df[1]
        df=df.split("User ID:")
        df=df[0]
        df=df.strip()
        dd=cc.split("User ID:")
        dd=dd[1]
        dd=dd.split("Product Price:")
        dd=dd[0]
        dd=dd.strip()
        conn = sqlite3.connect('orders.db')
        cursor = conn.execute("UPDATE COMPANY set status='Accepted' where oid='{}'".format(df))
        conn.commit() 
        conn = sqlite3.connect('usorders.db')
        cursor = conn.execute("UPDATE COMPANY set status='Rejected' where oid='{}'".format(df))
        conn.commit()
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=dd,text="Your order has been accepted by the Seller")     
        context.bot.edit_message_text(chat_id=update.effective_user.id,message_id=c,text="{}\n Status: Acepted".format(d),reply_markup=reply_markup)  
    elif a== '44': #Buyer's orders
      query.answer()  
      conn = sqlite3.connect('orders.db')
      cursor = conn.execute("SELECT ID from COMPANY where ID ='{}'".format(update.effective_user.id))
      jobs = cursor.fetchall()
      n=len(jobs)  
      n=n-10
      print(n)
      cursor = conn.execute("SELECT name,quantity,price,date,status,productID from COMPANY LIMIT 10 OFFSET {}".format(n))
      conn.commit()
      a="Orders Page 1:-\n\n"
      for row in cursor:
        print(row)
        g="\nProduct Name:  "+row[0]+"\nQuantity: "+row[1]+"\nProduct Price:  "+row[2]+"\nProduct ID:  "+row[5]+"\nDate: "+row[3]+"\nStatus: "+row[4]+"\n\n"
        a=a+g
        print(a)
      if n>0:
        keyboard =[[InlineKeyboardButton("Next page", callback_data="uspage")],[InlineKeyboardButton("🔙 Main Menu", callback_data="200")]]
      else:
        keyboard =[[InlineKeyboardButton("🔙 Main Menu", callback_data="200")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=a,reply_markup=reply_markup)
      return BUTTON
    elif a=="989":
        connection = sqlite3.connect("wallet.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT code,link,amount FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
        for names in cursor:
          inv=names[0]
          inv1=names[1]
          inv2=names[2]
          headers = {
              'Content-Type': 'application/json',
              'Authorization': 'Basic N2JiOWQxNmYtMGQ0ZS00YzNlLWJmNjQtMDk2ODYyMGIzNzFlOjdlOWRvYjBhUUxBMDg4YngzbTlndU5kSzFEcXJyWA=='
          }
          code=str(inv)
          response = requests.get('https://www.vivapayments.com/api/orders/{}'.format(code), headers=headers)
          r=response.text
          r=r.split('"StateId":')
          r=r[1]
          r=r.replace("}","")
          r=int(r)
          print(r)
          if r==1:
            conn = sqlite3.connect("wallet.db")  
            conn.execute("UPDATE COMPANY set amount = '0',link = '0', code='0' where ID = {}".format(int(update.effective_user.id)))
            conn.commit()
            conn.close()
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")],[InlineKeyboardButton("Add credit", callback_data="919")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Transcation Expired!",reply_markup=reply_markup)
            return BUTTON
          elif r==2:
              print("CANCEL generated new link")
              conn = sqlite3.connect("wallet.db")  
              conn.execute("UPDATE COMPANY set amount = '0',link = '0', code='0' where ID = {}".format(int(update.effective_user.id)))
              conn.commit()
              conn.close()
              keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")],[InlineKeyboardButton("Add credit", callback_data="919")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
              context.bot.send_message(chat_id=update.effective_user.id,text="Transcation Canceled!",reply_markup=reply_markup)
              return BUTTON
          elif r==3:
              connection = sqlite3.connect("wallet.db")  
              cursor = connection.cursor()  
              cursor.execute("SELECT balance,amount FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
              for names in cursor:
                inv=float(names[0])
                inv1=float(names[1])
              nm=inv1+inv
              bn=str(inv1)
              conn = sqlite3.connect("wallet.db")  
              conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(nm,int(update.effective_user.id)))
              conn.commit()
              conn.close()
              conn = sqlite3.connect("wallet.db")  
              conn.execute("UPDATE COMPANY set amount = '0',link = '0', code='0' where ID = {}".format(int(update.effective_user.id)))
              conn.commit()
              conn.close()
              keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
              context.bot.send_message(chat_id=update.effective_user.id,text="${} is added to your wallet".format(bn),reply_markup=reply_markup)
              return BUTTON
          else:
              print("pending")
              keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")],[InlineKeyboardButton("Check Again", callback_data="989")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
              context.bot.send_message(chat_id=update.effective_user.id,text="Transaction is pending.",reply_markup=reply_markup)
              return BUTTON

    elif a=="del":
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Send category name that you want to delete.",reply_markup=reply_markup)
      return DEL
    elif a=="cp":
      keyf=[["By Date"],["One Time use"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Select the category",reply_markup=reply_markup)  
      return C1

def men(update,context):
    msg=update.message.text
    if msg=="📝️ Products":
      keyf=[]
      connection = sqlite3.connect("cata.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT cat FROM COMPANY")
      for name in cursor:
        name=name[0]
        name=str(name)
        cv=[]
        cv.append(name)
        keyf.append(cv)
      keyf.append(["🔙 Main Menu"])
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Select the category",reply_markup=reply_markup)  
      return PD
    elif msg=="🛒 Cart":
      conn = sqlite3.connect('cart.db')
      cursor = conn.execute("SELECT name,Price,quantity from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
      conn.commit()
      global xru
      xru="Here is your Cart:\n\n"
      for row in cursor:  
        m=row[2]+" x "+row[0]+ "     "+row[1]+"\n"    
        xru=xru+m
      cursor = conn.execute("SELECT SUM(Price),SUM(quantity) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
      global ap
      ap=""
      global nj
      nj=""
      for row in cursor:
          ap=str(row[0])
      keyboard =[[InlineKeyboardButton("✅ Checkout", callback_data="25"),InlineKeyboardButton("❌ Cart", callback_data="29")],[InlineKeyboardButton("🔙 Cancel", callback_data="200")]]   
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=xru+"\nTotal"+" "+ap+"$" ,reply_markup=reply_markup)            
      return BUTTON
    elif msg=="🛍 My orders":
      conn = sqlite3.connect('orders.db')
      cursor = conn.execute("SELECT ID from COMPANY where ID ='{}'".format(update.effective_user.id))
      jobs = cursor.fetchall()
      n=len(jobs)  
      n=n-10
      print(n)
      cursor = conn.execute("SELECT name,quantity,price,date,status,productID from COMPANY LIMIT 10 OFFSET {}".format(n))
      conn.commit()
      a="Orders Page 1:-\n\n"
      for row in cursor:
        print(row)
        g="\nProduct Name:  "+row[0]+"\nQuantity: "+row[1]+"\nProduct Price:  "+row[2]+"\nProduct ID:  "+row[5]+"\nDate: "+row[3]+"\nStatus: "+row[4]+"\n\n"
        a=a+g
        print(a)
      if n>0:
        keyboard =[[InlineKeyboardButton("Next page", callback_data="uspage")],[InlineKeyboardButton("🔙 Main Menu", callback_data="200")]]
      else:
        keyboard =[[InlineKeyboardButton("🔙 Main Menu", callback_data="200")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=a,reply_markup=reply_markup)
      return BUTTON
    elif msg=="❓ Help/Support":
      context.bot.send_message(chat_id=update.effective_user.id,text=hel[0])
      return BUTTON
    elif msg=="🙋 FAQ":
      context.bot.send_message(chat_id=update.effective_user.id,text=fa[0])
      return BUTTON
    elif msg=="💳 Add Credit":
        keyboard =[[InlineKeyboardButton("Stripe",callback_data="cash")],[InlineKeyboardButton("Crypto", callback_data="9956")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Select the method to add the credit",reply_markup=reply_markup)
        return BUTTON 
            
    elif msg=="💳 Wallet":
      connection = sqlite3.connect("wallet.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT balance,name FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
      for names in cursor:
        inv=names[0]
        gsr=names[1]
        inv=float(inv)
        inv=round(inv,5)
        inv=str(inv)
        context.bot.send_message(chat_id=update.effective_user.id,text="Wallet ID:  "+str(update.effective_user.id)+"\n\nName:  "+gsr+"\nBalance:  "+"$"+inv)
    elif msg=="🔙 Main Menu":
      keyf=[["📝️ Products","🛒 Cart"],["💳 Wallet","💳 Add Credit"],["🛍 My orders","❓ Help/Support"],["🙋 FAQ","🤝 Affiliate Program"]]
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Main Menu",reply_markup=reply_markup)
    elif msg == "🤝 Affiliate Program":
      connection = sqlite3.connect("wallet.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT tref FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
      for names in cursor:
        inv=names[0]
      msg = 'https://t.me/astre_ts_bot?start={}'.format(str(update.effective_user.id))
      context.bot.send_message(chat_id=update.effective_user.id, text="Total Referrals: {}\nReferral Link:  {}\n\nInvite Your Friends to get get commisions when they buy the product".format(inv,msg),disable_web_page_preview=True)
        
def trade(update,context):
    msg=update.message.text
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set tradview = '{}' where ID = {}".format(msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    context.bot.send_message(chat_id=update.effective_user.id,text="Send your email address")
    return ADDRESS


def crp(update,context):
    msg=update.message.text
    try:
      msg=int(msg)
      aa=msg
      print("check")
      conn = sqlite3.connect("wallet.db")    
      cursor =conn.execute("SELECT code,link FROM COMPANY where ID= {}".format(int(update.effective_user.id)) )
      conn.commit()
      print("check")
      for names in cursor:
        inv=names[0]
        inv1=names[1]
        print("check")
        if inv=="0":
            headers = {
                'Content-Type': 'application/json',
                'Authorization': 'Basic N2JiOWQxNmYtMGQ0ZS00YzNlLWJmNjQtMDk2ODYyMGIzNzFlOjdlOWRvYjBhUUxBMDg4YngzbTlndU5kSzFEcXJyWA=='
            }
            gg=aa*100
            amount=gg
            data = { "amount": amount, "customerTrns": "Telegram Bot Shop Payment", "requestLang": "en-GB","paymentTimeOut": 1800, "merchantTrns": "Telegram bot shop payment"}
            data=str(data)
            response = requests.post('https://www.vivapayments.com/api/orders', headers=headers, data=data)
            r= response.text
            r=r.split('{"OrderCode":')
            r=r[1]
            r=r.split(',"ErrorCode"')
            code=r[0]
            linka="https://www.vivapayments.com//web/newtransaction.aspx?ref={}".format(code)
            print(code)
            print(linka)
            conn = sqlite3.connect("wallet.db")  
            conn.execute("UPDATE COMPANY set amount = '{}',link = '{}', code='{}' where ID = {}".format(aa,linka,code,int(update.effective_user.id)))
            conn.commit()
            conn.close()
            keyboard =[[InlineKeyboardButton("Make a Payment",url=linka)],[InlineKeyboardButton("I have paid", callback_data="989")],[InlineKeyboardButton("🔙 Back", callback_data="200")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="Click on below button to make payment and after making payment wait for few minutes then come back and click on i have paid.",reply_markup=reply_markup)
            return BUTTON
        else:
            headers = {
                'Content-Type': 'application/json',
                'Authorization': 'Basic N2JiOWQxNmYtMGQ0ZS00YzNlLWJmNjQtMDk2ODYyMGIzNzFlOjdlOWRvYjBhUUxBMDg4YngzbTlndU5kSzFEcXJyWA=='
            }
            code=str(inv)
            response = requests.get('https://www.vivapayments.com/api/orders/{}'.format(code), headers=headers)
            r=response.text
            r=r.split('"StateId":')
            r=r[1]
            r=r.replace("}","")
            r=int(r)
            print(r)
            if r==1:
              conn = sqlite3.connect("wallet.db")  
              conn.execute("UPDATE COMPANY set amount = '0',link = '0', code='0' where ID = {}".format(int(update.effective_user.id)))
              conn.commit()
              conn.close()
              keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")],[InlineKeyboardButton("Add credit", callback_data="9956")]]
              reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
              context.bot.send_message(chat_id=update.effective_user.id,text="Transcation Expired!",reply_markup=reply_markup)
              return BUTTON
            elif r==2:
                print("CANCEL generated new link")
                conn = sqlite3.connect("wallet.db")  
                conn.execute("UPDATE COMPANY set amount = '0',link = '0', code='0' where ID = {}".format(int(update.effective_user.id)))
                conn.commit()
                conn.close()
                keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")],[InlineKeyboardButton("Add credit", callback_data="9956")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                context.bot.send_message(chat_id=update.effective_user.id,text="Transcation Canceled!",reply_markup=reply_markup)
                return BUTTON
            elif r==3:
                print("PAID ADD CREDIT TO WALLET")
                connection = sqlite3.connect("wallet.db")  
                cursor = connection.cursor()  
                cursor.execute("SELECT balance,amount,refby  FROMCOMPANY where id= {}".format(int(update.effective_user.id)) )
                for names in cursor:
                  inv=float(names[0])
                  inv1=float(names[1])
                  reffby=names[2]
                  cursor.execute("SELECT balance  FROM COMPANY where id= {}".format(int(reffby)) )
                  for names in cursor:
                    bala=float(names[0])
                    balan=0.05*inv1
                    bala=bala+balan
                    conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(bala,int(reffby)))
                    conn.commit()
                nm=inv1+inv
                bn=str(inv1) 
                conn = sqlite3.connect("wallet.db")  
                conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(nm,int(update.effective_user.id)))
                conn.commit()
                conn.close()
                conn = sqlite3.connect("wallet.db")  
                conn.execute("UPDATE COMPANY set amount = '0',link = '0', code='0' where ID = {}".format(int(update.effective_user.id)))
                conn.commit()
                conn.close()
                keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                context.bot.send_message(chat_id=update.effective_user.id,text="${} is added to your wallet".format(bn),reply_markup=reply_markup)
                return BUTTON
            else:
                keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")],[InlineKeyboardButton("Check Again", callback_data="989")]]
                reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
                context.bot.send_message(chat_id=update.effective_user.id,text="Transcation is pending",reply_markup=reply_markup)
                return BUTTON
         
    except:
        context.bot.send_message(chat_id=update.effective_user.id,text='Invalid amount! send in digits only')
        return CRP

def address(update,context):
    msg=str(update.message.text)
    bnm=update.message.from_user
    ms=bnm.first_name
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("SELECT COUNT(Price),SUM(Price) from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
    conn.commit()
    global summ
    summ=" "
    global cou
    cou=" "
    for row in cursor: 
      summ=float(row[1])
      cou=row[0]
    connection = sqlite3.connect("wallet.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT balance,tradview FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
    for names in cursor:
      inv=float(names[0])
      rdus=names[1]
    bn=inv-summ
    bn=str(bn)
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set balance = '{}',email='{}' where ID = {}".format(bn,msg,int(update.effective_user.id)))
    conn.commit()
    conn.close()
    if msg=="None": 
      msg1=str(update.message.location.longitude)
      msg2=str(update.message.location.latitude)
      msg="Longtude:  "+msg1+"  , "+"Latitude:  "+msg2
    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("SELECT ID,name, quantity,Price,productID from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
    conn.commit()
    for row in cursor:
        v= random.randint (0,999999)
        dat=date.today()
        naak=row[1]
        deepe=row[3]
        conn = sqlite3.connect('usorders.db')
        conn.execute("INSERT INTO COMPANY (ID,name,price,oid,address,date,quantity,pnam,productID,status) \
                VALUES ('{}', '{}','{}', '{}', '{}', '{}', '{}','{}','{}','{}')".format(str(update.effective_user.id),row[1],row[3],str(v),msg,dat,row[2],ms,row[4],"pending")) 
        conn.commit()
        conn = sqlite3.connect('orders.db')
        conn.execute("INSERT INTO COMPANY (ID,name,price,oid,address,date,quantity,pnam,productID,status) \
                VALUES ('{}', '{}','{}', '{}', '{}', '{}', '{}','{}','{}','{}')".format(str(update.effective_user.id),row[1],row[3],str(v),msg,dat,row[2],ms,row[4],"pending")) 
        conn.commit()

    conn = sqlite3.connect('cart.db')
    cursor = conn.execute("DELETE  from COMPANY where ID= '{}'".format(str(update.effective_user.id)))
    conn.commit()
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="You order has been placed you will recieve conformation message shortly",reply_markup=reply_markup)
    eml="Name: {}\nTrading View Username: {}\nEmail Address: {}\nPrice: {}".format(ms,rdus,msg,deepe)
    port = 465  # For SSL
    smtp_server = "smtp.gmail.com"
    sender_email = "astre.trading2@gmail.com"  # Enter your address
    receiver_email = "astre.trading@gmail.com"  # Enter receiver address
    password = "juggrzxgdmcojadj"
    message = """\
    Subject: Trading View

    """+eml

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, message)   
   
    return BUTTON
def modf(update,context):
    msg=update.message.text
    global mpd
    mpd=msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Send new file to replace",reply_markup=reply_markup)
    return MODFF
def adfile(update,context):
    msg=update.message.text
    global eopl
    eopl=msg
    connection = sqlite3.connect("shop.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT productID FROM COMPANY where productID= '{}'".format(msg))
    jobs = cursor.fetchall()
    if len(jobs) !=0:   
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Send new file to add",reply_markup=reply_markup)
      return ADFILEE
    else:
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Invalid ProductID",reply_markup=reply_markup)
      return ADFILE  
def adfilee(update,context):
    msg=update.message.document.file_id
    yu= random.randint (0,999999)
    conn = sqlite3.connect('file.db')
    conn.execute("INSERT INTO COMPANY (ID,file,fid) \
                      VALUES ('{}', '{}','{}')".format(eopl,msg,yu))
    conn.commit()
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="FIle added",reply_markup=reply_markup)
def modff(update,context):
    sd=update.message.document.file_id
    filef = context.bot.getFile(update.message.document.file_id)
    filef.download('{}.txt'.format(mpd))
    conn = sqlite3.connect('shop.db')
    cursor = conn.execute("UPDATE COMPANY set file='{}' where productID='{}'".format(sd,mpd))
    conn.commit()

    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Done",reply_markup=reply_markup)
    return BUTTON
def modify(update,context):
    global mopid
    msg=update.message.text
    mopid=msg
    conn = sqlite3.connect('shop.db')
    cursor = conn.execute("SELECT  productID from COMPANY")
    conn.commit()
    for names in cursor:
      inv=names[0]
      if inv in msg:
            context.bot.send_message(chat_id=update.effective_user.id,text="Send product name")
            return ABAD

      else:
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="ProductID is wrong",reply_markup=reply_markup)
            return BUTTON
def ham(update,context):
  msg=update.message.text
  if msg=="Cancel":
    keyboard =[[InlineKeyboardButton("➕ Product", callback_data="1"),InlineKeyboardButton("❌ Product", callback_data="2"),
                InlineKeyboardButton("⚙️ Modify Product", callback_data="32")],[InlineKeyboardButton("⚙️ Delete Category", callback_data="del"),
                InlineKeyboardButton("⚙️ ADD Category", callback_data="90")],[InlineKeyboardButton("🛒 Orders", callback_data="99"),
                InlineKeyboardButton("🛒 ADD File", callback_data="1file"),InlineKeyboardButton("Managers List", callback_data="lman")],
                [InlineKeyboardButton("🙋‍♂️ Welome message", callback_data="912"),InlineKeyboardButton("❓ Set Help / Support", callback_data="913"),
                InlineKeyboardButton("Edit FAQ", callback_data="faq")],[InlineKeyboardButton("🔊 Announcement", callback_data="916"),
                InlineKeyboardButton("Referral Data", callback_data="refd")],[InlineKeyboardButton("Modify the Product's file", callback_data="3344"),
                InlineKeyboardButton("Change Photo", callback_data="3678")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Welcome to admin Dashboard" ,reply_markup=reply_markup)
  elif msg=="Pending Orders":
    conn = sqlite3.connect('orders.db')
    cursor = conn.execute("SELECT ID,name,price,productID,address,date,pnam,quantity,oid from COMPANY where status ='pending'")
    conn.commit()
    for row in cursor:
      print(row)
      g="\nProduct Name:  "+row[1]+"\nQuantity: "+row[7]+"\nCustomer Name:  "+row[6]+"\nProduct ID:  "+row[3]+"\nOrder ID:  "+row[8]+"\nUser ID:  "+row[0]+"\nProduct Price:  "+row[2]+ "\nBuyer's Address:  "+row[4]+"\nDate: "+row[5]
      keyboard =[[InlineKeyboardButton("Accept", callback_data="120"),InlineKeyboardButton("Reject", callback_data="130")],[InlineKeyboardButton("🔙 Cancel", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Pending order:\n"+g,reply_markup=reply_markup)
    return BUTTON
  elif msg=="Order History":
      conn = sqlite3.connect('orders.db')
      cursor = conn.execute("SELECT ID from COMPANY")
      jobs = cursor.fetchall()
      n=len(jobs)
      n=n-10
      print(n)
      cursor = conn.execute("SELECT name,quantity,price,date,pnam,address from COMPANY LIMIT 10 OFFSET {}".format(n))
      conn.commit()
      a="Orders Page 1:-\n\n"
      for row in cursor:
        print(row)
        g="\nProduct Name:  "+row[0]+"\nQuantity: "+row[1]+"\nCustomer Name:  "+row[4]+"\nProduct Price:  "+row[2]+ "\nBuyer's Address:  "+row[5]+"\nDate: "+row[3]+"\n\n"
        a=a+g
      if n>0:
        keyboard =[[InlineKeyboardButton("Next page", callback_data="page")],[InlineKeyboardButton("🔙 Main Menu", callback_data="100")]]
      else:
        keyboard =[[InlineKeyboardButton("🔙 Main Menu", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text=a,reply_markup=reply_markup)
      return BUTTON
  elif msg=="Digital Orders":
    conn = sqlite3.connect('Qorder.db')
    cursor = conn.execute("SELECT ID,price,oid,type,date,pnam,productID from COMPANY")
    conn.commit()
    global ty
    ty=""
    for row in cursor:
      print(row)
      g="\nProduct Name:  "+row[5]+"\nProduct ID:  "+row[6]+"\nUser ID:  "+row[0]+"\nProduct Price:  "+row[1]+"$"+"\nDate: "+row[4]+"\n"
      ty=ty+g
    keyboard =[[InlineKeyboardButton("🔙 Cancel", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Orders\n\n"+ty ,reply_markup=reply_markup)
      
    return BUTTON
def delete(update,context):
    msg=update.message.text
    conn = sqlite3.connect('shop.db')
    cursor = conn.execute("DELETE  from COMPANY where productID='{}'".format(msg))
    conn.commit()
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Product Deleted",reply_markup=reply_markup)
    return BUTTON

def ch(update,context):
    a=update.message.text
    b=a.split(",")
    for names in b:
      conn = sqlite3.connect('cata.db')
      conn.execute("INSERT INTO COMPANY (cat) \
              VALUES ('{}')".format(names)) 
      conn.commit()
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Category added successfully",reply_markup=reply_markup)
    return BUTTON
    
def sh(update,context):
       a=update.message.text
       wc.pop(0)
       wc.append(a)
       keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
       reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
       context.bot.send_message(chat_id=update.effective_user.id,text="Welcome message updated",reply_markup=reply_markup)
       return BUTTON
def je(update,context):
        msg=update.message.text
        conn = sqlite3.connect('users.db')
        cursor = conn.execute("SELECT ID from COMPANY")
        conn.commit()
        for row in cursor:
            aa=row[0]
            try:
                context.bot.send_message(chat_id=aa,text=msg)
            except:
                pass
        keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
        reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Message Sent",reply_markup=reply_markup)
        return BUTTON
       
def wa(update,context):
      a=update.message.text
      myfile = "cata.txt"
      with open(myfile, "r+") as f:
        data = f.read()
        f.seek(0)
        f.write(re.sub(r"<string>ABC</string>(\s+)<string>(.*)</string>", r"<xyz>ABC</xyz>\1<xyz>\2</xyz>", a))
        f.truncate()
      keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
      reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Category Changed",reply_markup=reply_markup)
      return BUTTON
def trc(update,context):
  global msgay
  msgay=update.message.text
  keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
  reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
  context.bot.send_message(chat_id=update.effective_user.id,text='Send Amount to send in USD',reply_markup=reply_markup)
  return AMO
def amo(update,context):
  msg=update.message.text
  try:
    suma=float(msg)
    connection = sqlite3.connect("wallet.db")  
    cursor = connection.cursor()  
    cursor.execute("SELECT balance,refby FROM COMPANY where id= {}".format(int(msgay)) )
    for names in cursor:
      inv=float(names[0])
      reffby=names[1]
      cursor.execute("SELECT balance  FROM COMPANY where id= {}".format(int(reffby)) )
      for names in cursor:
        bala=float(names[0])
        balan=0.05*suma
        bala=bala+balan
        conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(bala,int(reffby)))
        conn.commit()
    bn=inv+suma
    bn=str(bn)
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(bn,int(msgay)))
    conn.commit()
    conn.close()
    context.bot.send_message(chat_id=int(msgay),text="${} is added to your wallet".format(str(msg)))
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text='Amount Added successfully',reply_markup=reply_markup)
    return BUTTON
  except:
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text='Transcation failed! make sure you send amount in digits or correct wallet ID',reply_markup=reply_markup)
    return BUTTON

def sho(update,context): 
       a=update.message.text
       hel.pop(0)
       hel.append(a)
       keyboard =[[InlineKeyboardButton("🔙 Backk", callback_data="100")]]
       reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
       context.bot.send_message(chat_id=update.effective_user.id,text="Help/Support updated",reply_markup=reply_markup)
       return BUTTON
def faq(update,context): 
       a=update.message.text
       fa.pop(0)
       fa.append(a)
       keyboard =[[InlineKeyboardButton("🔙 Backk", callback_data="100")]]
       reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
       context.bot.send_message(chat_id=update.effective_user.id,text="Help/Support updated",reply_markup=reply_markup)
       return BUTTON
def akey(update,context): 
       a=update.message.photo[-1].file_id
       print(a)
       phoy.pop(0)
       phoy.append(a)
       keyboard =[[InlineKeyboardButton("🔙 Backk", callback_data="100")]]
       reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
       context.bot.send_message(chat_id=update.effective_user.id,text="photo updated",reply_markup=reply_markup)
       return BUTTON
def cart(update,context):
    msg=update.message.text
    conn = sqlite3.connect('shop.db')
    cursor = conn.execute("SELECT productID from COMPANY")
    conn.commit()
    for names in cursor:
      inv=names[0]
      if inv in msg:
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("SELECT name, description, price,productID   from COMPANY where productID= '{}'".format(inv))
        for row in cursor:
            conn = sqlite3.connect('cart.db')
            conn.execute("INSERT INTO COMPANY (ID,name,productID,price) \
                    VALUES ({}, '{}','{}', '{}')".format(int(update.effective_user.id),row[3],row[2],row[1]))
            
            context.bot.send_message(chat_id=update.effective_user.id,text="Product added to cart 🛒")
        break
      else:
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text="ProduchID is wrong",reply_markup=reply_markup)
    return BUTTON

def pd(update,context):
    msg=update.message.text
    conn = sqlite3.connect('shop.db')
    cursor = conn.execute("SELECT category from COMPANY")
    conn.commit()
  
    for names in cursor:
      inv=names[0]
      
      if inv in msg:
        conn = sqlite3.connect('shop.db')
        cursor = conn.execute("SELECT name, description, price,productID,photoid,file,type    from COMPANY where category= '{}'".format(inv))
        for row in cursor:   
                               
          m="Product name:  "+row[0]+"\nProduct ID:  "+row[3]+"\nPrice:  "+row[2]+"$"+"\nDescripton:  "+row[1]+"\nProduct Type:  "+row[6] +"\nQuantity:  1.0"
          keyboard =[[InlineKeyboardButton("➕ Add", callback_data="16"),InlineKeyboardButton("🔙 Cancel", callback_data="200")],[InlineKeyboardButton("➕ Quantity", callback_data="160"),InlineKeyboardButton("➖ Quantity", callback_data="161")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
          context.bot.send_photo(chat_id=update.effective_user.id,photo=row[4],caption=m,reply_markup=reply_markup)
        break

    return BUTTON

def pdstock(update,context):
    msg=update.message.text
    conn = sqlite3.connect('coming.db')
    cursor = conn.execute("SELECT category from COMPANY")
    conn.commit()
    for names in cursor:
      inv=names[0]
      
      if inv in msg:
        conn = sqlite3.connect('coming.db')
        cursor = conn.execute("SELECT name, description, price,productID,photoid  from COMPANY where category= '{}'".format(inv))
        conn.commit()
        for row in cursor:   
                               
          m="product name:  "+row[0]+"\nProduct ID:  "+row[3]+"\nPrice:  "+row[2]+"$"+"\nDescripton:  "+row[1] +"\nQuantity:  1.0"
          keyboard =[[InlineKeyboardButton("➕ Add", callback_data="16"),InlineKeyboardButton("🔙 Cancel", callback_data="200")],[InlineKeyboardButton("➕ Quantity", callback_data="160"),InlineKeyboardButton("➖ Quantity", callback_data="161")]]
          reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
          context.bot.send_photo(chat_id=update.effective_user.id,photo=row[4],caption=m,reply_markup=reply_markup)
        break

    return BUTTON
def photo(update,context):

            msg=update.message.photo[-1].file_id
            global pha
            pha =msg

            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send decription of product',reply_markup=reply_markup)

            return DES


def ab(update,context):
            msg=update.message.text
            global nm
            nm =msg
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='send product logo.',reply_markup=reply_markup)
            return PHO

def stock(update,context):
            msg=update.message.text
            global qm
            qm =msg

            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='send product logo.',reply_markup=reply_markup)
            return PHOSTOCK
def photostock(update,context):

            msg=update.message.photo[-1].file_id
            global gj
            gj =msg

            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send decription of product',reply_markup=reply_markup)
            return DESSTOCK
def desstock(update,context):
            msg=update.message.text
            global vn
            vn =msg
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send price of product',reply_markup=reply_markup)
            return PRSTOCK

def prstock(update,context):
    msg=update.message.text
    global cm
    cm =msg
    try:
      pri=float(cm)
      keyf=[]
      file= open("cata.txt","r+")  
      xyy=file.read()
      v=xyy.split(",")
      for name in v:
        cv=[]
        cv.append(name)
        keyf.append(cv)
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Select the category",reply_markup=reply_markup)  
      return PIDSTOCK
    except:
        context.bot.send_message(chat_id=update.effective_user.id,text='Invalid price! send in digits only')
        return PRSTOCK
  
def pidstock(update,context):
            msg=update.message.text
            global hm
            hm =msg

            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send Special Product ID',reply_markup=reply_markup)

            return CATEGORYSTOCK
def categorystock(update,context):
  msg=update.message.text
  conn = sqlite3.connect('coming.db')
  cursor=conn.execute("SELECT productID from COMPANY")
  conn.commit()
  for row in cursor:
    am=row[0]
    print(am)
    if am in msg:
      context.bot.send_message(chat_id=update.effective_user.id,text="This ProductID is already included. Try with different ID")
      return CATEGORYSTOCK
  yu= random.randint (0,999999)
  conn = sqlite3.connect('coming.db')
  conn.execute("INSERT INTO COMPANY (name,description,price,productID,photoid,category) \
                    VALUES ('{}', '{}','{}', '{}','{}','{}')".format(qm,vn,cm,msg,gj,hm))
  conn.commit()

  keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
  reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
  context.bot.send_message(chat_id=update.effective_user.id,text="Product added" ,reply_markup=reply_markup)
  return BUTTON
def des(update,context):
    msg=update.message.text
    global ds
    ds =msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text='Send price of product',reply_markup=reply_markup)
    return PR 
def pr(update,context):
    msg=update.message.text
    global pri
    pri =msg
    try:
      pri=float(pri)
      keyf=[]
      connection = sqlite3.connect("cata.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT cat FROM COMPANY")
      for name in cursor:
        name=name[0]
        name=str(name)
        cv=[]
        cv.append(name)
        keyf.append(cv)
      reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
      context.bot.send_message(chat_id=update.effective_user.id,text="Select the category",reply_markup=reply_markup)  
      return PID  
    except:
      context.bot.send_message(chat_id=update.effective_user.id,text='Invalid price! send in digits only')
      return PR 

def pid(update,context):
    msg=update.message.text             
    global ctr
    ctr =msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text='Send Special Product ID',reply_markup=reply_markup)
    return TYPP
def typp(update,context):
  msg=update.message.text
  global typr
  typr=msg
  keyf=[["Physical"],["Digital"]]
  reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
  context.bot.send_message(chat_id=update.effective_user.id,text="Select the Product Type",reply_markup=reply_markup)  
  return CATEGORY

def aman(update,context):
    msg=update.message.text
    man.append(msg)
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Manager added Successfully",reply_markup=reply_markup)
    return BUTTON
def rman(update,context):
  try:
    msg=update.message.text
    man.remove(msg)
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Manager removed Successfully",reply_markup=reply_markup)
    return BUTTON
  except:
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Manager not found send correct ID",reply_markup=reply_markup)
    return RMAN


def category(update,context):
  msg=update.message.text
  global bkl
  bkl=msg
  if msg=="Physical":
    yu= random.randint (0,999999)
    conn = sqlite3.connect('shop.db')
    conn.execute("INSERT INTO COMPANY (name,description,price,productID,photoid,category,file,type) \
                      VALUES ('{}', '{}','{}', '{}','{}','{}','{}','{}')".format(nm,ds,pri,msg,pha,ctr,"0",msg))
    conn.commit()

    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Product added" ,reply_markup=reply_markup)
    return BUTTON
  elif msg=="Digital":
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Send the Product File",reply_markup=reply_markup)
    return FCAT
def fcat(update,context):
    msg=update.message.document.file_id
    global fill
    fill="0"
    file1 = context.bot.getFile(update.message.document.file_id)
    file1.download('{}.txt'.format(typr))
    yu= random.randint (0,999999)
    conn = sqlite3.connect('shop.db')
    conn.execute("INSERT INTO COMPANY (name,description,price,productID,photoid,category,file,type) \
                      VALUES ('{}', '{}','{}', '{}','{}','{}','{}','{}')".format(nm,ds,pri,typr,pha,ctr,msg,bkl))
    conn.commit()

    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Product added" ,reply_markup=reply_markup)
    return BUTTON


def photoad(update,context):
    msg=update.message.photo[-1].file_id
    global pha
    pha =msg
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text='Send decription of product',reply_markup=reply_markup)
    return DESAD
def (update,context):
    chat_id = update.effective_user.id
    msg=update.message.text
    try:
      msg=int(msg)
      title = "Pay through Stripe"
      description = "Wallet"
      # select a payload just for you to recognize its the donation from your bot
      payload = "Custom-Payload"
      # In order to get a provider_token see https://core.telegram.org/bots/payments#getting-a-token
      provider_token = "350862534:LIVE:MmEzNGYxMjU3MmRi"
      currency = "USD"
      #price in dollars
      price = msg
      # price * 100 so as to include 2 decimal points
      prices = [LabeledPrice("Test", price * 100)]
      # optionally pass need_name=True, need_phone_number=True,
      # need_email=True, need_shipping_address=True, is_flexible=True
      context.bot.send_invoice(
        
        chat_id, title, description, payload, provider_token,  currency, prices, 
      )
      return QWER
    except:
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send in digits only.',reply_markup=reply_markup)
            return COU
def precheckout_callback(update, context):
    query = update.pre_checkout_query
    # check the payload, is this from your bot?
    if query.invoice_payload != 'Custom-Payload':
        # answer False pre_checkout_query
        query.answer(ok=False, error_message="Something went wrong...")
    else:
        query.answer(ok=True)
def qwer(update, context):
    # do something after successfully receiving payment?
    cx=update.message.successful_payment.total_amount
    print(cx)
    gh=cx/100
    kl=str(gh)
    vb=gh
    conn = sqlite3.connect("wallet.db") 
    cursor=conn.execute("SELECT balance,refby,name FROM COMPANY where id= {}".format(int(update.effective_user.id)) )       
    for names in cursor:
      inv=float(names[0])
      reffby=names[1]
      thu=names[2]
      print(reffby)
      cursor.execute("SELECT balance  FROM COMPANY where id= {}".format(int(reffby)) )
      for names in cursor:
        bala=float(names[0])
        balan=0.05*vb
        bala=bala+balan
        print(bala)
        conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(bala,int(reffby)))
        conn.commit()
    bn=vb+inv
    bn=str(bn)
    print(bn)
    conn = sqlite3.connect("wallet.db")  
    conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(bn,int(update.effective_user.id)))
    conn.commit()
    conn.close()    
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="200")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text='{} added in your Wallet'.format(gh),reply_markup=reply_markup)
    port = 465  
    smtp_server = "smtp.gmail.com"
    sender_email = "astre.trading2@gmail.com"  # Enter your address
    receiver_email = "astre.trading@gmail.com"  # Enter receiver address
    password = "juggrzxgdmcojadj"

    eml="Wallet Notification:\n{} deposited {}$".format(thu,kl)
    message = """\
    Subject: Trading View

    """+eml
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, message)
    return BUTTON
def abad(update,context):
            msg=update.message.text
            global nm
            nm =msg
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='send product logo.',reply_markup=reply_markup)
            return PHOAD
def pidad(update,context):
            msg=update.message.text
            global ctr
            ctr =msg
            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send Special Product ID',reply_markup=reply_markup)

            return CATEGORYAD

def desad(update,context):
            msg=update.message.text
            global ds
            ds =msg

            keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
            reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
            context.bot.send_message(chat_id=update.effective_user.id,text='Send price of product',reply_markup=reply_markup)
            return PRAD

def prad(update,context):
        msg=update.message.text
        global pri
        pri =msg
        keyf=[]
        file= open("cata.txt","r+")  
        xyy=file.read()
        v=xyy.split(",")
        for name in v:
          cv=[]
          cv.append(name)
          keyf.append(cv)
        reply_markup = ReplyKeyboardMarkup(keyf,resize_keyboard=True)
        context.bot.send_message(chat_id=update.effective_user.id,text="Select the category",reply_markup=reply_markup)  
  
        return PIDAD
def cod(update,context):
  msg=update.message.text
  connection = sqlite3.connect("coupon.db")  
  cursor = connection.cursor()  
  cursor.execute("SELECT date,amount FROM COMPANY where code= '{}'".format(msg)) 
  jobs = cursor.fetchall()
  if len(jobs) ==0:
    context.bot.send_message(chat_id=update.effective_user.id,text="Coupon not found")
  else:
    for names in jobs:
      dt=names[0]
      am=names[1]
    if dt=="0":
      connection = sqlite3.connect("wallet.db")  
      cursor = connection.cursor()  
      cursor.execute("SELECT balance FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
      for names in cursor:
        inv=float(names[0])
      bn=float(am)+inv
      bn=str(bn)
      conn = sqlite3.connect("wallet.db")  
      conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(bn,int(update.effective_user.id)))
      conn.commit()
      conn.close()
      conn = sqlite3.connect("coupon.db")  
      conn.execute("DELETE from COMPANY where code = '{}'".format(msg))
      conn.commit()
      conn.close()
      context.bot.send_message(chat_id=update.effective_user.id,text="${} ADDED TO YOUR WALLET".format(am))
    else:
      x=datetime.datetime.now()
      y=datetime.datetime.strptime(dt, '%Y-%m-%d %H:%M:%S')
      if x>y:
        conn = sqlite3.connect("coupon.db")  
        conn.execute("DELETE from COMPANY where code = '{}'".format(msg))
        conn.commit()
        conn.close()
        context.bot.send_message(chat_id=update.effective_user.id,text="Coupon Expired!")
      else:
        connection = sqlite3.connect("wallet.db")  
        cursor = connection.cursor()  
        cursor.execute("SELECT balance FROM COMPANY where id= {}".format(int(update.effective_user.id)) )
        for names in cursor:
          inv=float(names[0])
        bn=float(am)+inv
        bn=str(bn)
        conn = sqlite3.connect("wallet.db")  
        conn.execute("UPDATE COMPANY set balance = '{}' where ID = {}".format(bn,int(update.effective_user.id)))
        conn.commit()
        conn.close()
        context.bot.send_message(chat_id=update.effective_user.id,text="${} ADDED TO YOUR WALLET".format(am))
def deh(update,context):
  msg=update.message.text
  try:
    conn = sqlite3.connect("cata.db")  
    conn.execute("DELETE from COMPANY where cat = '{}'".format(msg))
    conn.commit()
    conn.close()
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Category Deleted Sccessfully.",reply_markup=reply_markup)
    return BUTTON
  except:
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Category not found",reply_markup=reply_markup)
    return DEL
def c1(update,context):
  msg=update.message.text
  if msg=="One Time use":
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Send credit Amount for that coupon.",reply_markup=reply_markup)
    return C2
  elif msg=="By Date":
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Send Expiray Date for the Coupon format: YYYY-MM-DD",reply_markup=reply_markup)
    return C3
def c3(update,context):
  msg=update.message.text
  try:
    global bir
    bir = datetime.datetime.strptime(msg, '%Y-%m-%d')
    print(bir)
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Send credit Amount for that coupon.",reply_markup=reply_markup)
    return C4

  except:
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Invalid Amount!",reply_markup=reply_markup)
    return C2
def c4(update,context):
  msg=update.message.text
  try:
    mg=float(msg)
    ida=update.message.message_id
    res= ''.join(random.choice(string.ascii_letters) for i in range(8))
    conn = sqlite3.connect('coupon.db')
    conn.execute("INSERT INTO COMPANY (ID,code,date,amount) \
            VALUES ('{}', '{}','{}', '{}')".format(ida,res,bir,msg)) 
    conn.commit()
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Coupon Cod: {}\nAmount: {}\nExpiry Date: {}\n\nGenerated".format(res,msg,bir),reply_markup=reply_markup)
    return BUTTON
  except:
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Invalid Amount!",reply_markup=reply_markup)
    return C2
def c2(update,context):
  msg=update.message.text
  try:
    mg=float(msg)
    ida=update.message.message_id
    res= ''.join(random.choice(string.ascii_letters) for i in range(8))
    conn = sqlite3.connect('coupon.db')
    conn.execute("INSERT INTO COMPANY (ID,code,date,amount) \
            VALUES ('{}', '{}','{}', '{}')".format(ida,res,"0",msg)) 
    conn.commit()
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Coupon Cod: {}\nAmount: {}\nOne Time Used\n\nGenerated".format(res,msg),reply_markup=reply_markup)
    return BUTTON

  except:
    keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
    reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True,one_time_keyboard=True)
    context.bot.send_message(chat_id=update.effective_user.id,text="Invalid Amount!",reply_markup=reply_markup)
    return C2

def categoryad(update,context):
  msg=update.message.text
  conn = sqlite3.connect('shop.db')
  
  cursor = conn.execute("DELETE  from COMPANY where productID='{}'".format(mopid))
  conn.commit()
  conn = sqlite3.connect('shop.db')
  conn.execute("INSERT INTO COMPANY (name,description,price,productID,photoid,category) \
                    VALUES ( '{}','{}', '{}','{}','{}','{}')".format(nm,ds,pri,msg,pha,ctr))
  conn.commit()
  
  keyboard =[[InlineKeyboardButton("🔙 Back", callback_data="100")]]
  reply_markup = InlineKeyboardMarkup(keyboard,resize_keyboard=True)
  context.bot.send_message(chat_id=update.effective_user.id,text="Product modified" ,reply_markup=reply_markup)
  return BUTTON
def precheckout_callback(update, context):
    query = update.pre_checkout_query
    # check the payload, is this from your bot?
    if query.invoice_payload != 'Custom-Payload':
        # answer False pre_checkout_query
        query.answer(ok=False, error_message="Something went wrong...")
    else:
        query.answer(ok=True)
def cancel(update, context):
    user = update.message.from_user
    return ConversationHandler.END
def main():                        
  updater = Updater("5406893588:AAF7APYgDHTX1UHAsImdHgrgIXlvfur4-kc", use_context=True)
  dp = updater.dispatcher
  conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start),MessageHandler(Filters.regex('^(📝️ Products|🛒 Cart|💳 Wallet|💳 Add Credit|🛍 My orders|❓ Help/Support|🔙 Main Menu|🤝 Affiliate Program|🙋 FAQ)$'), men),CallbackQueryHandler(button)],

        states={
            
            BUTTON: [CallbackQueryHandler(button),MessageHandler(Filters.text, pd)],
            AB: [MessageHandler(Filters.text, ab),CallbackQueryHandler(button)],
            HAM: [MessageHandler(Filters.text, ham)],
            COD: [MessageHandler(Filters.text, cod)],
            DES: [MessageHandler(Filters.text, des),CallbackQueryHandler(button)],
            CATEGORY: [MessageHandler(Filters.text, category),CallbackQueryHandler(button)],
            PR: [MessageHandler(Filters.text, pr),CallbackQueryHandler(button)],
            PD: [MessageHandler(Filters.text, pd),CallbackQueryHandler(button)],
            AMO: [MessageHandler(Filters.text, amo),CallbackQueryHandler(button)],
            QWER  : [MessageHandler(Filters.successful_payment, qwer)],   
            TRC: [MessageHandler(Filters.text, trc),CallbackQueryHandler(button)],
            PHO: [MessageHandler(Filters.photo, photo),CallbackQueryHandler(button)],
            PHOAD: [MessageHandler(Filters.photo, photoad),CallbackQueryHandler(button)],
            COU: [MessageHandler(Filters.text, cou)],
            PID: [MessageHandler(Filters.text, pid),CallbackQueryHandler(button)],
            AMAN: [MessageHandler(Filters.text, aman),CallbackQueryHandler(button)],
            RMAN: [MessageHandler(Filters.text, rman),CallbackQueryHandler(button)],
            PIDAD: [MessageHandler(Filters.text, pidad),CallbackQueryHandler(button)],
            TRADE: [MessageHandler(Filters.text, trade),CallbackQueryHandler(button)],
            CART: [MessageHandler(Filters.text, cart),CallbackQueryHandler(button)],
            DELETE: [MessageHandler(Filters.text, delete),CallbackQueryHandler(button)],
            MODIFY: [MessageHandler(Filters.text, modify),CallbackQueryHandler(button)],
            CATEGORYAD: [MessageHandler(Filters.text, categoryad),CallbackQueryHandler(button)],
            PRAD: [MessageHandler(Filters.text, prad),CallbackQueryHandler(button)],
            DESAD: [MessageHandler(Filters.text, desad),CallbackQueryHandler(button)],
            ABAD: [MessageHandler(Filters.text, abad),CallbackQueryHandler(button)],
            CH: [MessageHandler(Filters.text, ch),CallbackQueryHandler(button)],    
            SH: [MessageHandler(Filters.text, sh),CallbackQueryHandler(button)],
            C1: [MessageHandler(Filters.text, c1),CallbackQueryHandler(button)],
            C2: [MessageHandler(Filters.text, c2),CallbackQueryHandler(button)],
            C3: [MessageHandler(Filters.text, c3),CallbackQueryHandler(button)],
            C4: [MessageHandler(Filters.text, c4),CallbackQueryHandler(button)],
            SHO: [MessageHandler(Filters.text, sho),CallbackQueryHandler(button)],
            FAQ: [MessageHandler(Filters.text, faq),CallbackQueryHandler(button)],
            JE: [MessageHandler(Filters.text, je),CallbackQueryHandler(button)],
            WA: [MessageHandler(Filters.text, wa),CallbackQueryHandler(button)],
            DEL: [MessageHandler(Filters.text, deh),CallbackQueryHandler(button)],
            STOCK: [MessageHandler(Filters.text, stock),CallbackQueryHandler(button)],
            CRP: [MessageHandler(Filters.text, crp),CallbackQueryHandler(button)],
            TYPP: [MessageHandler(Filters.text, typp),CallbackQueryHandler(button)],
            MODFF: [MessageHandler(Filters.document, modff),CallbackQueryHandler(button)],
            ADFILE: [MessageHandler(Filters.text, adfile),CallbackQueryHandler(button)],
            ADFILEE: [MessageHandler(Filters.document, adfilee),CallbackQueryHandler(button)],
            MODF: [MessageHandler(Filters.text, modf),CallbackQueryHandler(button)],
            AKEY: [MessageHandler(Filters.photo, akey),CallbackQueryHandler(button)],
            FCAT: [MessageHandler(Filters.document, fcat),CallbackQueryHandler(button)],
            ADDRESS: [MessageHandler(Filters.text|Filters.location, address),CallbackQueryHandler(button)],
            PHOSTOCK: [MessageHandler(Filters.photo, photostock),CallbackQueryHandler(button)],
            CATEGORYSTOCK: [MessageHandler(Filters.text, categorystock),CallbackQueryHandler(button)],
            PRSTOCK: [MessageHandler(Filters.text, prstock),CallbackQueryHandler(button)],
            PIDSTOCK: [MessageHandler(Filters.text, pidstock),CallbackQueryHandler(button)],
            DESSTOCK: [MessageHandler(Filters.text, desstock),CallbackQueryHandler(button)],
            PDSTOCK: [MessageHandler(Filters.text, pdstock),CallbackQueryHandler(button)],

        },  
        fallbacks=[CommandHandler('cancel', cancel)],
        allow_reentry=True
    )
  dp.add_handler(conv_handler)
  dp.add_handler(PreCheckoutQueryHandler(precheckout_callback))
  updater.start_polling()
  updater.idle()
    
if __name__ == '__main__':
    main()